"""E3 — Docs and Agreements workbook.

Replaces the malformed Doc Tracker 2025.csv. Tabs: Agreements, Commitment
Letters, Deliverables, Templates, Lookups.
"""

from gsg_sheets import (
    BRAND,
    add_dropdown,
    add_kpi_row,
    add_lookups_tab,
    add_section_header,
    band_rows,
    conditional_formatting_date,
    conditional_formatting_status,
    freeze_header,
    new_workbook,
    save_workbook,
    set_column_widths,
    setup_dashboard_tab,
    tab_color,
    write_header,
)
from gsg_sheets.styles import apply_style, editable_style, formula_style
from gsg_sheets.taxonomies import AGREEMENT_STATUS, AGREEMENT_TYPE, INTERVENTION_TYPES
from openpyxl.styles import Alignment, Font


FILENAME = "E3 - Docs and Agreements.xlsx"


AGREEMENTS_HEADERS = [
    "agreement_id",
    "company_id",
    "company_name",
    "agreement_type",
    "signed_date",
    "signatory_name",
    "signatory_title",
    "gsg_signatory",
    "drive_url",
    "status",
    "related_intervention",
    "assignment_id",
    "notes",
    "updated_at",
    "updated_by",
]

COMMITMENT_HEADERS = [
    "letter_id",
    "company_id",
    "company_name",
    "related_to",
    "signatory_name",
    "signatory_title",
    "signed_date",
    "drive_url",
    "status",
    "notes",
]

DELIVERABLES_HEADERS = [
    "deliverable_id",
    "company_id",
    "company_name",
    "assignment_id",
    "intervention_type",
    "title",
    "due_date",
    "status",
    "drive_url",
    "owner_email",
    "notes",
]

TEMPLATES_HEADERS = [
    "template_id",
    "template_name",
    "purpose",
    "drive_url",
    "last_updated",
    "owner_email",
    "notes",
]

SEED_TEMPLATES = [
    ("TMP-001", "Master Joint Project Support Agreement (MJPSA)",
     "Primary bilateral agreement between GSG and selected companies", "", "", "", ""),
    ("TMP-002", "Commitment Letter — Conference Travel",
     "Company acceptance of conference travel support under Dutch or SIDA fund",
     "", "", "", ""),
    ("TMP-003", "C-Suite Coaching Engagement Letter",
     "Advisor-company engagement for C-Suite coaching track", "", "", "", ""),
    ("TMP-004", "Legal Support Referral Agreement",
     "Three-way agreement GSG, legal partner, company", "", "", "", ""),
    ("TMP-005", "Market Access Service Agreement",
     "Scope of services for MA interventions (registration, MKG agency, resource placement)",
     "", "", "", ""),
]


def _build_agreements(ws):
    tab_color(ws, BRAND["red"])
    write_header(ws, AGREEMENTS_HEADERS)
    freeze_header(ws)
    style_f = formula_style()
    style_e = editable_style()

    for row in range(2, 1001):
        cell = ws.cell(row=row, column=1, value=f'=IF(C{row}<>"", "AG-E3-"&TEXT(ROW()-1,"0000"), "")')
        apply_style(cell, style_f)

    add_dropdown(ws, "D", "=agreement_type")
    add_dropdown(ws, "J", "=agreement_status")
    add_dropdown(ws, "K", "=intervention_types")

    editable = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]
    for row in range(2, 1001):
        for col in editable:
            apply_style(ws.cell(row=row, column=col), style_e)

    # Phase K polish
    conditional_formatting_status(
        ws, "J",
        good=["Executed", "Countersigned", "Signed"],
        warn=["Sent"],
        bad=[],
        neutral=["Drafted"],
        last_row=1000,
    )
    conditional_formatting_date(ws, "E", last_row=1000)
    band_rows(ws, last_col_letter="O", last_row=1000)

    set_column_widths(
        ws,
        {1: 14, 2: 12, 3: 28, 4: 22, 5: 14, 6: 22, 7: 20, 8: 22, 9: 36,
         10: 16, 11: 22, 12: 14, 13: 40, 14: 18, 15: 20},
    )


def _build_commitments(ws):
    tab_color(ws, BRAND["orange"])
    write_header(ws, COMMITMENT_HEADERS)
    freeze_header(ws)
    style_f = formula_style()
    style_e = editable_style()

    for row in range(2, 501):
        cell = ws.cell(row=row, column=1, value=f'=IF(C{row}<>"", "CL-E3-"&TEXT(ROW()-1,"000"), "")')
        apply_style(cell, style_f)

    add_dropdown(ws, "I", "=agreement_status")

    editable = [2, 3, 4, 5, 6, 7, 8, 9, 10]
    for row in range(2, 501):
        for col in editable:
            apply_style(ws.cell(row=row, column=col), style_e)

    set_column_widths(
        ws,
        {1: 12, 2: 12, 3: 28, 4: 30, 5: 22, 6: 20, 7: 14, 8: 36, 9: 16, 10: 40},
    )


def _build_deliverables(ws):
    tab_color(ws, BRAND["teal"])
    write_header(ws, DELIVERABLES_HEADERS)
    freeze_header(ws)
    style_f = formula_style()
    style_e = editable_style()

    for row in range(2, 1001):
        cell = ws.cell(row=row, column=1, value=f'=IF(C{row}<>"", "DLV-E3-"&TEXT(ROW()-1,"0000"), "")')
        apply_style(cell, style_f)

    add_dropdown(ws, "E", "=intervention_types")
    add_dropdown(ws, "H", '"Planned,In Progress,Submitted,Accepted,Returned,Rejected"')

    editable = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11]
    for row in range(2, 1001):
        for col in editable:
            apply_style(ws.cell(row=row, column=col), style_e)

    set_column_widths(
        ws,
        {1: 14, 2: 12, 3: 28, 4: 14, 5: 22, 6: 30, 7: 14, 8: 14, 9: 36,
         10: 26, 11: 40},
    )


def _build_templates(ws):
    tab_color(ws, BRAND["navy"])
    write_header(ws, TEMPLATES_HEADERS)
    freeze_header(ws)
    style_e = editable_style()
    for offset, row_data in enumerate(SEED_TEMPLATES, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=offset, column=col_idx, value=value)
            apply_style(cell, style_e)
    set_column_widths(ws, {1: 12, 2: 38, 3: 44, 4: 36, 5: 14, 6: 26, 7: 40})


def _build_dashboard(ws):
    """Mirrors pages/docs/DocsPage.tsx in-sheet."""
    setup_dashboard_tab(ws, brand_tab_color=BRAND["red"])

    title = ws.cell(row=1, column=1, value="Docs and Agreements Dashboard")
    title.font = Font(name="Source Sans Pro", bold=True, size=18, color=BRAND["navy"])
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    sub = ws.cell(row=2, column=1, value="Live mirror of the Docs module in the Elevate Portal.")
    sub.font = Font(name="Source Sans Pro", size=10, color=BRAND["muted_text"])
    sub.alignment = Alignment(horizontal="left", indent=1)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)

    row = 4
    row = add_section_header(ws, row, "Top metrics")
    row = add_kpi_row(ws, row, [
        {"label": "Agreements", "formula": "=COUNTA(Agreements!C2:C1000)", "tone": "navy"},
        {"label": "Executed", "formula": "=COUNTIF(Agreements!J2:J1000,\"Executed\")", "tone": "green"},
        {"label": "In flight", "formula": "=COUNTIFS(Agreements!J2:J1000,\"Sent\")+COUNTIFS(Agreements!J2:J1000,\"Signed\")+COUNTIFS(Agreements!J2:J1000,\"Countersigned\")", "tone": "amber"},
        {"label": "Commitment letters", "formula": "=COUNTA('Commitment Letters'!C2:C500)", "tone": "teal"},
    ])

    # By status
    row = add_section_header(ws, row, "By status")
    for status in AGREEMENT_STATUS:
        ws.cell(row=row, column=1, value=status).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
        ws.cell(row=row, column=2, value=f'=COUNTIF(Agreements!J2:J1000,"{status}")').font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        bar = ws.cell(
            row=row, column=3,
            value=f'=IFERROR(REPT("█",MIN(40,ROUND(COUNTIF(Agreements!J2:J1000,"{status}")/MAX(1,COUNTA(Agreements!C2:C1000))*40,0))),"")',
        )
        bar.font = Font(name="Menlo", size=11, color=BRAND["red"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        row += 1
    row += 1

    # By type
    row = add_section_header(ws, row, "By type")
    for atype in AGREEMENT_TYPE:
        ws.cell(row=row, column=1, value=atype).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
        ws.cell(row=row, column=2, value=f'=COUNTIF(Agreements!D2:D1000,"{atype}")').font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        bar = ws.cell(
            row=row, column=3,
            value=f'=IFERROR(REPT("█",MIN(40,ROUND(COUNTIF(Agreements!D2:D1000,"{atype}")/MAX(1,COUNTA(Agreements!D2:D1000))*40,0))),"")',
        )
        bar.font = Font(name="Menlo", size=11, color=BRAND["teal"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        row += 1


def build() -> str:
    wb = new_workbook()
    add_lookups_tab(
        wb,
        {
            "agreement_type": AGREEMENT_TYPE,
            "agreement_status": AGREEMENT_STATUS,
            "intervention_types": INTERVENTION_TYPES,
        },
    )

    _build_dashboard(wb.create_sheet("Dashboard"))
    _build_agreements(wb.create_sheet("Agreements"))
    _build_commitments(wb.create_sheet("Commitment Letters"))
    _build_deliverables(wb.create_sheet("Deliverables"))
    _build_templates(wb.create_sheet("Templates"))

    order = ["Dashboard", "Agreements", "Commitment Letters", "Deliverables", "Templates", "Lookups"]
    for idx, name in enumerate(order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    return str(save_workbook(wb, FILENAME))


if __name__ == "__main__":
    print(build())
