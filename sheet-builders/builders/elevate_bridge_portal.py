"""E3 — Elevate Bridge Programme workbook.

Owns the full selection + upskilling programme that produced the 32 admitted
freelancers (Oct 2025 - Apr 2026). The new ElevateBridge HQ at
/elevatebridge in the portal reads this workbook. The existing
"E3 - Freelancers ElevateBridge" workbook (matching engine pool of 203)
stays as-is; this one is parallel.

Tabs:
  Applicants          - one row per of the 203, normalised
  Form Responses      - raw form mirror (read-only lineage)
  S1 Killing Factor   - S1 outcome + income breakdown
  S2 Tracks Sorting   - registered vs assigned track
  S3 SSI              - LinkedIn Social Selling Index
  S3 Response Scoring - long form, one row per (applicant, criterion)
  Interview Scoring   - long form, one row per (applicant, question)
  Final Decisions     - admitted / waitlisted / withdrew / dropped
  Scoring Rubrics     - rubric descriptors (drives the Scoring tab UI)
  Mentors             - 3 mentors for the three training tracks
  Training Sessions   - all sessions with recording / passcode / curriculum
  Session Attendance  - matrix of (session, applicant) -> attended
  Top Performers      - pre-computed final ranking
  ActivityLog         - portal-written audit trail
  Lookups             - named ranges for data validation
"""

from gsg_sheets import (
    BRAND,
    add_dropdown,
    add_kpi_row,
    add_lookups_tab,
    add_section_header,
    band_rows,
    conditional_formatting_status,
    freeze_header,
    new_workbook,
    save_workbook,
    set_column_widths,
    setup_dashboard_tab,
    tab_color,
    write_header,
)
from gsg_sheets.styles import apply_style, editable_style, formula_style
from openpyxl.styles import Alignment, Font


FILENAME = "E3 - Elevate Bridge Programme.xlsx"


# Schema. Every list here is the EXACT ordered header row written to its
# tab. The portal's src/types/elevateBridge.ts mirrors these field names.

APPLICANTS_HEADERS = [
    "applicant_id",
    "full_name_en",
    "full_name_ar",
    "email",
    "phone",
    "gender",
    "dob",
    "location",
    "region",
    "specialization",
    "education_level",
    "university_name",
    "university_major",
    "has_disability",
    "electricity_reliability",
    "internet_reliability",
    "track_registered",
    "track_assigned",
    "current_stage",
    "decision",
    "killing_factor_result",
    "killing_factor_reason",
    "ssi_score",
    "response_score_fl",
    "response_score_sm",
    "interview_score_fl",
    "interview_score_sm",
    "total_score",
    "hours_per_week",
    "session_commitment",
    "notes",
    "updated_at",
    "updated_by",
]

RESPONSES_HEADERS = [
    "applicant_id",
    "timestamp",
    "consent",
    "safeguarding_ack",
    "full_name_en",
    "full_name_ar",
    "email",
    "phone",
    "gender",
    "dob",
    "location",
    "specialization",
    "education_level",
    "university_major",
    "university_name",
    "disability_details",
    "how_heard",
    "track_registered",
    "session_commitment",
    "hours_per_week",
    "raw_blob",
]

STAGE1_HEADERS = [
    "applicant_id",
    "full_name",
    "email",
    "phone",
    "location",
    "track_registered",
    "killing_factor_result",
    "reason",
    "total_upwork_income_i",
    "upwork_income_jh",
    "upwork_income_agency",
    "sm_income_i",
    "sm_income_agency",
    "total_income_normalized",
    "notes",
]

STAGE2_HEADERS = [
    "applicant_id",
    "full_name",
    "email",
    "track_registered",
    "track_assigned",
    "skills_category",
    "jh_sector_upwork",
    "jh_sector_sm",
    "notes",
]

STAGE3_SSI_HEADERS = [
    "applicant_id",
    "full_name",
    "email",
    "linkedin_url",
    "ssi_score",
    "ssi_deadline_met",
    "notes",
]

STAGE3_RESPONSE_HEADERS = [
    "score_id",
    "applicant_id",
    "full_name",
    "track",
    "category",
    "criterion_key",
    "criterion_label",
    "weight",
    "sub_weight",
    "score",
    "notes",
    "scored_by",
    "scored_at",
]

INTERVIEW_HEADERS = [
    "interview_id",
    "applicant_id",
    "full_name",
    "track",
    "q_number",
    "category",
    "criterion_label",
    "weight",
    "sub_weight",
    "score",
    "attended",
    "notes",
    "scored_by",
    "scored_at",
]

DECISIONS_HEADERS = [
    "applicant_id",
    "full_name",
    "email",
    "track",
    "final_score",
    "decision",
    "decision_date",
    "decision_by",
    "notes",
]

RUBRICS_HEADERS = [
    "rubric_id",
    "track",
    "stage",
    "category",
    "criterion_key",
    "criterion_label",
    "weight",
    "sub_weight",
    "score_5",
    "score_4",
    "score_3",
    "score_2",
    "score_1",
    "notes",
]

MENTORS_HEADERS = [
    "mentor_id",
    "full_name",
    "email",
    "whatsapp",
    "track",
    "hourly_rate",
    "total_hours",
    "budget_total",
    "bio",
]

SESSIONS_HEADERS = [
    "session_id",
    "track",
    "session_num",
    "date",
    "topic",
    "recording_url",
    "passcode",
    "curriculum_url",
    "hours",
    "status",
    "notes",
    "updated_at",
    "updated_by",
]

ATTENDANCE_HEADERS = [
    "attendance_id",
    "session_id",
    "applicant_id",
    "full_name",
    "attended",
    "notes",
    "updated_at",
    "updated_by",
]

MATCHES_HEADERS = [
    "match_id",
    "applicant_id",
    "freelancer_name",
    "freelancer_email",
    "company_id",
    "company_name",
    "track",                # FL | SM | FL+SM
    "status",               # Proposed | Engaged | Producing | Completed | Cancelled
    "hours_per_week",
    "start_date",
    "end_date",
    "notes",
    "created_by",
    "created_at",
    "updated_at",
    "updated_by",
]

TOP_PERFORMERS_HEADERS = [
    "applicant_id",
    "overall_rank",
    "city_rank",
    "area",
    "track",
    "full_name_en",
    "full_name_ar",
    "email",
    "phone",
    "gender",
    "dob",
    "location",
    "specialization",
    "education_level",
    "total_earnings",
    "performance_score",
    "profile_url",
    "withdrew",
]

ACTIVITY_HEADERS = [
    "activity_id",
    "timestamp",
    "user_email",
    "entity_type",
    "entity_id",
    "action",
    "field",
    "old_value",
    "new_value",
    "details",
]


# Lookup lists for dropdowns.
TRACKS = ["FL", "SM", "FL+SM", ""]
STAGES = ["Applied", "S1 Filter", "S2 Sort", "S3 Scoring", "Interview", "Decision"]
DECISIONS = ["Admitted", "Waitlisted", "Withdrew", "Dropped", "Disqualified", ""]
REGIONS = ["West Bank", "Gaza Strip", "Outside Palestine"]
KILL_RESULT = ["Pass", "Fail"]
SSI_DEADLINE = ["Yes", "No", "Extended"]
SESSION_STATUSES = ["Scheduled", "Completed", "Cancelled", "Rescheduled"]
MATCH_STATUSES = ["Proposed", "Engaged", "Producing", "Completed", "Cancelled"]
ATTENDED = ["Yes", "No", "Late"]
YESNO = ["Yes", "No"]
RUBRIC_STAGES = ["Response", "Interview", "SSI"]
EB_ACTIONS = [
    "applicant_updated", "decision_changed", "score_edited", "interview_scored",
    "session_created", "session_updated", "attendance_marked", "mentor_updated",
    "rubric_updated", "top_performer_updated", "import_external", "export",
]
ENTITY_TYPES = ["applicant", "score", "interview", "session", "attendance", "decision", "mentor", "rubric"]


def _write_long_table(ws, headers, last_row=2000, last_col_letter=None, editable_cols=None, dropdowns=None):
    """Standardised long-table builder: header, freeze, band, dropdowns, edit styling."""
    tab_color(ws, BRAND["navy"])
    write_header(ws, headers)
    freeze_header(ws)
    style_e = editable_style()
    if editable_cols is None:
        editable_cols = list(range(1, len(headers) + 1))
    for row in range(2, last_row + 1):
        for col in editable_cols:
            apply_style(ws.cell(row=row, column=col), style_e)
    if dropdowns:
        for col_letter, source in dropdowns.items():
            add_dropdown(ws, col_letter, source, last_row=last_row)
    if last_col_letter:
        band_rows(ws, last_col_letter=last_col_letter, last_row=last_row)


def _build_applicants(ws):
    _write_long_table(
        ws,
        APPLICANTS_HEADERS,
        last_row=500,
        last_col_letter="AG",
        dropdowns={
            "I":  "=eb_regions",         # region
            "Q":  "=eb_tracks",          # track_registered
            "R":  "=eb_tracks",          # track_assigned
            "S":  "=eb_stages",          # current_stage
            "T":  "=eb_decisions",       # decision
            "U":  "=eb_kill_result",     # killing_factor_result
        },
    )
    conditional_formatting_status(
        ws, "T",
        good=["Admitted"],
        warn=["Waitlisted"],
        bad=["Withdrew", "Dropped", "Disqualified"],
        last_row=500,
    )
    set_column_widths(
        ws,
        {1: 14, 2: 26, 3: 24, 4: 28, 5: 16, 6: 8, 7: 12, 8: 18, 9: 14, 10: 18,
         11: 16, 12: 24, 13: 22, 14: 10, 15: 8, 16: 8, 17: 10, 18: 10, 19: 12,
         20: 12, 21: 14, 22: 22, 23: 10, 24: 12, 25: 12, 26: 12, 27: 12,
         28: 12, 29: 10, 30: 12, 31: 32, 32: 18, 33: 20},
    )


def _build_responses(ws):
    tab_color(ws, BRAND["muted_text"])
    write_header(ws, RESPONSES_HEADERS)
    freeze_header(ws)
    band_rows(ws, last_col_letter="U", last_row=500)
    set_column_widths(ws, {i: 18 for i in range(1, len(RESPONSES_HEADERS) + 1)})


def _build_stage1(ws):
    _write_long_table(
        ws,
        STAGE1_HEADERS,
        last_row=300,
        last_col_letter="O",
        dropdowns={"G": "=eb_kill_result"},
    )
    conditional_formatting_status(
        ws, "G",
        good=["Pass"], bad=["Fail"], last_row=300,
    )
    set_column_widths(ws, {1: 14, 2: 26, 3: 28, 4: 16, 5: 18, 6: 10, 7: 12,
                           8: 30, 9: 14, 10: 14, 11: 14, 12: 14, 13: 14, 14: 14, 15: 30})


def _build_stage2(ws):
    _write_long_table(
        ws,
        STAGE2_HEADERS,
        last_row=300,
        last_col_letter="I",
        dropdowns={"D": "=eb_tracks", "E": "=eb_tracks"},
    )
    set_column_widths(ws, {1: 14, 2: 26, 3: 28, 4: 14, 5: 14, 6: 18, 7: 22, 8: 22, 9: 30})


def _build_stage3_ssi(ws):
    _write_long_table(
        ws,
        STAGE3_SSI_HEADERS,
        last_row=200,
        last_col_letter="G",
        dropdowns={"F": "=eb_ssi_deadline"},
    )
    set_column_widths(ws, {1: 14, 2: 26, 3: 28, 4: 32, 5: 10, 6: 16, 7: 30})


def _build_stage3_response(ws):
    _write_long_table(
        ws,
        STAGE3_RESPONSE_HEADERS,
        last_row=5000,
        last_col_letter="M",
        dropdowns={"D": "=eb_tracks", "J": '"1,2,3,4,5"'},
    )
    set_column_widths(ws, {1: 22, 2: 14, 3: 26, 4: 8, 5: 20, 6: 22, 7: 28,
                           8: 8, 9: 10, 10: 8, 11: 30, 12: 24, 13: 18})


def _build_interview(ws):
    _write_long_table(
        ws,
        INTERVIEW_HEADERS,
        last_row=5000,
        last_col_letter="N",
        dropdowns={"D": "=eb_tracks", "J": '"1,2,3,4,5,Yes,No"', "K": "=eb_attended"},
    )
    set_column_widths(ws, {1: 22, 2: 14, 3: 26, 4: 8, 5: 8, 6: 22, 7: 28,
                           8: 8, 9: 10, 10: 10, 11: 12, 12: 30, 13: 24, 14: 18})


def _build_decisions(ws):
    _write_long_table(
        ws,
        DECISIONS_HEADERS,
        last_row=300,
        last_col_letter="I",
        dropdowns={"D": "=eb_tracks", "F": "=eb_decisions"},
    )
    conditional_formatting_status(
        ws, "F",
        good=["Admitted"], warn=["Waitlisted"],
        bad=["Withdrew", "Dropped", "Disqualified"], last_row=300,
    )
    set_column_widths(ws, {1: 14, 2: 26, 3: 28, 4: 10, 5: 12, 6: 14, 7: 14, 8: 22, 9: 30})


def _build_rubrics(ws):
    _write_long_table(
        ws,
        RUBRICS_HEADERS,
        last_row=300,
        last_col_letter="N",
        dropdowns={"B": "=eb_tracks", "C": "=eb_rubric_stages"},
    )
    set_column_widths(ws, {1: 20, 2: 8, 3: 12, 4: 22, 5: 22, 6: 30, 7: 8, 8: 10,
                           9: 28, 10: 28, 11: 28, 12: 28, 13: 28, 14: 24})


def _build_mentors(ws):
    _write_long_table(
        ws,
        MENTORS_HEADERS,
        last_row=20,
        last_col_letter="I",
    )
    set_column_widths(ws, {1: 14, 2: 28, 3: 28, 4: 18, 5: 36, 6: 10, 7: 12, 8: 12, 9: 40})


def _build_sessions(ws):
    _write_long_table(
        ws,
        SESSIONS_HEADERS,
        last_row=100,
        last_col_letter="M",
        dropdowns={"J": "=eb_session_statuses"},
    )
    conditional_formatting_status(
        ws, "J",
        good=["Completed"], warn=["Rescheduled"], bad=["Cancelled"],
        neutral=["Scheduled"], last_row=100,
    )
    set_column_widths(ws, {1: 22, 2: 36, 3: 8, 4: 12, 5: 32, 6: 42, 7: 18, 8: 42,
                           9: 8, 10: 14, 11: 28, 12: 18, 13: 20})


def _build_attendance(ws):
    _write_long_table(
        ws,
        ATTENDANCE_HEADERS,
        last_row=2000,
        last_col_letter="H",
        dropdowns={"E": "=eb_attended"},
    )
    conditional_formatting_status(
        ws, "E",
        good=["Yes"], warn=["Late"], bad=["No"], last_row=2000,
    )
    set_column_widths(ws, {1: 26, 2: 22, 3: 14, 4: 28, 5: 10, 6: 30, 7: 18, 8: 20})


def _build_matches(ws):
    _write_long_table(
        ws,
        MATCHES_HEADERS,
        last_row=500,
        last_col_letter="P",
        dropdowns={"G": "=eb_tracks", "H": "=eb_match_statuses"},
    )
    conditional_formatting_status(
        ws, "H",
        good=["Producing", "Engaged", "Completed"],
        warn=["Proposed"],
        bad=["Cancelled"],
        last_row=500,
    )
    set_column_widths(ws, {1: 22, 2: 14, 3: 26, 4: 28, 5: 14, 6: 28, 7: 8,
                           8: 12, 9: 8, 10: 12, 11: 12, 12: 30, 13: 24, 14: 18, 15: 18, 16: 20})


def _build_top_performers(ws):
    _write_long_table(
        ws,
        TOP_PERFORMERS_HEADERS,
        last_row=60,
        last_col_letter="R",
        dropdowns={"E": "=eb_tracks", "R": "=eb_yesno"},
    )
    set_column_widths(ws, {1: 14, 2: 6, 3: 6, 4: 14, 5: 8, 6: 28, 7: 24, 8: 28,
                           9: 16, 10: 8, 11: 12, 12: 18, 13: 24, 14: 16, 15: 14,
                           16: 12, 17: 36, 18: 10})


def _build_activity(ws):
    tab_color(ws, BRAND["navy"])
    write_header(ws, ACTIVITY_HEADERS)
    freeze_header(ws)
    add_dropdown(ws, "D", "=eb_entity_types", last_row=4000)
    add_dropdown(ws, "F", "=eb_actions", last_row=4000)
    band_rows(ws, last_col_letter="J", last_row=4000)
    set_column_widths(ws, {1: 32, 2: 22, 3: 26, 4: 14, 5: 22, 6: 22, 7: 18, 8: 24, 9: 24, 10: 40})


def _build_dashboard(ws):
    setup_dashboard_tab(ws, brand_tab_color=BRAND["red"])

    title = ws.cell(row=1, column=1, value="ElevateBridge — Programme HQ")
    title.font = Font(name="Source Sans Pro", bold=True, size=18, color=BRAND["navy"])
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    sub = ws.cell(
        row=2, column=1,
        value="Selection + upskilling pipeline that produced the 32 admitted freelancers. "
              "Oct 2025 - Apr 2026. Mirror of the portal's /elevatebridge module.",
    )
    sub.font = Font(name="Source Sans Pro", size=10, color=BRAND["muted_text"])
    sub.alignment = Alignment(horizontal="left", indent=1)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)

    row = 4
    row = add_section_header(ws, row, "Selection funnel")
    row = add_kpi_row(ws, row, [
        {"label": "Applicants",       "formula": "=COUNTA(Applicants!B2:B500)", "tone": "navy"},
        {"label": "Waitlisted",       "formula": '=COUNTIF(Applicants!U2:U500,"Fail")', "tone": "amber"},
        {"label": "Qualified Stage 2","formula": '=COUNTIF(Applicants!R2:R500,"FL")+COUNTIF(Applicants!R2:R500,"SM")+COUNTIF(Applicants!R2:R500,"FL+SM")', "tone": "teal"},
        {"label": "Admitted",         "formula": '=COUNTIF(Applicants!T2:T500,"Admitted")', "tone": "green"},
    ])

    row = add_section_header(ws, row, "Capacity building")
    row = add_kpi_row(ws, row, [
        {"label": "Sessions",   "formula": '=COUNTIF(\'Training Sessions\'!J2:J100,"Completed")&" / "&COUNTA(\'Training Sessions\'!A2:A100)', "tone": "navy"},
        {"label": "Hours done", "formula": '=SUMIF(\'Training Sessions\'!J2:J100,"Completed",\'Training Sessions\'!I2:I100)', "tone": "teal"},
        {"label": "Budget",     "formula": "=SUM(Mentors!H2:H20)", "tone": "navy", "value_fmt": "$#,##0"},
    ])

    row = add_section_header(ws, row, "By track (assigned)")
    for track in ["FL", "SM", "FL+SM"]:
        ws.cell(row=row, column=1, value=track).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
        ws.cell(row=row, column=2, value=f'=COUNTIF(Applicants!R2:R500,"{track}")').font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        bar = ws.cell(
            row=row, column=3,
            value=f'=IFERROR(REPT("█",MIN(40,ROUND(COUNTIF(Applicants!R2:R500,"{track}")/MAX(1,COUNTA(Applicants!B2:B500))*40,0))),"")',
        )
        bar.font = Font(name="Menlo", size=11, color=BRAND["red"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        row += 1


def build() -> str:
    wb = new_workbook()
    add_lookups_tab(
        wb,
        {
            "eb_tracks": TRACKS,
            "eb_stages": STAGES,
            "eb_decisions": DECISIONS,
            "eb_regions": REGIONS,
            "eb_kill_result": KILL_RESULT,
            "eb_ssi_deadline": SSI_DEADLINE,
            "eb_session_statuses": SESSION_STATUSES,
            "eb_match_statuses": MATCH_STATUSES,
            "eb_attended": ATTENDED,
            "eb_yesno": YESNO,
            "eb_rubric_stages": RUBRIC_STAGES,
            "eb_actions": EB_ACTIONS,
            "eb_entity_types": ENTITY_TYPES,
        },
    )

    _build_dashboard(wb.create_sheet("Dashboard"))
    _build_applicants(wb.create_sheet("Applicants"))
    _build_responses(wb.create_sheet("Form Responses"))
    _build_stage1(wb.create_sheet("S1 Killing Factor"))
    _build_stage2(wb.create_sheet("S2 Tracks Sorting"))
    _build_stage3_ssi(wb.create_sheet("S3 SSI"))
    _build_stage3_response(wb.create_sheet("S3 Response Scoring"))
    _build_interview(wb.create_sheet("Interview Scoring"))
    _build_decisions(wb.create_sheet("Final Decisions"))
    _build_rubrics(wb.create_sheet("Scoring Rubrics"))
    _build_mentors(wb.create_sheet("Mentors"))
    _build_sessions(wb.create_sheet("Training Sessions"))
    _build_attendance(wb.create_sheet("Session Attendance"))
    _build_top_performers(wb.create_sheet("Top Performers"))
    _build_matches(wb.create_sheet("Matches"))
    _build_activity(wb.create_sheet("ActivityLog"))

    order = [
        "Dashboard", "Applicants", "Form Responses", "S1 Killing Factor",
        "S2 Tracks Sorting", "S3 SSI", "S3 Response Scoring",
        "Interview Scoring", "Final Decisions", "Scoring Rubrics",
        "Mentors", "Training Sessions", "Session Attendance",
        "Top Performers", "Matches", "ActivityLog", "Lookups",
    ]
    for idx, name in enumerate(order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    return str(save_workbook(wb, FILENAME))


if __name__ == "__main__":
    print(build())
