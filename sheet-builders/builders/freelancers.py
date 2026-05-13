"""E3 — Freelancers (ElevateBridge) workbook.

ElevateBridge is a sub-intervention under Market Access. Consolidates the
35-tab legacy ElevateBridge Freelancers Application Responses workbook into
one canonical schema. Tabs: Freelancers, Track Assignments, Income Tracking,
Assessments, Lookups.
"""

from gsg_sheets import (
    BRAND,
    add_dropdown,
    add_kpi_row,
    add_lookups_tab,
    add_section_header,
    band_rows,
    conditional_formatting_status,
    freeze_header,
    new_workbook,
    save_workbook,
    set_column_widths,
    setup_dashboard_tab,
    tab_color,
    write_header,
)
from gsg_sheets.styles import apply_style, editable_style, formula_style
from gsg_sheets.taxonomies import (
    FREELANCER_ROLE,
    FREELANCER_STATUS,
    FREELANCER_TRACK,
)
from openpyxl.styles import Alignment, Font


FILENAME = "E3 - Freelancers ElevateBridge.xlsx"


FREELANCERS_HEADERS = [
    "freelancer_id",
    "full_name",
    "email",
    "phone",
    "location",
    "track",
    "role_profile",
    "assigned_mentor",       # the freelance industry mentor (external)
    "company_id",            # Cohort 3 company match (after graduation)
    "status",                # pipeline status — see FREELANCER_STATUS
    "start_date",
    "source_sheet",
    "notes",                 # general note from intake / mentor
    # Tracker columns appended on the same row, mirroring the Advisors
    # workbook so the kanban + drawer + audit log can drive the full
    # ElevateBridge funnel from the portal without a cross-tab join.
    "assignee_email",        # GSG team member triaging this freelancer
    "ack_sent",              # Yes/No — welcome email confirmed sent
    "assessment_date",
    "decision_date",
    "tracker_notes",         # private triage notes, separate from intake notes
    "updated_at",
    "updated_by",
]

FOLLOWUPS_HEADERS = [
    "followup_id",
    "freelancer_id",
    "due_date",
    "type",
    "assignee_email",
    "status",
    "notes",
    "created_by",
    "created_at",
    "completed_at",
    "updated_at",
    "updated_by",
]

ACTIVITY_HEADERS = [
    "activity_id",
    "timestamp",
    "user_email",
    "freelancer_id",
    "action",
    "field",
    "old_value",
    "new_value",
    "details",
]

COMMENTS_HEADERS = [
    "comment_id",
    "freelancer_id",
    "author_email",
    "body",
    "visibility",
    "created_at",
    "updated_at",
    "updated_by",
]

FOLLOWUP_TYPES = ["Email", "Call", "Meeting", "Other"]
FOLLOWUP_STATUSES = ["Open", "Done", "Snoozed"]
COMMENT_VISIBILITY = ["Team", "Admins"]
ACTIVITY_ACTIONS = ["status_change", "tracker_edit", "comment", "followup", "form_import", "assessment", "income"]

TRACK_ASSIGNMENTS_HEADERS = [
    "assignment_id",
    "freelancer_id",
    "freelancer_name",
    "track",
    "cohort_group",
    "mentor_email",
    "start_date",
    "end_date",
    "status",
    "notes",
]

INCOME_HEADERS = [
    "record_id",
    "freelancer_id",
    "freelancer_name",
    "month",
    "platform",
    "gross_income_usd",
    "verified_by",
    "verified_at",
    "notes",
]

ASSESSMENTS_HEADERS = [
    "assessment_id",
    "freelancer_id",
    "freelancer_name",
    "assessment_type",
    "assessor_email",
    "date",
    "score",
    "outcome",
    "notes",
]


def _build_freelancers(ws):
    tab_color(ws, BRAND["red"])
    write_header(ws, FREELANCERS_HEADERS)
    freeze_header(ws)
    style_f = formula_style()
    style_e = editable_style()

    for row in range(2, 2001):
        cell = ws.cell(row=row, column=1, value=f'=IF(B{row}<>"", "FL-E3-"&TEXT(ROW()-1,"0000"), "")')
        apply_style(cell, style_f)

    add_dropdown(ws, "F", "=freelancer_track")
    add_dropdown(ws, "G", "=freelancer_role")
    add_dropdown(ws, "J", "=freelancer_status")
    add_dropdown(ws, "P", '"Yes,No"')              # ack_sent (col 16 = P)
    # status (col J), ack_sent (col P), assessment_date (col Q), decision_date (col R)

    # Editable: everything except auto-id (1) and audit (last 2 cols).
    n_cols = len(FREELANCERS_HEADERS)
    editable = list(range(2, n_cols - 1))  # 2..n-2 inclusive
    for row in range(2, 2001):
        for col in editable:
            apply_style(ws.cell(row=row, column=col), style_e)

    # Phase K polish — colors map to the matching-engine workflow.
    conditional_formatting_status(
        ws, "J",
        good=["Producing", "Active"],
        warn=["Matched", "Released"],
        bad=["Dropped"],
        neutral=["Available", "On Hold", "Archived"],
        last_row=2000,
    )
    # Date columns: assessment / decision / start_date should highlight
    # when overdue.
    from gsg_sheets import conditional_formatting_date
    conditional_formatting_date(ws, "Q", last_row=2000)  # assessment_date
    conditional_formatting_date(ws, "R", last_row=2000)  # decision_date
    band_rows(ws, last_col_letter="T", last_row=2000)

    set_column_widths(
        ws,
        {
            1: 14,   # freelancer_id
            2: 26,   # full_name
            3: 28,   # email
            4: 16,   # phone
            5: 16,   # location
            6: 14,   # track
            7: 14,   # role_profile
            8: 22,   # assigned_mentor
            9: 12,   # company_id
            10: 14,  # status
            11: 14,  # start_date
            12: 26,  # source_sheet
            13: 40,  # notes
            14: 24,  # assignee_email
            15: 10,  # ack_sent
            16: 14,  # assessment_date
            17: 14,  # decision_date
            18: 32,  # tracker_notes
            19: 18,  # updated_at
            20: 20,  # updated_by
        },
    )


def _build_freelancer_followups(ws):
    tab_color(ws, BRAND["orange"])
    write_header(ws, FOLLOWUPS_HEADERS)
    freeze_header(ws)
    for row in range(2, 1001):
        ws.cell(row=row, column=1, value=f'=IF(B{row}<>"", "FLU-E3-"&TEXT(ROW()-1,"0000"), "")')
    add_dropdown(ws, "D", "=fl_followup_types", last_row=1000)
    add_dropdown(ws, "F", "=fl_followup_statuses", last_row=1000)
    conditional_formatting_status(
        ws, "F",
        good=["Done"],
        warn=["Snoozed"],
        neutral=["Open"],
        last_row=1000,
    )
    from gsg_sheets import conditional_formatting_date
    conditional_formatting_date(ws, "C", last_row=1000)
    band_rows(ws, last_col_letter="L", last_row=1000)
    set_column_widths(
        ws,
        {1: 14, 2: 14, 3: 14, 4: 12, 5: 24, 6: 12, 7: 36, 8: 24, 9: 18, 10: 18, 11: 18, 12: 20},
    )
    style_e = editable_style()
    for row in range(2, 1001):
        for col in [2, 3, 4, 5, 6, 7]:
            apply_style(ws.cell(row=row, column=col), style_e)


def _build_freelancer_activity(ws):
    tab_color(ws, BRAND["navy"])
    write_header(ws, ACTIVITY_HEADERS)
    freeze_header(ws)
    for row in range(2, 4001):
        ws.cell(row=row, column=1, value=f'=IF(B{row}<>"", "FACT-"&TEXT(ROW()-1,"00000"), "")')
    add_dropdown(ws, "E", "=fl_activity_actions", last_row=4000)
    band_rows(ws, last_col_letter="I", last_row=4000)
    set_column_widths(ws, {1: 14, 2: 20, 3: 26, 4: 14, 5: 18, 6: 18, 7: 24, 8: 24, 9: 36})


def _build_freelancer_comments(ws):
    tab_color(ws, BRAND["teal"])
    write_header(ws, COMMENTS_HEADERS)
    freeze_header(ws)
    for row in range(2, 2001):
        ws.cell(row=row, column=1, value=f'=IF(B{row}<>"", "FCMT-"&TEXT(ROW()-1,"00000"), "")')
    add_dropdown(ws, "E", "=fl_comment_visibility", last_row=2000)
    band_rows(ws, last_col_letter="H", last_row=2000)
    set_column_widths(ws, {1: 14, 2: 14, 3: 26, 4: 60, 5: 12, 6: 18, 7: 18, 8: 20})
    style_e = editable_style()
    for row in range(2, 2001):
        for col in [2, 3, 4, 5]:
            apply_style(ws.cell(row=row, column=col), style_e)


def _build_track_assignments(ws):
    tab_color(ws, BRAND["teal"])
    write_header(ws, TRACK_ASSIGNMENTS_HEADERS)
    freeze_header(ws)
    style_f = formula_style()
    style_e = editable_style()

    for row in range(2, 2001):
        cell = ws.cell(row=row, column=1, value=f'=IF(B{row}<>"", "FLA-E3-"&TEXT(ROW()-1,"0000"), "")')
        apply_style(cell, style_f)

    add_dropdown(ws, "D", "=freelancer_track")
    add_dropdown(ws, "I", "=freelancer_status")

    editable = [2, 3, 4, 5, 6, 7, 8, 9, 10]
    for row in range(2, 2001):
        for col in editable:
            apply_style(ws.cell(row=row, column=col), style_e)

    set_column_widths(
        ws,
        {1: 14, 2: 14, 3: 26, 4: 14, 5: 16, 6: 26, 7: 14, 8: 14, 9: 14, 10: 40},
    )


def _build_income(ws):
    tab_color(ws, BRAND["orange"])
    write_header(ws, INCOME_HEADERS)
    freeze_header(ws)
    style_f = formula_style()
    style_e = editable_style()

    for row in range(2, 5001):
        cell = ws.cell(row=row, column=1, value=f'=IF(B{row}<>"", "INC-E3-"&TEXT(ROW()-1,"00000"), "")')
        apply_style(cell, style_f)

    add_dropdown(ws, "E", '"Upwork,Freelancer,Fiverr,Toptal,LinkedIn,Other"')

    editable = [2, 3, 4, 5, 6, 7, 8, 9]
    for row in range(2, 5001):
        for col in editable:
            apply_style(ws.cell(row=row, column=col), style_e)

    set_column_widths(
        ws,
        {1: 16, 2: 14, 3: 26, 4: 10, 5: 16, 6: 14, 7: 22, 8: 16, 9: 40},
    )


def _build_assessments(ws):
    tab_color(ws, BRAND["navy"])
    write_header(ws, ASSESSMENTS_HEADERS)
    freeze_header(ws)
    style_f = formula_style()
    style_e = editable_style()

    for row in range(2, 2001):
        cell = ws.cell(row=row, column=1, value=f'=IF(B{row}<>"", "ASM-E3-"&TEXT(ROW()-1,"0000"), "")')
        apply_style(cell, style_f)

    add_dropdown(ws, "D", '"Technical,Business,Language,Mock Client,Final"')
    add_dropdown(ws, "H", '"Passed,Borderline,Failed,Pending"')

    editable = [2, 3, 4, 5, 6, 7, 8, 9]
    for row in range(2, 2001):
        for col in editable:
            apply_style(ws.cell(row=row, column=col), style_e)

    set_column_widths(
        ws,
        {1: 14, 2: 14, 3: 26, 4: 18, 5: 26, 6: 14, 7: 10, 8: 14, 9: 40},
    )


def _build_dashboard(ws):
    """Mirrors pages/freelancers/FreelancersPage.tsx in-sheet."""
    setup_dashboard_tab(ws, brand_tab_color=BRAND["red"])

    title = ws.cell(row=1, column=1, value="ElevateBridge Freelancers Dashboard")
    title.font = Font(name="Source Sans Pro", bold=True, size=18, color=BRAND["navy"])
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    sub = ws.cell(row=2, column=1, value="Live mirror of the ElevateBridge module in the Elevate Portal.")
    sub.font = Font(name="Source Sans Pro", size=10, color=BRAND["muted_text"])
    sub.alignment = Alignment(horizontal="left", indent=1)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)

    row = 4
    row = add_section_header(ws, row, "Top metrics")
    row = add_kpi_row(ws, row, [
        {"label": "Freelancers", "formula": "=COUNTA(Freelancers!B2:B2000)", "tone": "navy"},
        {"label": "In Program", "formula": "=COUNTIF(Freelancers!J2:J2000,\"In Program\")", "tone": "green"},
        {"label": "Applicants", "formula": "=COUNTIF(Freelancers!J2:J2000,\"Applicant\")", "tone": "amber"},
        {"label": "Income (USD)", "formula": "=SUM('Income Tracking'!F2:F5000)", "tone": "teal", "value_fmt": "$#,##0"},
    ])

    row = add_section_header(ws, row, "By status")
    for status in FREELANCER_STATUS:
        ws.cell(row=row, column=1, value=status).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
        ws.cell(row=row, column=2, value=f'=COUNTIF(Freelancers!J2:J2000,"{status}")').font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        bar = ws.cell(
            row=row, column=3,
            value=f'=IFERROR(REPT("█",MIN(40,ROUND(COUNTIF(Freelancers!J2:J2000,"{status}")/MAX(1,COUNTA(Freelancers!B2:B2000))*40,0))),"")',
        )
        bar.font = Font(name="Menlo", size=11, color=BRAND["red"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        row += 1
    row += 1

    row = add_section_header(ws, row, "By track")
    for track in FREELANCER_TRACK:
        ws.cell(row=row, column=1, value=track).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
        ws.cell(row=row, column=2, value=f'=COUNTIF(Freelancers!F2:F2000,"{track}")').font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        bar = ws.cell(
            row=row, column=3,
            value=f'=IFERROR(REPT("█",MIN(40,ROUND(COUNTIF(Freelancers!F2:F2000,"{track}")/MAX(1,COUNTA(Freelancers!F2:F2000))*40,0))),"")',
        )
        bar.font = Font(name="Menlo", size=11, color=BRAND["teal"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        row += 1


def build() -> str:
    wb = new_workbook()
    add_lookups_tab(
        wb,
        {
            "freelancer_track": FREELANCER_TRACK,
            "freelancer_role": FREELANCER_ROLE,
            "freelancer_status": FREELANCER_STATUS,
            "fl_followup_types": FOLLOWUP_TYPES,
            "fl_followup_statuses": FOLLOWUP_STATUSES,
            "fl_comment_visibility": COMMENT_VISIBILITY,
            "fl_activity_actions": ACTIVITY_ACTIONS,
        },
    )

    _build_dashboard(wb.create_sheet("Dashboard"))
    _build_freelancers(wb.create_sheet("Freelancers"))
    _build_freelancer_followups(wb.create_sheet("FollowUps"))
    _build_freelancer_activity(wb.create_sheet("ActivityLog"))
    _build_freelancer_comments(wb.create_sheet("Comments"))
    _build_track_assignments(wb.create_sheet("Track Assignments"))
    _build_income(wb.create_sheet("Income Tracking"))
    _build_assessments(wb.create_sheet("Assessments"))

    order = ["Dashboard", "Freelancers", "FollowUps", "ActivityLog", "Comments", "Track Assignments", "Income Tracking", "Assessments", "Lookups"]
    for idx, name in enumerate(order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    return str(save_workbook(wb, FILENAME))


if __name__ == "__main__":
    print(build())
