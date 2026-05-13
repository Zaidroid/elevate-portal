"""E3 — Conferences and Travel workbook.

Tabs: Conference Catalogue, Scoring Matrix, Company x Conference Tracker,
Commitment Letters, Summary Dashboard, Lookups.

Seeds with 15-conference catalogue placeholder rows and pre-loads Web Summit
Qatar 2026 commitments for the four SIDA-supported companies.
"""

from datetime import date

from gsg_sheets import (
    BRAND,
    add_dropdown,
    add_kpi_row,
    add_lookups_tab,
    add_section_header,
    band_rows,
    conditional_formatting_date,
    conditional_formatting_status,
    freeze_header,
    new_workbook,
    save_workbook,
    set_column_widths,
    setup_dashboard_tab,
    tab_color,
    write_header,
    write_rows,
)
from gsg_sheets.styles import apply_style, cell_style, editable_style, formula_style
from openpyxl.styles import Alignment, Font
from gsg_sheets.taxonomies import (
    CONFERENCE_DECISION,
    CONFERENCE_TIER,
    FUND_CODES,
)


FILENAME = "E3 - Conferences and Travel.xlsx"


CATALOGUE_HEADERS = [
    "conference_id",
    "name",
    "city",
    "country",
    "start_date",
    "end_date",
    "website",
    "tier",
    "fund_eligible",
    "estimated_cost_per_company_usd",
    "status",
    "notes",
]

SCORING_HEADERS = [
    "company_id",
    "company_name",
    "conference_id",
    "industry_fit",
    "market_fit",
    "readiness",
    "cost_benefit",
    "total_score",
    "rank_bucket",
    "notes",
]

TRACKER_HEADERS = [
    "tracking_id",
    "company_id",
    "company_name",
    "conference_id",
    "conference_name",
    "fit_score",
    "decision",
    "signatory_name",
    "signatory_title",
    "commitment_letter_url",
    "travel_dates",
    "flight_booked",
    "visa_status",
    "payment_id",
    "fund_code",
    "owner_email",
    "notes",
    "updated_at",
    "updated_by",
]

COMMITMENT_HEADERS = [
    "letter_id",
    "company_id",
    "company_name",
    "conference_id",
    "conference_name",
    "signatory_name",
    "signatory_title",
    "signed_date",
    "drive_url",
    "fund_code",
    "notes",
]


SEED_CATALOGUE = [
    # (name, city, country, start, end, website, tier, fund_eligible, cost, status, notes)
    ("Web Summit Qatar 2026", "Doha", "Qatar", date(2026, 2, 23), date(2026, 2, 26),
     "https://qatar.websummit.com", "T1", "SIDA", 8000, "Active", "GSG delegation confirmed"),
    ("We Make Future (WMF) 2026", "Bologna", "Italy", date(2026, 6, 24), date(2026, 6, 26),
     "https://www.wemakefuture.it", "T1", "Both", 6500, "Tracked", "Outreach pending"),
    ("GITEX Global 2026", "Dubai", "UAE", date(2026, 10, 12), date(2026, 10, 16),
     "https://www.gitex.com", "T1", "Both", 7500, "Tracked", ""),
    ("LEAP 2026", "Riyadh", "Saudi Arabia", date(2026, 2, 9), date(2026, 2, 12),
     "https://onegiantleap.com", "T1", "Both", 7000, "Tracked", ""),
    ("TechBBQ 2026", "Copenhagen", "Denmark", date(2026, 9, 16), date(2026, 9, 17),
     "https://techbbq.org", "T2", "Dutch", 5500, "Tracked", ""),
    ("Slush 2026", "Helsinki", "Finland", date(2026, 11, 18), date(2026, 11, 19),
     "https://www.slush.org", "T1", "Both", 7500, "Tracked", ""),
    ("VivaTech 2026", "Paris", "France", date(2026, 6, 11), date(2026, 6, 14),
     "https://vivatechnology.com", "T1", "Both", 8000, "Tracked", ""),
    ("Web Summit Lisbon 2026", "Lisbon", "Portugal", date(2026, 11, 9), date(2026, 11, 12),
     "https://websummit.com", "T1", "Both", 8500, "Tracked", ""),
    ("TNW Conference 2026", "Amsterdam", "Netherlands", date(2026, 6, 18), date(2026, 6, 19),
     "https://thenextweb.com/conference", "T2", "Dutch", 5500, "Tracked", ""),
    ("Step Conference 2026", "Dubai", "UAE", date(2026, 2, 25), date(2026, 2, 26),
     "https://stepconference.com", "T2", "Both", 4500, "Tracked", ""),
    ("Expand North Star 2026", "Dubai", "UAE", date(2026, 10, 13), date(2026, 10, 16),
     "https://expandnorthstar.com", "T2", "Both", 5000, "Tracked", ""),
    ("Seamless Middle East 2026", "Dubai", "UAE", date(2026, 5, 20), date(2026, 5, 22),
     "https://seamless-me.com", "T2", "Both", 4200, "Tracked", ""),
    ("Arab Health 2026", "Dubai", "UAE", date(2026, 1, 27), date(2026, 1, 30),
     "https://www.arabhealthonline.com", "T2", "Both", 5500, "Tracked", "HealthTech only"),
    ("Money 20/20 Europe 2026", "Amsterdam", "Netherlands", date(2026, 6, 2), date(2026, 6, 4),
     "https://www.money2020.com", "T1", "Dutch", 7200, "Tracked", "FinTech only"),
    ("Africa Tech Festival 2026", "Cape Town", "South Africa", date(2026, 11, 10), date(2026, 11, 13),
     "https://africatechfestival.com", "T2", "Both", 6800, "Tracked", ""),
]


WEB_SUMMIT_QATAR_COMPANIES = [
    # (company_name, signatory_name, signatory_title)
    ("Radix Technologies", "Adel Jodalah", "CEO"),
    ("Sada Intelligence", "Yousef Ashhab", "CEO"),
    ("Haweya Information Technology", "Mohammed Qudaih", "CEO"),
    ("Tech 360 for Digital Solutions (Farabio)", "Imad Temeiza", "CEO"),
]


def _build_catalogue(ws):
    tab_color(ws, BRAND["red"])
    write_header(ws, CATALOGUE_HEADERS)
    freeze_header(ws)

    style_f = formula_style()
    for row in range(2, 201):
        cell = ws.cell(row=row, column=1, value=f'=IF(B{row}<>"", "CONF-E3-"&TEXT(ROW()-1,"000"), "")')
        apply_style(cell, style_f)

    start_row = 2
    data_rows = [
        [
            "",  # id auto-formula handles it
            name, city, country, start, end, website, tier, fund, cost, status, notes,
        ]
        for name, city, country, start, end, website, tier, fund, cost, status, notes in SEED_CATALOGUE
    ]
    write_rows(ws, data_rows, start_row=start_row, editable_cols=list(range(2, 13)))

    add_dropdown(ws, "H", "=conference_tier")
    add_dropdown(ws, "I", '"Dutch,SIDA,Both"')
    add_dropdown(ws, "K", '"Tracked,Active,Past"')

    set_column_widths(
        ws,
        {1: 14, 2: 32, 3: 16, 4: 18, 5: 14, 6: 14, 7: 36, 8: 8, 9: 12,
         10: 14, 11: 10, 12: 40},
    )


def _build_scoring(ws):
    tab_color(ws, BRAND["teal"])
    write_header(ws, SCORING_HEADERS)
    freeze_header(ws)
    style_f = formula_style()
    style_e = editable_style()

    for row in range(2, 1001):
        # total_score = industry_fit + market_fit + readiness + cost_benefit
        cell = ws.cell(
            row=row,
            column=8,
            value=f'=IFERROR(SUM(D{row}:G{row}), "")',
        )
        apply_style(cell, style_f)
        # rank_bucket: High (>=16), Medium (>=11), Low otherwise. Scale 1-5 per field x4 = max 20.
        cell = ws.cell(
            row=row,
            column=9,
            value=(
                f'=IF(H{row}="","",'
                f'IF(H{row}>=16,"High Fit",'
                f'IF(H{row}>=11,"Medium Fit","Low Fit")))'
            ),
        )
        apply_style(cell, style_f)
        for col in [1, 2, 3, 4, 5, 6, 7, 10]:
            apply_style(ws.cell(row=row, column=col), style_e)

    add_dropdown(ws, "D", '"1,2,3,4,5"')
    add_dropdown(ws, "E", '"1,2,3,4,5"')
    add_dropdown(ws, "F", '"1,2,3,4,5"')
    add_dropdown(ws, "G", '"1,2,3,4,5"')

    set_column_widths(ws, {1: 12, 2: 28, 3: 14, 4: 12, 5: 12, 6: 12, 7: 14, 8: 12, 9: 14, 10: 40})


def _build_tracker(ws):
    tab_color(ws, BRAND["orange"])
    write_header(ws, TRACKER_HEADERS)
    freeze_header(ws)
    style_f = formula_style()
    style_e = editable_style()

    for row in range(2, 1001):
        cell = ws.cell(row=row, column=1, value=f'=IF(B{row}<>"", "TRK-"&TEXT(ROW()-1,"000"), "")')
        apply_style(cell, style_f)

    # Pre-load Web Summit Qatar 2026 rows (company_id left blank; joins by name until Company Master is wired).
    start_row = 2
    for offset, (name, signatory, title) in enumerate(WEB_SUMMIT_QATAR_COMPANIES):
        row = start_row + offset
        ws.cell(row=row, column=3).value = name
        ws.cell(row=row, column=5).value = "Web Summit Qatar 2026"
        ws.cell(row=row, column=7).value = "Committed"
        ws.cell(row=row, column=8).value = signatory
        ws.cell(row=row, column=9).value = title
        ws.cell(row=row, column=15).value = "91763"  # SIDA
        ws.cell(row=row, column=16).value = "zsalem33@gmail.com"
        ws.cell(row=row, column=17).value = "Commitment letter signed; payment docs sent to Khamis Eweis"

    editable = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17]
    for row in range(2, 1001):
        for col in editable:
            apply_style(ws.cell(row=row, column=col), style_e)

    add_dropdown(ws, "G", "=conference_decision")
    add_dropdown(ws, "L", '"Yes,No,Pending"')
    add_dropdown(ws, "M", '"Not required,Pending,Applied,Approved,Denied"')
    add_dropdown(ws, "O", "=fund_codes")

    # Phase K polish
    conditional_formatting_status(
        ws, "G",
        good=["Attended", "Committed"],
        warn=["Nominated"],
        bad=["Withdrawn"],
        last_row=1000,
    )
    conditional_formatting_status(
        ws, "M",
        good=["Approved", "Not required"],
        warn=["Applied", "Pending"],
        bad=["Denied"],
        last_row=1000,
    )
    conditional_formatting_date(ws, "K", last_row=1000)
    band_rows(ws, last_col_letter="S", last_row=1000)

    set_column_widths(
        ws,
        {
            1: 12, 2: 12, 3: 28, 4: 14, 5: 28, 6: 10, 7: 14, 8: 22,
            9: 18, 10: 36, 11: 20, 12: 14, 13: 18, 14: 16, 15: 12,
            16: 26, 17: 40, 18: 18, 19: 20,
        },
    )


def _build_commitments(ws):
    tab_color(ws, BRAND["navy"])
    write_header(ws, COMMITMENT_HEADERS)
    freeze_header(ws)
    style_f = formula_style()
    style_e = editable_style()

    for row in range(2, 501):
        cell = ws.cell(row=row, column=1, value=f'=IF(B{row}<>"", "CL-E3-"&TEXT(ROW()-1,"000"), "")')
        apply_style(cell, style_f)

    # Pre-load Web Summit Qatar commitments
    for offset, (name, signatory, title) in enumerate(WEB_SUMMIT_QATAR_COMPANIES):
        row = 2 + offset
        ws.cell(row=row, column=3).value = name
        ws.cell(row=row, column=5).value = "Web Summit Qatar 2026"
        ws.cell(row=row, column=6).value = signatory
        ws.cell(row=row, column=7).value = title
        ws.cell(row=row, column=10).value = "91763"

    add_dropdown(ws, "J", "=fund_codes")

    editable = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11]
    for row in range(2, 501):
        for col in editable:
            apply_style(ws.cell(row=row, column=col), style_e)

    set_column_widths(
        ws,
        {1: 12, 2: 12, 3: 28, 4: 14, 5: 28, 6: 22, 7: 18, 8: 14, 9: 36,
         10: 12, 11: 40},
    )


def _build_dashboard(ws):
    """Mirrors pages/conferences/ConferencesPage.tsx in-sheet."""
    setup_dashboard_tab(ws, brand_tab_color=BRAND["red"])

    title = ws.cell(row=1, column=1, value="Conferences Dashboard")
    title.font = Font(name="Source Sans Pro", bold=True, size=18, color=BRAND["navy"])
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    sub = ws.cell(row=2, column=1, value="Live mirror of the Conferences module in the Elevate Portal.")
    sub.font = Font(name="Source Sans Pro", size=10, color=BRAND["muted_text"])
    sub.alignment = Alignment(horizontal="left", indent=1)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)

    row = 4
    row = add_section_header(ws, row, "Top metrics")
    row = add_kpi_row(ws, row, [
        {"label": "Conferences tracked", "formula": "=COUNTA('Conference Catalogue'!B2:B200)", "tone": "navy"},
        {"label": "Committed", "formula": "=COUNTIF('Company x Conference Tracker'!G2:G1000,\"Committed\")", "tone": "amber"},
        {"label": "Attended", "formula": "=COUNTIF('Company x Conference Tracker'!G2:G1000,\"Attended\")", "tone": "green"},
        {"label": "Commitment letters", "formula": "=COUNTA('Commitment Letters'!C2:C500)", "tone": "teal"},
    ])

    # By tier
    row = add_section_header(ws, row, "By tier")
    for tier in ["T1", "T2", "T3"]:
        ws.cell(row=row, column=1, value=tier).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
        ws.cell(row=row, column=2, value=f'=COUNTIF(\'Conference Catalogue\'!H2:H200,"{tier}")').font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        bar = ws.cell(
            row=row, column=3,
            value=f'=IFERROR(REPT("█",MIN(40,ROUND(COUNTIF(\'Conference Catalogue\'!H2:H200,"{tier}")/MAX(1,COUNTA(\'Conference Catalogue\'!H2:H200))*40,0))),"")',
        )
        bar.font = Font(name="Menlo", size=11, color=BRAND["red"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        row += 1
    row += 1

    # By fund eligibility
    row = add_section_header(ws, row, "By fund eligibility")
    for fund in ["Dutch", "SIDA", "Both"]:
        ws.cell(row=row, column=1, value=fund).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
        ws.cell(row=row, column=2, value=f'=COUNTIF(\'Conference Catalogue\'!I2:I200,"{fund}")').font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        bar = ws.cell(
            row=row, column=3,
            value=f'=IFERROR(REPT("█",MIN(40,ROUND(COUNTIF(\'Conference Catalogue\'!I2:I200,"{fund}")/MAX(1,COUNTA(\'Conference Catalogue\'!I2:I200))*40,0))),"")',
        )
        bar.font = Font(name="Menlo", size=11, color=BRAND["teal"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        row += 1
    row += 1

    # Tracker decisions
    row = add_section_header(ws, row, "Tracker decisions")
    for decision in ["Nominated", "Committed", "Withdrawn", "Attended"]:
        ws.cell(row=row, column=1, value=decision).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
        ws.cell(row=row, column=2, value=f'=COUNTIF(\'Company x Conference Tracker\'!G2:G1000,"{decision}")').font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        bar = ws.cell(
            row=row, column=3,
            value=f'=IFERROR(REPT("█",MIN(40,ROUND(COUNTIF(\'Company x Conference Tracker\'!G2:G1000,"{decision}")/MAX(1,COUNTA(\'Company x Conference Tracker\'!G2:G1000))*40,0))),"")',
        )
        bar.font = Font(name="Menlo", size=11, color=BRAND["orange"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        row += 1


def build() -> str:
    wb = new_workbook()
    add_lookups_tab(
        wb,
        {
            "conference_tier": CONFERENCE_TIER,
            "conference_decision": CONFERENCE_DECISION,
            "fund_codes": FUND_CODES,
        },
    )

    _build_dashboard(wb.create_sheet("Dashboard"))
    _build_catalogue(wb.create_sheet("Conference Catalogue"))
    _build_scoring(wb.create_sheet("Scoring Matrix"))
    _build_tracker(wb.create_sheet("Company x Conference Tracker"))
    _build_commitments(wb.create_sheet("Commitment Letters"))

    order = [
        "Dashboard",
        "Conference Catalogue",
        "Scoring Matrix",
        "Company x Conference Tracker",
        "Commitment Letters",
        "Lookups",
    ]
    for idx, name in enumerate(order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    return str(save_workbook(wb, FILENAME))


if __name__ == "__main__":
    print(build())
