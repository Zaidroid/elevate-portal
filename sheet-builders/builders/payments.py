"""E3 — Payments Tracker workbook.

Tabs: Payments, Advisor Fees, Vendor Fees, Participant Stipends, Summary, Lookups.
"""

from gsg_sheets import (
    BRAND,
    add_dropdown,
    add_kpi_row,
    add_lookups_tab,
    add_section_header,
    band_rows,
    conditional_formatting_date,
    conditional_formatting_status,
    freeze_header,
    new_workbook,
    save_workbook,
    set_column_widths,
    setup_dashboard_tab,
    tab_color,
    write_header,
)
from gsg_sheets.styles import apply_style, cell_style, editable_style, formula_style
from gsg_sheets.taxonomies import (
    FUND_CODES,
    INTERVENTION_TYPES,
    PAYEE_TYPE,
    PAYMENT_STATUS,
)
from openpyxl.styles import Alignment, Font, PatternFill


FILENAME = "E3 - Payments Tracker.xlsx"

PAYMENTS_HEADERS = [
    "payment_id",
    "pr_id",
    "company_id",
    "assignment_id",
    "payee_type",
    "payee_name",
    "intervention_type",
    "fund_code",
    "amount_usd",
    "currency",
    "payment_date",
    "status",
    "finance_contact",
    "invoice_url",
    "receipt_url",
    "needs_review",
    "notes",
    "updated_at",
    "updated_by",
]

ADVISOR_FEES_HEADERS = [
    "fee_id",
    "advisor_name",
    "advisor_email",
    "company_id",
    "engagement_type",
    "hourly_rate_usd",
    "hours",
    "total_usd",
    "precedent_ref",
    "fund_code",
    "intervention_type",
    "payment_id",
    "status",
    "notes",
]

VENDOR_FEES_HEADERS = [
    "fee_id",
    "vendor_name",
    "vendor_contact",
    "service_description",
    "pr_id",
    "amount_usd",
    "currency",
    "fund_code",
    "intervention_type",
    "payment_id",
    "status",
    "notes",
]

PARTICIPANT_STIPENDS_HEADERS = [
    "stipend_id",
    "participant_name",
    "participant_email",
    "company_id",
    "program_component",
    "amount_usd",
    "fund_code",
    "month",
    "payment_id",
    "status",
    "notes",
]


def _style_editable_rows(ws, cols, start_row=2, end_row=500):
    style_e = editable_style()
    for row in range(start_row, end_row + 1):
        for col in cols:
            cell = ws.cell(row=row, column=col)
            apply_style(cell, style_e)


def _build_payments(ws):
    tab_color(ws, BRAND["red"])
    write_header(ws, PAYMENTS_HEADERS)
    freeze_header(ws)

    style_f = formula_style()
    for row in range(2, 501):
        cell = ws.cell(
            row=row,
            column=1,
            value=f'=IF(F{row}<>"", "PAY-E3-"&TEXT(ROW()-1,"0000"), "")',
        )
        apply_style(cell, style_f)
        ws.cell(row=row, column=13).value = "Khamis Eweis"
        ws.cell(row=row, column=10).value = "USD"

    add_dropdown(ws, "E", "=payee_type")
    add_dropdown(ws, "G", "=intervention_types")
    add_dropdown(ws, "H", "=fund_codes")
    add_dropdown(ws, "J", '"USD,ILS,EUR"')
    add_dropdown(ws, "L", "=payment_status")
    add_dropdown(ws, "P", '"TRUE,FALSE"')  # needs_review flag

    editable = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17]
    _style_editable_rows(ws, editable)

    # Phase K polish: status badges, overdue dates, needs_review red, banded rows.
    conditional_formatting_status(
        ws, "L",
        good=["Paid", "Approved"],
        warn=["Pending Approval", "Sent to Finance"],
        bad=["Rejected"],
        last_row=500,
    )
    conditional_formatting_date(ws, "K", last_row=500)
    conditional_formatting_status(ws, "P", bad=["TRUE"], good=["FALSE"], last_row=500)
    band_rows(ws, last_col_letter="S", last_row=500)

    set_column_widths(
        ws,
        {
            1: 16, 2: 16, 3: 12, 4: 12, 5: 14, 6: 26, 7: 22, 8: 18,
            9: 14, 10: 10, 11: 14, 12: 18, 13: 18, 14: 36, 15: 36,
            16: 14, 17: 40, 18: 18, 19: 20,
        },
    )


def _build_advisor_fees(ws):
    tab_color(ws, BRAND["teal"])
    write_header(ws, ADVISOR_FEES_HEADERS)
    freeze_header(ws)

    style_f = formula_style()
    for row in range(2, 301):
        ws.cell(row=row, column=1, value=f'=IF(B{row}<>"", "AF-E3-"&TEXT(ROW()-1,"000"), "")')
        apply_style(ws.cell(row=row, column=1), style_f)
        cell = ws.cell(row=row, column=8, value=f'=IF(AND(F{row}<>"",G{row}<>""), F{row}*G{row}, "")')
        apply_style(cell, style_f)
        # Default precedent reference. Per the master context, C-Suite Local
        # Advisor precedent is 30h x $55 = $1,650.
        ws.cell(row=row, column=9).value = "C-Suite Local Advisor: 30h x $55 = $1,650"

    add_dropdown(ws, "E", '"C-Suite Coaching,Domain Advisor,Technical Mentor,Board Advisor,Legal,Financial,Other"')
    add_dropdown(ws, "J", "=fund_codes")
    add_dropdown(ws, "K", "=intervention_types")
    add_dropdown(ws, "M", "=payment_status")

    _style_editable_rows(ws, [2, 3, 4, 5, 6, 7, 9, 10, 11, 12, 13, 14], end_row=300)

    set_column_widths(
        ws,
        {1: 14, 2: 26, 3: 28, 4: 12, 5: 22, 6: 14, 7: 10, 8: 14,
         9: 40, 10: 18, 11: 22, 12: 16, 13: 16, 14: 40},
    )


def _build_vendor_fees(ws):
    tab_color(ws, BRAND["orange"])
    write_header(ws, VENDOR_FEES_HEADERS)
    freeze_header(ws)

    style_f = formula_style()
    for row in range(2, 301):
        ws.cell(row=row, column=1, value=f'=IF(B{row}<>"", "VF-E3-"&TEXT(ROW()-1,"000"), "")')
        apply_style(ws.cell(row=row, column=1), style_f)
        ws.cell(row=row, column=7).value = "USD"

    add_dropdown(ws, "H", "=fund_codes")
    add_dropdown(ws, "I", "=intervention_types")
    add_dropdown(ws, "G", '"USD,ILS,EUR"')
    add_dropdown(ws, "K", "=payment_status")

    _style_editable_rows(ws, [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12], end_row=300)
    set_column_widths(
        ws,
        {1: 14, 2: 26, 3: 22, 4: 36, 5: 14, 6: 14, 7: 10, 8: 18, 9: 22,
         10: 16, 11: 16, 12: 40},
    )


def _build_participant_stipends(ws):
    tab_color(ws, BRAND["navy"])
    write_header(ws, PARTICIPANT_STIPENDS_HEADERS)
    freeze_header(ws)

    style_f = formula_style()
    for row in range(2, 501):
        ws.cell(row=row, column=1, value=f'=IF(B{row}<>"", "ST-E3-"&TEXT(ROW()-1,"0000"), "")')
        apply_style(ws.cell(row=row, column=1), style_f)

    add_dropdown(ws, "G", "=fund_codes")
    add_dropdown(ws, "J", "=payment_status")

    _style_editable_rows(ws, [2, 3, 4, 5, 6, 7, 8, 9, 10, 11])
    set_column_widths(
        ws,
        {1: 14, 2: 26, 3: 28, 4: 12, 5: 22, 6: 14, 7: 18, 8: 12,
         9: 16, 10: 16, 11: 40},
    )


def _build_summary(ws):
    tab_color(ws, BRAND["navy"])
    write_header(
        ws,
        ["Fund code", "Intervention", "Total paid (USD)", "Pending approval (USD)", "Sent to finance (USD)"],
    )
    freeze_header(ws)
    style_f = formula_style()
    row = 2
    for fund in FUND_CODES:
        for intervention in INTERVENTION_TYPES:
            ws.cell(row=row, column=1, value=fund)
            ws.cell(row=row, column=2, value=intervention)
            cell = ws.cell(
                row=row,
                column=3,
                value=(
                    f'=SUMIFS(Payments!I:I,Payments!H:H,A{row},'
                    f'Payments!G:G,B{row},Payments!L:L,"Paid")'
                ),
            )
            apply_style(cell, style_f)
            cell = ws.cell(
                row=row,
                column=4,
                value=(
                    f'=SUMIFS(Payments!I:I,Payments!H:H,A{row},'
                    f'Payments!G:G,B{row},Payments!L:L,"Pending Approval")'
                ),
            )
            apply_style(cell, style_f)
            cell = ws.cell(
                row=row,
                column=5,
                value=(
                    f'=SUMIFS(Payments!I:I,Payments!H:H,A{row},'
                    f'Payments!G:G,B{row},Payments!L:L,"Sent to Finance")'
                ),
            )
            apply_style(cell, style_f)
            row += 1

    set_column_widths(ws, {1: 18, 2: 22, 3: 18, 4: 22, 5: 22})


def _build_dashboard(ws):
    """Mirrors pages/payments/PaymentsPage.tsx in-sheet."""
    setup_dashboard_tab(ws, brand_tab_color=BRAND["red"])

    title = ws.cell(row=1, column=1, value="Payments Dashboard")
    title.font = Font(name="Source Sans Pro", bold=True, size=18, color=BRAND["navy"])
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    sub = ws.cell(row=2, column=1, value="Live mirror of the Payments module in the Elevate Portal.")
    sub.font = Font(name="Source Sans Pro", size=10, color=BRAND["muted_text"])
    sub.alignment = Alignment(horizontal="left", indent=1)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)

    row = 4
    row = add_section_header(ws, row, "Top metrics")
    row = add_kpi_row(ws, row, [
        {"label": "Payments", "formula": "=COUNTA(Payments!F2:F500)", "tone": "navy"},
        {"label": "Total paid (USD)", "formula": "=SUMIFS(Payments!I2:I500,Payments!L2:L500,\"Paid\")", "tone": "green", "value_fmt": "$#,##0"},
        {"label": "Pending approval", "formula": "=SUMIFS(Payments!I2:I500,Payments!L2:L500,\"Pending Approval\")", "tone": "amber", "value_fmt": "$#,##0"},
        {"label": "Needs review", "formula": "=COUNTIF(Payments!P2:P500,\"TRUE\")", "tone": "red"},
    ])

    # By status breakdown
    row = add_section_header(ws, row, "By status")
    for status in PAYMENT_STATUS:
        ws.cell(row=row, column=1, value=status)
        ws.cell(row=row, column=1).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
        ws.cell(row=row, column=2, value=f'=SUMIFS(Payments!I2:I500,Payments!L2:L500,"{status}")')
        ws.cell(row=row, column=2).number_format = "$#,##0"
        ws.cell(row=row, column=2).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        bar = ws.cell(
            row=row, column=3,
            value=(
                f'=IFERROR(REPT("█",MIN(40,ROUND('
                f'SUMIFS(Payments!I2:I500,Payments!L2:L500,"{status}")/'
                f'MAX(1,SUM(Payments!I2:I500))*40,0))),"")'
            ),
        )
        bar.font = Font(name="Menlo", size=11, color=BRAND["red"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        row += 1
    row += 1

    # By fund
    row = add_section_header(ws, row, "By fund")
    for fund in FUND_CODES:
        ws.cell(row=row, column=1, value=fund)
        ws.cell(row=row, column=1).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
        ws.cell(row=row, column=2, value=f'=SUMIFS(Payments!I2:I500,Payments!H2:H500,"{fund}")')
        ws.cell(row=row, column=2).number_format = "$#,##0"
        ws.cell(row=row, column=2).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        bar = ws.cell(
            row=row, column=3,
            value=(
                f'=IFERROR(REPT("█",MIN(40,ROUND('
                f'SUMIFS(Payments!I2:I500,Payments!H2:H500,"{fund}")/'
                f'MAX(1,SUM(Payments!I2:I500))*40,0))),"")'
            ),
        )
        bar.font = Font(name="Menlo", size=11, color=BRAND["teal"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        row += 1
    row += 1

    # Needs review queue summary
    row = add_section_header(ws, row, "Needs review")
    ws.cell(row=row, column=1, value="Rows flagged needs_review = TRUE")
    ws.cell(row=row, column=1).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
    ws.cell(row=row, column=2, value="=COUNTIF(Payments!P2:P500,\"TRUE\")")
    ws.cell(row=row, column=2).font = Font(name="Source Sans Pro", size=12, color=BRAND["red"], bold=True)


def build() -> str:
    wb = new_workbook()
    add_lookups_tab(
        wb,
        {
            "intervention_types": INTERVENTION_TYPES,
            "fund_codes": FUND_CODES,
            "payee_type": PAYEE_TYPE,
            "payment_status": PAYMENT_STATUS,
        },
    )

    _build_dashboard(wb.create_sheet("Dashboard"))
    _build_payments(wb.create_sheet("Payments"))
    _build_advisor_fees(wb.create_sheet("Advisor Fees"))
    _build_vendor_fees(wb.create_sheet("Vendor Fees"))
    _build_participant_stipends(wb.create_sheet("Participant Stipends"))
    _build_summary(wb.create_sheet("Summary"))

    order = ["Dashboard", "Payments", "Advisor Fees", "Vendor Fees", "Participant Stipends", "Summary", "Lookups"]
    for idx, name in enumerate(order):
        wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    return str(save_workbook(wb, FILENAME))


if __name__ == "__main__":
    print(build())
