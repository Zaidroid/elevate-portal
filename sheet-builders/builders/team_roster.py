"""E3 — Team Roster workbook.

Source of truth for portal authentication. Fallbacks to the hardcoded roster
in elevate-portal/src/config/team.ts if the sheet is unavailable at login time.

Tabs: Team, Roles, Lookups.
"""

from gsg_sheets import (
    BRAND,
    add_dropdown,
    add_kpi_row,
    add_lookups_tab,
    add_section_header,
    band_rows,
    conditional_formatting_status,
    freeze_header,
    new_workbook,
    save_workbook,
    set_column_widths,
    setup_dashboard_tab,
    tab_color,
    write_header,
    write_rows,
)
from gsg_sheets.styles import apply_style, editable_style
from openpyxl.styles import Alignment, Font


FILENAME = "E3 - Team Roster.xlsx"

TEAM_HEADERS = [
    "email",
    "name",
    "role",
    "active",
    "title",
    "team",
    "notes",
    "updated_at",
    "updated_by",
]

# Seed from existing hardcoded rosters across elevate-portal / selection-tool / Advisors.
SEED_ROSTER = [
    # (email, name, role, active, title, team, notes)
    ("zaid@gazaskygeeks.com", "Zaid Salem", "admin", "Yes", "Market Access Officer", "Companies", ""),
    ("zsalem33@gmail.com", "Zaid Salem", "admin", "Yes", "Market Access Officer", "Companies", "Personal Google account used during build"),
    ("israa@gazaskygeeks.com", "Israa Hamoudeh", "user", "Yes", "Team Lead", "Companies", ""),
    ("doaa@gazaskygeeks.com", "Doaa Younis", "user", "Yes", "TTH and Upskilling", "Companies", ""),
    ("ayesh@gazaskygeeks.com", "Mohammed Ayesh", "user", "Yes", "Market Access and Legal", "Companies", ""),
    ("raouf@gazaskygeeks.com", "Raouf Said", "user", "Yes", "Co-working Spaces", "Companies", ""),
    ("muna@gazaskygeeks.com", "Muna Mahroum", "user", "Yes", "Pre-TTH", "Companies", ""),
    ("mzourob@gazaskygeeks.com", "Mohammed Zourob", "user", "Yes", "ElevateBridge / Freelancers", "Companies", ""),
    ("rand@gazaskygeeks.com", "Rand Safi", "admin", "Yes", "Senior Program Manager", "Leadership", "Initiated Travel Committee"),
]

ROLES_HEADERS = ["role", "description", "can_approve_payments", "can_edit_roster", "can_export_reports"]
ROLES = [
    ("admin", "Full access to all modules, approvals, and exports", "Yes", "Yes", "Yes"),
    ("user", "Standard team member, create and edit own module records", "No", "No", "Yes"),
    ("viewer", "Read-only access to dashboards and reports", "No", "No", "No"),
]


def _build_team(ws):
    tab_color(ws, BRAND["red"])
    write_header(ws, TEAM_HEADERS)
    freeze_header(ws)
    style_e = editable_style()

    rows = [list(r) + ["", ""] for r in SEED_ROSTER]
    end = write_rows(ws, rows, start_row=2, editable_cols=[1, 2, 3, 4, 5, 6, 7])

    # Apply editable styling to remaining blank rows up to 200.
    for row in range(end + 1, 201):
        for col in range(1, 8):
            apply_style(ws.cell(row=row, column=col), style_e)

    add_dropdown(ws, "C", "=roles")
    add_dropdown(ws, "D", '"Yes,No"')

    # Phase K polish
    conditional_formatting_status(
        ws, "C", good=["admin"], warn=["user"], neutral=["viewer"], last_row=200
    )
    conditional_formatting_status(ws, "D", good=["Yes"], bad=["No"], last_row=200)
    band_rows(ws, last_col_letter="I", last_row=200)

    set_column_widths(
        ws,
        {1: 30, 2: 24, 3: 12, 4: 10, 5: 26, 6: 18, 7: 36, 8: 18, 9: 20},
    )


def _build_roles(ws):
    tab_color(ws, BRAND["teal"])
    write_header(ws, ROLES_HEADERS)
    freeze_header(ws)
    style_e = editable_style()
    write_rows(ws, ROLES, start_row=2, editable_cols=[2, 3, 4, 5])
    add_dropdown(ws, "C", '"Yes,No"')
    add_dropdown(ws, "D", '"Yes,No"')
    add_dropdown(ws, "E", '"Yes,No"')
    set_column_widths(ws, {1: 14, 2: 44, 3: 22, 4: 18, 5: 22})


def _build_dashboard(ws):
    """Mirrors pages/team/TeamPage.tsx in-sheet."""
    setup_dashboard_tab(ws, brand_tab_color=BRAND["red"])

    title = ws.cell(row=1, column=1, value="Team Roster Dashboard")
    title.font = Font(name="Source Sans Pro", bold=True, size=18, color=BRAND["navy"])
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    sub = ws.cell(row=2, column=1, value="Live mirror of the Team Roster module in the Elevate Portal.")
    sub.font = Font(name="Source Sans Pro", size=10, color=BRAND["muted_text"])
    sub.alignment = Alignment(horizontal="left", indent=1)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)

    row = 4
    row = add_section_header(ws, row, "Top metrics")
    row = add_kpi_row(ws, row, [
        {"label": "Members", "formula": "=COUNTA(Team!A2:A200)", "tone": "navy"},
        {"label": "Admins", "formula": '=COUNTIF(Team!C2:C200,"admin")', "tone": "red"},
        {"label": "Active", "formula": '=COUNTIF(Team!D2:D200,"Yes")', "tone": "green"},
        {"label": "Inactive", "formula": '=COUNTIF(Team!D2:D200,"No")', "tone": "slate"},
    ])

    row = add_section_header(ws, row, "By role")
    for role in ["admin", "user", "viewer"]:
        ws.cell(row=row, column=1, value=role.capitalize()).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
        ws.cell(row=row, column=2, value=f'=COUNTIF(Team!C2:C200,"{role}")').font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        bar = ws.cell(
            row=row, column=3,
            value=f'=IFERROR(REPT("█",MIN(40,ROUND(COUNTIF(Team!C2:C200,"{role}")/MAX(1,COUNTA(Team!C2:C200))*40,0))),"")',
        )
        bar.font = Font(name="Menlo", size=11, color=BRAND["red"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        row += 1


def build() -> str:
    wb = new_workbook()
    add_lookups_tab(wb, {"roles": ["admin", "user", "viewer"]})
    _build_dashboard(wb.create_sheet("Dashboard"))
    _build_team(wb.create_sheet("Team"))
    _build_roles(wb.create_sheet("Roles"))

    order = ["Dashboard", "Team", "Roles", "Lookups"]
    for idx, name in enumerate(order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    return str(save_workbook(wb, FILENAME))


if __name__ == "__main__":
    print(build())
