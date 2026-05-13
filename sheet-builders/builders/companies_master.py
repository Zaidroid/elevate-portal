"""E3 — Companies Master workbook.

Tabs (12, in display order):
    Dashboard, Companies, Contacts, Intervention Assignments,
    Reviews, Company Comments, Activity Log, Interview Aliases,
    Removed Companies, Pre-decision Recommendations,
    Status Log, Historical Interventions, Lookups

The 5 portal-managed tabs (Reviews / Company Comments / Activity Log /
Interview Aliases / Removed Companies / Pre-decision Recommendations) are
created here at build time with their canonical headers + tab styling
so the portal's `ensureSchema` call is a no-op when it first opens the
workbook. Without this, fresh workbooks land in a half-built state and
the team sees missing tabs until ensureSchema fires.
"""

from gsg_sheets import (
    BRAND,
    add_dropdown,
    add_kpi_row,
    add_lookups_tab,
    add_section_header,
    apply_row_style,
    band_rows,
    conditional_formatting_status,
    freeze_header,
    new_workbook,
    save_workbook,
    set_column_widths,
    setup_dashboard_tab,
    tab_color,
    write_header,
    write_rows,
)
from gsg_sheets.styles import cell_style, editable_style, note_style
from gsg_sheets.taxonomies import (
    ASSIGNMENT_STATUS,
    COHORT,
    COMPANY_STAGE,
    COMPANY_STATUS,
    FUND_CODES,
    GOVERNORATES,
    INTERVENTION_TYPES,
    SECTORS,
)
from openpyxl.styles import Alignment, Font


FILENAME = "E3 - Companies Master.xlsx"

COMPANIES_HEADERS = [
    "company_id",
    "company_name",
    "legal_name",
    "city",
    "governorate",
    "sector",
    "employee_count",
    "revenue_bracket",
    "international_revenue_pct",
    "readiness_score",
    "fund_code",
    "cohort",
    "status",
    "stage",
    "profile_manager_email",
    "selection_date",
    "onboarding_date",
    "primary_contact_id",
    "drive_folder_url",
    "notes",
    "updated_at",
    "updated_by",
]

CONTACTS_HEADERS = [
    "contact_id",
    "company_id",
    "full_name",
    "title",
    "email",
    "phone",
    "is_signatory",
    "notes",
    "updated_at",
    "updated_by",
]

ASSIGNMENTS_HEADERS = [
    "assignment_id",
    "company_id",
    "intervention_type",
    "sub_intervention",
    "fund_code",
    "start_date",
    "end_date",
    "owner_email",
    "status",
    "budget_usd",
    "procurement_pr_id",
    "notes",
    "updated_at",
    "updated_by",
]

STATUS_LOG_HEADERS = [
    "log_id",
    "company_id",
    "from_status",
    "to_status",
    "changed_at",
    "changed_by",
    "reason",
]

# ─── Portal-managed tab schemas (must mirror src/pages/companies/reviewTypes.ts) ──
# Reviews: one row per (reviewer, company) — captures the reviewer's
# proposed pillars, sub-interventions, decision verdict, and notes.
REVIEWS_HEADERS = [
    "review_id",
    "company_id",
    "reviewer_email",
    "decision",
    "proposed_pillars",
    "proposed_sub_interventions",
    "notes",
    "created_at",
    "updated_at",
]

# Company Comments: free-form per-company thread, used in the Review
# view's team thread tab + the Final Decision dossier.
COMMENTS_HEADERS = [
    "comment_id",
    "company_id",
    "author_email",
    "body",
    "created_at",
    "updated_at",
]

# Activity Log: append-only audit trail of every meaningful write.
ACTIVITY_HEADERS = [
    "activity_id",
    "timestamp",
    "user_email",
    "company_id",
    "action",
    "field",
    "old_value",
    "new_value",
    "details",
]

# Interview Aliases: maps schedule-name → applicant_company_name so the
# team's interview spreadsheet entries resolve to Source Data rows.
ALIAS_HEADERS = [
    "alias_id",
    "schedule_name",
    "applicant_company_name",
    "created_by",
    "created_at",
    "updated_at",
    "updated_by",
]

# Removed Companies: shared exclusion list — any name here is hidden
# from the review queue + materialize candidates + joined views.
REMOVED_HEADERS = [
    "removed_id",
    "company_name",
    "removed_by",
    "removed_at",
    "reason",
]

# Pre-decision Recommendations: structured per-company intervention
# recommendations made before the final decision session. Sourced from
# Israa's voting CSV / Raouf's notes docx / future seeds via the
# importer. The Final Decision view's pre-fill chain pulls from here
# right after existing locks.
PRE_DECISION_HEADERS = [
    "recommendation_id",
    "company_id",
    "company_name",
    "author_email",
    "pillar",
    "sub_intervention",
    "fund_hint",
    "note",
    "source",
    "created_at",
]


# Historical Interventions captures Elevate 1 and Elevate 2 company x intervention
# pairings imported from Companies Data Tracker (1).xlsx > 2. Participating Companies.
# Kept separate from the live Intervention Assignments tab so E3 roster queries stay clean.
HISTORICAL_INTERVENTIONS_HEADERS = [
    "historical_id",
    "cohort",
    "company_name",
    "company_id",
    "offering_name",
    "specialization",
    "cohort_name",
    "donor",
    "fund_code",
    "start_date",
    "end_date",
    "year",
    "agreement_link",
    "source",
    "notes",
    "updated_at",
    "updated_by",
]


def _build_companies(ws):
    tab_color(ws, BRAND["red"])
    write_header(ws, COMPANIES_HEADERS)
    freeze_header(ws)

    # Pre-generate an ID formula column. company_id auto-populates as E3-0001 when
    # company_name is entered.
    id_col = 1
    name_col = 2
    for row in range(2, 501):
        ws.cell(
            row=row,
            column=id_col,
            value=f'=IF(B{row}<>"", "E3-"&TEXT(ROW()-1, "0000"), "")',
        )

    # Data validation
    add_dropdown(ws, "E", "=governorates")  # governorate
    add_dropdown(ws, "F", "=sectors")
    add_dropdown(ws, "K", "=fund_codes")
    add_dropdown(ws, "L", f'"{COHORT}"')
    add_dropdown(ws, "M", "=company_status")
    add_dropdown(ws, "N", "=company_stage")  # funnel stage
    # profile_manager_email (col O) is free-text; portal will present a dropdown
    # sourced from the live Team Roster sheet filtered to Profile Manager tier.

    set_column_widths(
        ws,
        {
            1: 12,   # company_id
            2: 32,   # company_name
            3: 28,   # legal_name
            4: 16,   # city
            5: 18,   # governorate
            6: 20,   # sector
            7: 10,   # employee_count
            8: 18,   # revenue_bracket
            9: 12,   # international_revenue_pct
            10: 14,  # readiness_score
            11: 18,  # fund_code
            12: 10,  # cohort
            13: 14,  # status
            14: 16,  # stage
            15: 28,  # profile_manager_email
            16: 14,  # selection_date
            17: 14,  # onboarding_date
            18: 14,  # primary_contact_id
            19: 32,  # drive_folder_url
            20: 40,  # notes
            21: 18,  # updated_at
            22: 20,  # updated_by
        },
    )

    # Mark user-editable columns blue for rows 2..500 (everything except company_id
    # and updated_at/updated_by which the app manages).
    style_e = editable_style()
    for row in range(2, 501):
        for col in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]:
            cell = ws.cell(row=row, column=col)
            for attr, value in style_e.items():
                setattr(cell, attr, value)

    # Phase K polish: status tones, banded rows.
    conditional_formatting_status(
        ws, "M",
        good=["Active", "Graduated", "Onboarded"],
        warn=["Selected", "Interviewed"],
        bad=["Withdrawn"],
        neutral=["Applicant", "Shortlisted"],
        last_row=500,
    )
    band_rows(ws, last_col_letter="V", last_row=500)
    # Hide the ID + timestamp + audit columns by default. The team
    # works in human terms (name / sector / city) and only the portal
    # cares about the IDs. They're still here on un-hide if needed.
    _hide_columns(ws, [1, 18, 21, 22])  # company_id, primary_contact_id, updated_at, updated_by


def _build_contacts(ws):
    tab_color(ws, BRAND["teal"])
    write_header(ws, CONTACTS_HEADERS)
    freeze_header(ws)

    # contact_id formula: C3-0001 when full_name is entered
    for row in range(2, 1001):
        ws.cell(
            row=row,
            column=1,
            value=f'=IF(C{row}<>"", "C3-"&TEXT(ROW()-1, "0000"), "")',
        )

    add_dropdown(ws, "G", '"Yes,No"')  # is_signatory

    set_column_widths(
        ws,
        {1: 12, 2: 12, 3: 26, 4: 22, 5: 28, 6: 18, 7: 12, 8: 36, 9: 18, 10: 20},
    )

    style_e = editable_style()
    for row in range(2, 1001):
        for col in [2, 3, 4, 5, 6, 7, 8]:
            cell = ws.cell(row=row, column=col)
            for attr, value in style_e.items():
                setattr(cell, attr, value)
    # Hide IDs + audit cols on Contacts.
    _hide_columns(ws, [1, 2, 9, 10])  # contact_id, company_id, updated_at, updated_by


def _build_assignments(ws):
    tab_color(ws, BRAND["orange"])
    write_header(ws, ASSIGNMENTS_HEADERS)
    freeze_header(ws)

    for row in range(2, 1001):
        ws.cell(
            row=row,
            column=1,
            value=f'=IF(B{row}<>"", "A3-"&TEXT(ROW()-1, "0000"), "")',
        )

    add_dropdown(ws, "C", "=intervention_types")
    add_dropdown(ws, "E", "=fund_codes")
    add_dropdown(ws, "I", "=assignment_status")

    set_column_widths(
        ws,
        {
            1: 12,
            2: 12,
            3: 22,
            4: 22,
            5: 18,
            6: 14,
            7: 14,
            8: 28,
            9: 14,
            10: 14,
            11: 14,
            12: 40,
            13: 18,
            14: 20,
        },
    )

    style_e = editable_style()
    for row in range(2, 1001):
        for col in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]:
            cell = ws.cell(row=row, column=col)
            for attr, value in style_e.items():
                setattr(cell, attr, value)
    # Hide IDs + audit cols on Intervention Assignments.
    _hide_columns(ws, [1, 2, 13, 14])  # assignment_id, company_id, updated_at, updated_by


def _build_status_log(ws):
    tab_color(ws, BRAND["navy"])
    write_header(ws, STATUS_LOG_HEADERS)
    freeze_header(ws)
    set_column_widths(ws, {1: 12, 2: 12, 3: 14, 4: 14, 5: 20, 6: 26, 7: 40})
    _hide_columns(ws, [1, 2])  # log_id, company_id


def _build_historical_interventions(ws):
    tab_color(ws, BRAND["navy"])
    write_header(ws, HISTORICAL_INTERVENTIONS_HEADERS)
    freeze_header(ws)

    # historical_id auto-mints HI-0001 when company_name is filled.
    for row in range(2, 2001):
        ws.cell(
            row=row,
            column=1,
            value=f'=IF(C{row}<>"", "HI-"&TEXT(ROW()-1, "0000"), "")',
        )

    add_dropdown(ws, "B", '"E1,E2"')
    add_dropdown(ws, "H", '"Dutch,SIDA"')
    add_dropdown(ws, "I", "=fund_codes")

    set_column_widths(
        ws,
        {
            1: 12,   # historical_id
            2: 10,   # cohort
            3: 28,   # company_name
            4: 12,   # company_id (FK if match found)
            5: 26,   # offering_name
            6: 26,   # specialization
            7: 22,   # cohort_name
            8: 14,   # donor
            9: 14,   # fund_code
            10: 14,  # start_date
            11: 14,  # end_date
            12: 10,  # year
            13: 32,  # agreement_link
            14: 36,  # source
            15: 40,  # notes
            16: 18,  # updated_at
            17: 20,  # updated_by
        },
    )

    style_e = editable_style()
    for row in range(2, 2001):
        for col in range(2, 16):
            cell = ws.cell(row=row, column=col)
            for attr, value in style_e.items():
                setattr(cell, attr, value)
    _hide_columns(ws, [1, 4, 16, 17])  # historical_id, company_id, updated_at, updated_by


def _hide_columns(ws, columns):
    """Hide the given 1-indexed columns. Used for ID + audit columns the
    team rarely needs to see — keeps tabs human-readable."""
    from openpyxl.utils import get_column_letter
    for col_idx in columns:
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True


def _build_reviews(ws):
    """Reviews tab — one row per (reviewer, company)."""
    tab_color(ws, BRAND["teal"])
    write_header(ws, REVIEWS_HEADERS)
    freeze_header(ws)
    set_column_widths(
        ws,
        {
            1: 20,   # review_id
            2: 12,   # company_id
            3: 28,   # reviewer_email
            4: 14,   # decision
            5: 30,   # proposed_pillars
            6: 36,   # proposed_sub_interventions
            7: 50,   # notes
            8: 18,   # created_at
            9: 18,   # updated_at
        },
    )
    band_rows(ws, last_col_letter="I", last_row=2000)
    # Hide IDs + timestamps to keep the row readable.
    _hide_columns(ws, [1, 2, 8, 9])


def _build_comments(ws):
    """Company Comments tab — free-form per-company thread."""
    tab_color(ws, BRAND["navy"])
    write_header(ws, COMMENTS_HEADERS)
    freeze_header(ws)
    set_column_widths(
        ws,
        {
            1: 24,   # comment_id
            2: 12,   # company_id
            3: 28,   # author_email
            4: 80,   # body
            5: 18,   # created_at
            6: 18,   # updated_at
        },
    )
    band_rows(ws, last_col_letter="F", last_row=2000)
    _hide_columns(ws, [1, 2, 6])


def _build_activity(ws):
    """Activity Log tab — append-only audit trail."""
    tab_color(ws, BRAND["orange"])
    write_header(ws, ACTIVITY_HEADERS)
    freeze_header(ws)
    set_column_widths(
        ws,
        {
            1: 28,   # activity_id
            2: 18,   # timestamp
            3: 26,   # user_email
            4: 12,   # company_id
            5: 18,   # action
            6: 18,   # field
            7: 22,   # old_value
            8: 22,   # new_value
            9: 60,   # details
        },
    )
    band_rows(ws, last_col_letter="I", last_row=2000)
    _hide_columns(ws, [1, 4])


def _build_aliases(ws):
    """Interview Aliases tab — schedule_name → applicant_company_name overrides."""
    tab_color(ws, BRAND["teal"])
    write_header(ws, ALIAS_HEADERS)
    freeze_header(ws)
    set_column_widths(
        ws,
        {
            1: 20,   # alias_id
            2: 28,   # schedule_name
            3: 28,   # applicant_company_name
            4: 26,   # created_by
            5: 18,   # created_at
            6: 18,   # updated_at
            7: 26,   # updated_by
        },
    )
    band_rows(ws, last_col_letter="G", last_row=1000)
    _hide_columns(ws, [1, 5, 6, 7])


def _build_removed(ws):
    """Removed Companies tab — shared exclusion list."""
    tab_color(ws, BRAND["red"])
    write_header(ws, REMOVED_HEADERS)
    freeze_header(ws)
    set_column_widths(
        ws,
        {
            1: 20,   # removed_id
            2: 28,   # company_name
            3: 26,   # removed_by
            4: 18,   # removed_at
            5: 50,   # reason
        },
    )
    band_rows(ws, last_col_letter="E", last_row=500)
    _hide_columns(ws, [1])


def _build_pre_decisions(ws):
    """Pre-decision Recommendations — Israa CSV / Raouf docx / future seeds."""
    tab_color(ws, BRAND["orange"])
    write_header(ws, PRE_DECISION_HEADERS)
    freeze_header(ws)
    set_column_widths(
        ws,
        {
            1: 28,   # recommendation_id
            2: 12,   # company_id
            3: 28,   # company_name
            4: 26,   # author_email
            5: 16,   # pillar
            6: 22,   # sub_intervention
            7: 12,   # fund_hint
            8: 60,   # note
            9: 22,   # source
            10: 18,  # created_at
        },
    )
    add_dropdown(ws, "E", "=intervention_types")
    add_dropdown(ws, "G", "=fund_codes")
    band_rows(ws, last_col_letter="J", last_row=2000)
    _hide_columns(ws, [1, 2, 10])


def _build_dashboard(ws):
    """Mirrors pages/companies/CompaniesPage.tsx + the BoardPage funnel."""
    setup_dashboard_tab(ws, brand_tab_color=BRAND["red"])

    title = ws.cell(row=1, column=1, value="Companies Master Dashboard")
    title.font = Font(name="Source Sans Pro", bold=True, size=18, color=BRAND["navy"])
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    sub = ws.cell(row=2, column=1, value="Live mirror of the Companies module in the Elevate Portal.")
    sub.font = Font(name="Source Sans Pro", size=10, color=BRAND["muted_text"])
    sub.alignment = Alignment(horizontal="left", indent=1)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)

    row = 4
    row = add_section_header(ws, row, "Top metrics")
    row = add_kpi_row(ws, row, [
        {"label": "Companies", "formula": "=COUNTA(Companies!B2:B500)", "tone": "navy"},
        {"label": "Active", "formula": "=COUNTIF(Companies!M2:M2000,\"Active\")", "tone": "green"},
        {"label": "Onboarded", "formula": "=COUNTIF(Companies!M2:M2000,\"Onboarded\")", "tone": "teal"},
        {"label": "Withdrawn", "formula": "=COUNTIF(Companies!M2:M2000,\"Withdrawn\")", "tone": "red"},
    ])

    # Funnel
    row = add_section_header(ws, row, "Funnel by status")
    for status in COMPANY_STATUS:
        ws.cell(row=row, column=1, value=status).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
        ws.cell(row=row, column=2, value=f'=COUNTIF(Companies!M2:M2000,"{status}")').font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        bar = ws.cell(
            row=row, column=3,
            value=(
                f'=IFERROR(REPT("█",MIN(40,ROUND(COUNTIF(Companies!M2:M2000,"{status}")/'
                f'MAX(1,COUNTA(Companies!M2:M2000))*40,0))),"")'
            ),
        )
        bar.font = Font(name="Menlo", size=11, color=BRAND["red"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        row += 1
    row += 1

    # By fund
    row = add_section_header(ws, row, "By fund")
    for fund in FUND_CODES:
        ws.cell(row=row, column=1, value=fund).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
        ws.cell(row=row, column=2, value=f'=COUNTIF(Companies!K2:K2000,"{fund}")').font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        bar = ws.cell(
            row=row, column=3,
            value=(
                f'=IFERROR(REPT("█",MIN(40,ROUND(COUNTIF(Companies!K2:K2000,"{fund}")/'
                f'MAX(1,COUNTA(Companies!K2:K2000))*40,0))),"")'
            ),
        )
        bar.font = Font(name="Menlo", size=11, color=BRAND["teal"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        row += 1
    row += 1

    # Assignments by intervention type
    row = add_section_header(ws, row, "Active assignments by intervention")
    for intervention in INTERVENTION_TYPES:
        ws.cell(row=row, column=1, value=intervention).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
        ws.cell(row=row, column=2, value=f'=COUNTIF(\'Intervention Assignments\'!C2:C2000,"{intervention}")').font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        bar = ws.cell(
            row=row, column=3,
            value=(
                f'=IFERROR(REPT("█",MIN(40,ROUND(COUNTIF(\'Intervention Assignments\'!C2:C2000,"{intervention}")/'
                f'MAX(1,COUNTA(\'Intervention Assignments\'!C2:C2000))*40,0))),"")'
            ),
        )
        bar.font = Font(name="Menlo", size=11, color=BRAND["orange"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        row += 1


def build() -> str:
    wb = new_workbook()

    dashboard = wb.create_sheet("Dashboard")
    companies = wb.create_sheet("Companies")
    contacts = wb.create_sheet("Contacts")
    assignments = wb.create_sheet("Intervention Assignments")
    # Portal-managed tabs — created with canonical headers so the
    # ensureSchema pass on first portal mount is a no-op.
    reviews = wb.create_sheet("Reviews")
    comments = wb.create_sheet("Company Comments")
    activity = wb.create_sheet("Activity Log")
    aliases = wb.create_sheet("Interview Aliases")
    removed = wb.create_sheet("Removed Companies")
    pre_decisions = wb.create_sheet("Pre-decision Recommendations")
    log = wb.create_sheet("Status Log")
    historical = wb.create_sheet("Historical Interventions")

    # Lookups first so named ranges exist before other tabs reference them.
    add_lookups_tab(
        wb,
        {
            "intervention_types": INTERVENTION_TYPES,
            "fund_codes": FUND_CODES,
            "company_status": COMPANY_STATUS,
            "company_stage": COMPANY_STAGE,
            "assignment_status": ASSIGNMENT_STATUS,
            "governorates": GOVERNORATES,
            "sectors": SECTORS,
        },
    )

    _build_companies(companies)
    _build_contacts(contacts)
    _build_assignments(assignments)
    _build_reviews(reviews)
    _build_comments(comments)
    _build_activity(activity)
    _build_aliases(aliases)
    _build_removed(removed)
    _build_pre_decisions(pre_decisions)
    _build_status_log(log)
    _build_historical_interventions(historical)
    _build_dashboard(dashboard)

    # Reorder: Dashboard first, then data + portal-managed tabs, Lookups last.
    desired = [
        "Dashboard",
        "Companies",
        "Contacts",
        "Intervention Assignments",
        "Reviews",
        "Company Comments",
        "Activity Log",
        "Interview Aliases",
        "Removed Companies",
        "Pre-decision Recommendations",
        "Status Log",
        "Historical Interventions",
        "Lookups",
    ]
    for idx, name in enumerate(desired):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    path = save_workbook(wb, FILENAME)
    return str(path)


if __name__ == "__main__":
    print(build())
