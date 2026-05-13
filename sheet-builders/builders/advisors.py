"""E3 - Non-Technical Advisors workbook.

Owns the full Advisor pipeline: intake, Stage 1 + Stage 2 scoring, triage
(via tracker columns appended to the Advisors row, no cross-tab join),
follow-ups, activity log, and comments thread. Mentors stay as a side tab.

Tab order: Dashboard, Advisors, FollowUps, ActivityLog, Comments, Mentors,
Lookups.

The Advisors tab schema is the union of:
  - Original form-response columns from the legacy "Form Responses 1" sheet
    (preserved verbatim so the migrator can copy data 1:1).
  - Pipeline tracker columns (pipeline_status, assignee, dates, notes) so
    triage state lives on the same row as the advisor — the kanban and the
    table read from a single tab.
  - Assignment FK columns (company_id, intervention_type, status, notes) for
    when an advisor gets matched to an Elevate company.
  - Computed scoring columns (stage1_score, stage1_pass, stage2_category,
    stage2_score) that the portal writes back on every change so the
    Dashboard tab can sort/filter on the live tier without re-running the
    engine in a sheet formula.
  - updated_at, updated_by audit pair.
"""

from gsg_sheets import (
    BRAND,
    add_bar_chart,
    add_dropdown,
    add_kpi_row,
    add_list_block,
    add_lookups_tab,
    add_pie_chart,
    add_progress_bar,
    add_section_header,
    band_rows,
    conditional_formatting_date,
    conditional_formatting_status,
    freeze_header,
    new_workbook,
    save_workbook,
    set_column_widths,
    setup_dashboard_tab,
    tab_color,
    write_header,
)
from gsg_sheets.styles import apply_style, editable_style
from gsg_sheets.taxonomies import ASSIGNMENT_STATUS, INTERVENTION_TYPES
from openpyxl.chart import Reference


FILENAME = "E3 - Non-Technical Advisors.xlsx"

PIPELINE_STATUSES = [
    "New",
    "Acknowledged",
    "Allocated",
    "Intro Scheduled",
    "Intro Done",
    "Assessment",
    "Approved",
    "Matched",
    "On Hold",
    "Rejected",
    # Archived = pre-Cohort 3 historicals or anything pulled out of the active
    # triage funnel. Hidden from the kanban by default.
    "Archived",
]

FOLLOWUP_TYPES = ["Email", "Call", "Meeting", "Other"]
FOLLOWUP_STATUSES = ["Open", "Done", "Snoozed"]
COMMENT_VISIBILITY = ["Team", "Admins"]
ACTIVITY_ACTIONS = ["status_change", "tracker_edit", "comment", "followup", "score_recompute"]
SCORE_TIERS = ["Pass", "Fail"]
CATEGORIES = ["CEO", "CTO", "COO", "Marketing", "AI", "Unqualified"]


# --- Advisors tab ---

ADVISORS_HEADERS = [
    "advisor_id",
    "timestamp",
    "full_name",
    "gender",
    "country",
    "email",
    "whatsapp",
    "linkedin",
    "tech_rating",
    "eco_rating",
    "exp_areas",
    "exp_detail",
    "c_level",
    "c_level_detail",
    "position",
    "employer",
    "years",
    "non_tech_subjects",
    "gsg_past",
    "paid_or_vol",
    "hourly_rate",
    "cv_link",
    "notes",
    "heard_from",
    "opportunities",
    "support_in",
    "support_via",
    "tech_specs",
    "newsletter",
    # Tracker columns appended on the same row.
    "pipeline_status",
    "assignee_email",
    "received_ack",
    "intro_scheduled_date",
    "intro_done_date",
    "assessment_date",
    "decision_date",
    "tracker_notes",
    # Assignment to an Elevate company (when matched).
    "assignment_company_id",
    "assignment_intervention_type",
    "assignment_status",
    "assignment_notes",
    # Computed scoring (portal writes back on change).
    "stage1_score",
    "stage1_pass",
    "stage2_category",
    "stage2_score",
    # Audit
    "updated_at",
    "updated_by",
]

# Columns we want the user to be able to type into. Excludes id, timestamp
# (set by the form), the four computed scoring columns, and the audit pair.
ADVISORS_EDITABLE_COLS = [
    *range(3, 30),    # form-response fields full_name..newsletter
    *range(30, 38),   # tracker columns
    *range(38, 42),   # assignment columns
]

# 1-based indexes for columns the conditional formatting and validators reach.
ADVISORS_COL = {
    "advisor_id": 1,
    "country": 5,
    "tech_rating": 9,
    "eco_rating": 10,
    "c_level": 13,
    "years": 17,
    "paid_or_vol": 20,
    "pipeline_status": 30,
    "received_ack": 32,
    "intro_scheduled_date": 33,
    "intro_done_date": 34,
    "assessment_date": 35,
    "decision_date": 36,
    "assignment_intervention_type": 39,
    "assignment_status": 40,
    "stage1_pass": 43,
    "stage2_category": 44,
}

ADVISORS_LAST_ROW = 800


def _build_advisors(ws):
    tab_color(ws, BRAND["red"])
    write_header(ws, ADVISORS_HEADERS)
    freeze_header(ws)

    # advisor_id: ADV-E3-0001 mints when full_name (col C) is entered.
    for row in range(2, ADVISORS_LAST_ROW + 1):
        ws.cell(
            row=row, column=1,
            value=f'=IF(C{row}<>"", "ADV-E3-"&TEXT(ROW()-1, "0000"), "")',
        )

    # Validators
    add_dropdown(ws, "I", '"1,2,3,4,5"', last_row=ADVISORS_LAST_ROW)   # tech_rating
    add_dropdown(ws, "J", '"1,2,3,4,5"', last_row=ADVISORS_LAST_ROW)   # eco_rating
    add_dropdown(ws, "M", '"Yes,No"', last_row=ADVISORS_LAST_ROW)      # c_level
    add_dropdown(ws, "Q", '"less than 5,5-10,more than 10"', last_row=ADVISORS_LAST_ROW)
    add_dropdown(ws, "T", '"Paid,Volunteer,Either"', last_row=ADVISORS_LAST_ROW)
    add_dropdown(ws, "AD", "=pipeline_statuses", last_row=ADVISORS_LAST_ROW)  # pipeline_status (col 30 = AD)
    add_dropdown(ws, "AF", '"Yes,No"', last_row=ADVISORS_LAST_ROW)            # received_ack
    add_dropdown(ws, "AM", "=intervention_types", last_row=ADVISORS_LAST_ROW) # assignment_intervention_type (col 39 = AM)
    add_dropdown(ws, "AN", "=assignment_status", last_row=ADVISORS_LAST_ROW)  # assignment_status

    # Conditional formatting: pipeline_status tones, assignment_status tones,
    # date columns red-when-overdue, stage1_pass green/red.
    conditional_formatting_status(
        ws, "AD",
        good=["Approved", "Matched"],
        warn=["Intro Scheduled", "Intro Done", "Assessment", "Acknowledged", "Allocated"],
        bad=["Rejected"],
        neutral=["New", "On Hold"],
        last_row=ADVISORS_LAST_ROW,
    )
    conditional_formatting_status(
        ws, "AN",
        good=["Completed"],
        warn=["In Progress"],
        bad=["Cancelled"],
        neutral=["Planned"],
        last_row=ADVISORS_LAST_ROW,
    )
    for date_col_letter in ["AG", "AH", "AI", "AJ"]:  # intro/assessment/decision
        conditional_formatting_date(ws, date_col_letter, last_row=ADVISORS_LAST_ROW)

    # Banded rows
    band_rows(ws, last_col_letter="AT", last_row=ADVISORS_LAST_ROW)

    set_column_widths(
        ws,
        {
            1: 14,   # advisor_id
            2: 18,   # timestamp
            3: 26,   # full_name
            4: 10,   # gender
            5: 14,   # country
            6: 26,   # email
            7: 16,   # whatsapp
            8: 30,   # linkedin
            9: 8,    # tech_rating
            10: 8,   # eco_rating
            11: 30,  # exp_areas
            12: 40,  # exp_detail
            13: 8,   # c_level
            14: 30,  # c_level_detail
            15: 26,  # position
            16: 24,  # employer
            17: 14,  # years
            18: 30,  # non_tech_subjects
            19: 12,  # gsg_past
            20: 12,  # paid_or_vol
            21: 12,  # hourly_rate
            22: 30,  # cv_link
            23: 36,  # notes
            24: 16,  # heard_from
            25: 30,  # opportunities
            26: 30,  # support_in
            27: 26,  # support_via
            28: 30,  # tech_specs
            29: 12,  # newsletter
            30: 16,  # pipeline_status
            31: 24,  # assignee_email
            32: 12,  # received_ack
            33: 14,  # intro_scheduled_date
            34: 14,  # intro_done_date
            35: 14,  # assessment_date
            36: 14,  # decision_date
            37: 32,  # tracker_notes
            38: 14,  # assignment_company_id
            39: 22,  # assignment_intervention_type
            40: 14,  # assignment_status
            41: 32,  # assignment_notes
            42: 12,  # stage1_score
            43: 10,  # stage1_pass
            44: 14,  # stage2_category
            45: 12,  # stage2_score
            46: 18,  # updated_at
            47: 20,  # updated_by
        },
    )

    style_e = editable_style()
    for row in range(2, ADVISORS_LAST_ROW + 1):
        for col in ADVISORS_EDITABLE_COLS:
            apply_style(ws.cell(row=row, column=col), style_e)


# --- FollowUps tab ---

FOLLOWUPS_HEADERS = [
    "followup_id",
    "advisor_id",
    "due_date",
    "type",
    "assignee_email",
    "status",
    "notes",
    "created_by",
    "created_at",
    "completed_at",
    "updated_at",
    "updated_by",
]

FOLLOWUPS_LAST_ROW = 1000


def _build_followups(ws):
    tab_color(ws, BRAND["orange"])
    write_header(ws, FOLLOWUPS_HEADERS)
    freeze_header(ws)

    for row in range(2, FOLLOWUPS_LAST_ROW + 1):
        ws.cell(
            row=row, column=1,
            value=f'=IF(B{row}<>"", "FU-E3-"&TEXT(ROW()-1, "0000"), "")',
        )

    add_dropdown(ws, "D", "=followup_types", last_row=FOLLOWUPS_LAST_ROW)
    add_dropdown(ws, "F", "=followup_statuses", last_row=FOLLOWUPS_LAST_ROW)

    conditional_formatting_status(
        ws, "F",
        good=["Done"],
        warn=["Snoozed"],
        neutral=["Open"],
        last_row=FOLLOWUPS_LAST_ROW,
    )
    conditional_formatting_date(ws, "C", last_row=FOLLOWUPS_LAST_ROW)
    band_rows(ws, last_col_letter="L", last_row=FOLLOWUPS_LAST_ROW)

    set_column_widths(
        ws,
        {
            1: 14, 2: 14, 3: 14, 4: 12, 5: 24, 6: 12,
            7: 36, 8: 24, 9: 18, 10: 18, 11: 18, 12: 20,
        },
    )

    style_e = editable_style()
    for row in range(2, FOLLOWUPS_LAST_ROW + 1):
        for col in [2, 3, 4, 5, 6, 7]:
            apply_style(ws.cell(row=row, column=col), style_e)


# --- ActivityLog tab ---

ACTIVITY_HEADERS = [
    "activity_id",
    "timestamp",
    "user_email",
    "advisor_id",
    "action",
    "field",
    "old_value",
    "new_value",
    "details",
]

ACTIVITY_LAST_ROW = 4000


def _build_activity(ws):
    tab_color(ws, BRAND["navy"])
    write_header(ws, ACTIVITY_HEADERS)
    freeze_header(ws)

    for row in range(2, ACTIVITY_LAST_ROW + 1):
        ws.cell(
            row=row, column=1,
            value=f'=IF(B{row}<>"", "ACT-"&TEXT(ROW()-1, "00000"), "")',
        )

    add_dropdown(ws, "E", "=activity_actions", last_row=ACTIVITY_LAST_ROW)
    band_rows(ws, last_col_letter="I", last_row=ACTIVITY_LAST_ROW)
    set_column_widths(
        ws,
        {1: 14, 2: 20, 3: 26, 4: 14, 5: 18, 6: 18, 7: 24, 8: 24, 9: 36},
    )


# --- Comments tab ---

COMMENTS_HEADERS = [
    "comment_id",
    "advisor_id",
    "author_email",
    "body",
    "visibility",
    "created_at",
    "updated_at",
    "updated_by",
]

COMMENTS_LAST_ROW = 2000


def _build_comments(ws):
    tab_color(ws, BRAND["teal"])
    write_header(ws, COMMENTS_HEADERS)
    freeze_header(ws)

    for row in range(2, COMMENTS_LAST_ROW + 1):
        ws.cell(
            row=row, column=1,
            value=f'=IF(B{row}<>"", "CMT-"&TEXT(ROW()-1, "00000"), "")',
        )

    add_dropdown(ws, "E", "=comment_visibility", last_row=COMMENTS_LAST_ROW)
    band_rows(ws, last_col_letter="H", last_row=COMMENTS_LAST_ROW)
    set_column_widths(
        ws,
        {1: 14, 2: 14, 3: 26, 4: 60, 5: 12, 6: 18, 7: 18, 8: 20},
    )

    style_e = editable_style()
    for row in range(2, COMMENTS_LAST_ROW + 1):
        for col in [2, 3, 4, 5]:
            apply_style(ws.cell(row=row, column=col), style_e)


# --- Mentors tab (preserved from legacy) ---

MENTORS_HEADERS = [
    "mentor_id",
    "full_name",
    "email",
    "country",
    "expertise",
    "active",
    "notes",
    "updated_at",
    "updated_by",
]


def _build_mentors(ws):
    tab_color(ws, BRAND["orange"])
    write_header(ws, MENTORS_HEADERS)
    freeze_header(ws)

    for row in range(2, 201):
        ws.cell(
            row=row, column=1,
            value=f'=IF(B{row}<>"", "MNT-"&TEXT(ROW()-1, "0000"), "")',
        )

    add_dropdown(ws, "F", '"Yes,No"', last_row=200)
    band_rows(ws, last_col_letter="I", last_row=200)
    set_column_widths(
        ws,
        {1: 14, 2: 26, 3: 26, 4: 16, 5: 36, 6: 10, 7: 30, 8: 18, 9: 20},
    )

    style_e = editable_style()
    for row in range(2, 201):
        for col in [2, 3, 4, 5, 6, 7]:
            apply_style(ws.cell(row=row, column=col), style_e)


# --- Dashboard tab (mirrors pages/advisors/AdvisorDashboard) ---

# Cell-range references on the Advisors tab. Helps keep formulas readable.
A_COL_NAME = "C"
A_COL_COUNTRY = "E"
A_COL_PIPELINE = "AD"
A_COL_PASS = "AQ"      # stage1_pass — col 43 = AQ
A_COL_CAT = "AR"       # stage2_category — col 44 = AR
A_COL_S1_SCORE = "AP"  # stage1_score — col 42 = AP
A_RANGE = f"Advisors!$A$2:$AT${ADVISORS_LAST_ROW}"
A_NAME_RANGE = f"Advisors!${A_COL_NAME}$2:${A_COL_NAME}${ADVISORS_LAST_ROW}"
A_COUNTRY_RANGE = f"Advisors!${A_COL_COUNTRY}$2:${A_COL_COUNTRY}${ADVISORS_LAST_ROW}"
A_PIPELINE_RANGE = f"Advisors!${A_COL_PIPELINE}$2:${A_COL_PIPELINE}${ADVISORS_LAST_ROW}"
A_PASS_RANGE = f"Advisors!${A_COL_PASS}$2:${A_COL_PASS}${ADVISORS_LAST_ROW}"
A_CAT_RANGE = f"Advisors!${A_COL_CAT}$2:${A_COL_CAT}${ADVISORS_LAST_ROW}"
A_S1_SCORE_RANGE = f"Advisors!${A_COL_S1_SCORE}$2:${A_COL_S1_SCORE}${ADVISORS_LAST_ROW}"


def _build_dashboard(ws):
    setup_dashboard_tab(ws, brand_tab_color=BRAND["navy"])

    # ---- Header ----
    ws.cell(row=1, column=1, value="Advisor Pipeline Dashboard")
    title = ws.cell(row=1, column=1)
    from openpyxl.styles import Font, Alignment
    title.font = Font(name="Source Sans Pro", bold=True, size=18, color=BRAND["navy"])
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    ws.cell(row=2, column=1, value="Live mirror of the Advisor module in the Elevate Portal.")
    sub = ws.cell(row=2, column=1)
    sub.font = Font(name="Source Sans Pro", size=10, color=BRAND["muted_text"])
    sub.alignment = Alignment(horizontal="left", indent=1)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)

    # ---- Section: Top KPIs ----
    row = 4
    row = add_section_header(ws, row, "Top metrics")
    row = add_kpi_row(ws, row, [
        {
            "label": "Advisors",
            "formula": f'=COUNTA({A_NAME_RANGE})',
            "tone": "navy",
        },
        {
            "label": "Stage 1 pass",
            "formula": f'=COUNTIF({A_PASS_RANGE},"TRUE")+COUNTIF({A_PASS_RANGE},TRUE)',
            "tone": "green",
        },
        {
            "label": "Matched",
            "formula": f'=COUNTIF({A_PIPELINE_RANGE},"Matched")',
            "tone": "teal",
        },
        {
            "label": "Avg Stage 1",
            "formula": f'=IFERROR(ROUND(AVERAGE({A_S1_SCORE_RANGE}),0),0)',
            "tone": "amber",
            "value_fmt": "0",
        },
    ])

    # ---- Section: Pipeline funnel (counts per status) ----
    row = add_section_header(ws, row, "Pipeline funnel")
    funnel_start_row = row
    for idx, status in enumerate(PIPELINE_STATUSES):
        ws.cell(row=row, column=1, value=status)
        cell = ws.cell(row=row, column=1)
        cell.font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
        # count
        ws.cell(row=row, column=2, value=f'=COUNTIF({A_PIPELINE_RANGE},"{status}")')
        ws.cell(row=row, column=2).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        # bar
        bar_cell = ws.cell(
            row=row,
            column=3,
            value=(
                f'=IFERROR(REPT("█",MIN(40,ROUND(COUNTIF({A_PIPELINE_RANGE},"{status}")/'
                f'MAX(1,COUNTA({A_PIPELINE_RANGE}))*40,0))),"")'
            ),
        )
        bar_cell.font = Font(name="Menlo", size=11, color=BRAND["red"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        row += 1
    row += 1

    # ---- Section: By category (Stage 2) ----
    row = add_section_header(ws, row, "Category fit (Stage 2)")
    cat_table_start = row
    for cat in CATEGORIES:
        ws.cell(row=row, column=1, value=cat)
        ws.cell(row=row, column=1).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
        ws.cell(row=row, column=2, value=f'=COUNTIF({A_CAT_RANGE},"{cat}")')
        ws.cell(row=row, column=2).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        bar_cell = ws.cell(
            row=row, column=3,
            value=(
                f'=IFERROR(REPT("█",MIN(40,ROUND(COUNTIF({A_CAT_RANGE},"{cat}")/'
                f'MAX(1,COUNTA({A_CAT_RANGE}))*40,0))),"")'
            ),
        )
        bar_cell.font = Font(name="Menlo", size=11, color=BRAND["teal"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        row += 1
    row += 1

    # ---- Section: Follow-ups due this week ----
    row = add_section_header(ws, row, "Follow-ups (this week)")
    ws.cell(row=row, column=1, value="Open & due in <=7 days")
    ws.cell(row=row, column=1).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
    ws.cell(
        row=row, column=2,
        value=(
            f'=COUNTIFS(FollowUps!$F$2:$F${FOLLOWUPS_LAST_ROW},"Open",'
            f'FollowUps!$C$2:$C${FOLLOWUPS_LAST_ROW},">="&TODAY(),'
            f'FollowUps!$C$2:$C${FOLLOWUPS_LAST_ROW},"<="&(TODAY()+7))'
        ),
    )
    row += 1
    ws.cell(row=row, column=1, value="Overdue")
    ws.cell(row=row, column=1).font = Font(name="Source Sans Pro", size=10, color=BRAND["red"], bold=True)
    ws.cell(
        row=row, column=2,
        value=(
            f'=COUNTIFS(FollowUps!$F$2:$F${FOLLOWUPS_LAST_ROW},"Open",'
            f'FollowUps!$C$2:$C${FOLLOWUPS_LAST_ROW},"<"&TODAY())'
        ),
    )
    row += 2

    # ---- Charts ----
    # Pipeline funnel pie chart anchored to the right of the funnel section.
    cat_ref = Reference(ws, min_col=1, min_row=funnel_start_row, max_col=1,
                        max_row=funnel_start_row + len(PIPELINE_STATUSES) - 1)
    val_ref = Reference(ws, min_col=2, min_row=funnel_start_row, max_col=2,
                        max_row=funnel_start_row + len(PIPELINE_STATUSES) - 1)
    add_pie_chart(
        ws,
        title="Pipeline distribution",
        cat_ref=cat_ref,
        val_ref=val_ref,
        anchor="F" + str(funnel_start_row),
    )

    # Category bar chart
    cat2_ref = Reference(ws, min_col=1, min_row=cat_table_start, max_col=1,
                         max_row=cat_table_start + len(CATEGORIES) - 1)
    val2_ref = Reference(ws, min_col=2, min_row=cat_table_start, max_col=2,
                         max_row=cat_table_start + len(CATEGORIES) - 1)
    add_bar_chart(
        ws,
        title="Category fit",
        cat_ref=cat2_ref,
        val_ref=val2_ref,
        anchor="F" + str(cat_table_start),
    )


def build() -> str:
    wb = new_workbook()

    add_lookups_tab(
        wb,
        {
            "pipeline_statuses": PIPELINE_STATUSES,
            "followup_types": FOLLOWUP_TYPES,
            "followup_statuses": FOLLOWUP_STATUSES,
            "comment_visibility": COMMENT_VISIBILITY,
            "activity_actions": ACTIVITY_ACTIONS,
            "intervention_types": INTERVENTION_TYPES,
            "assignment_status": ASSIGNMENT_STATUS,
            "categories": CATEGORIES,
            "score_tiers": SCORE_TIERS,
        },
    )

    dashboard = wb.create_sheet("Dashboard")
    advisors = wb.create_sheet("Advisors")
    followups = wb.create_sheet("FollowUps")
    activity = wb.create_sheet("ActivityLog")
    comments = wb.create_sheet("Comments")
    mentors = wb.create_sheet("Mentors")

    _build_advisors(advisors)
    _build_followups(followups)
    _build_activity(activity)
    _build_comments(comments)
    _build_mentors(mentors)
    _build_dashboard(dashboard)

    # Tab order: Dashboard first.
    desired = ["Dashboard", "Advisors", "FollowUps", "ActivityLog", "Comments", "Mentors", "Lookups"]
    for idx, name in enumerate(desired):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    return str(save_workbook(wb, FILENAME))


if __name__ == "__main__":
    print(build())
