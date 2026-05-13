"""E3 — Procurement Plan workbook.

Mirrors the structure of GSG_WB_Procurement_Plan_2025-2026.xlsx but rebuilt
for Cohort 3 with intervention mapping, fund code, and computed thresholds
and deadlines.

Tabs: Q1 2026, Q2 2026, Q3 2026, Q4 2026, Annual Summary, Lookups.
"""

from gsg_sheets import (
    BRAND,
    add_dropdown,
    add_kpi_row,
    add_lookups_tab,
    add_section_header,
    band_rows,
    conditional_formatting_date,
    conditional_formatting_status,
    conditional_formatting_threshold,
    freeze_header,
    new_workbook,
    save_workbook,
    set_column_widths,
    setup_dashboard_tab,
    tab_color,
    write_header,
)
from gsg_sheets.styles import apply_style, cell_style, editable_style, formula_style
from gsg_sheets.taxonomies import (
    FUND_CODES,
    INTERVENTION_TYPES,
    PROCUREMENT_STATUS,
)
from openpyxl.styles import Alignment, Font


FILENAME = "E3 - Procurement Plan.xlsx"

QUARTERS = ["Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026"]

QUARTER_HEADERS = [
    "pr_id",
    "activity",
    "intervention_type",
    "company_id",
    "office_code",
    "gl_account",
    "fund_code",
    "lin_code",
    "item_description",
    "unit",
    "qty",
    "unit_cost_usd",
    "total_cost_usd",
    "threshold_class",
    "sla_working_days",
    "target_award_date",
    "pr_submit_date",
    "pr_deadline",
    "local_international",
    "requester_email",
    "status",
    "procurement_contact",
    "notes",
    "updated_at",
    "updated_by",
]

# Column indexes (1-based) for the quarter tab.
COL = {header: i + 1 for i, header in enumerate(QUARTER_HEADERS)}


def _fmt_formulas(row: int, quarter_tag: str) -> dict:
    """Return {col_index: formula_string} for the formula-driven columns."""
    qty = f"K{row}"
    unit = f"L{row}"
    total = f"M{row}"
    cls = f"N{row}"
    sla = f"O{row}"
    target = f"P{row}"
    return {
        COL["pr_id"]: f'=IF(B{row}<>"", "PR-{quarter_tag}-"&TEXT(ROW()-1,"000"), "")',
        COL["total_cost_usd"]: f'=IF(AND({qty}<>"",{unit}<>""), {qty}*{unit}, "")',
        COL["threshold_class"]: (
            f'=IF({total}="","",'
            f'IF({total}<5000,"Micro",'
            f'IF({total}<25000,"Small",'
            f'IF({total}<150000,"Standard","High Value"))))'
        ),
        COL["sla_working_days"]: (
            f'=IFS({cls}="Micro",5,{cls}="Small",10,'
            f'{cls}="Standard",25,{cls}="High Value",35,TRUE,"")'
        ),
        COL["pr_deadline"]: (
            f'=IF(AND({target}<>"",{sla}<>""),'
            f'WORKDAY({target},-{sla}),"")'
        ),
    }


def _build_quarter(ws, quarter_index: int):
    quarter_tag = f"Q{quarter_index}"
    tab_color(ws, [BRAND["red"], BRAND["orange"], BRAND["teal"], BRAND["navy"]][quarter_index - 1])
    write_header(ws, QUARTER_HEADERS)
    freeze_header(ws)

    style_e = editable_style()
    style_f = formula_style()

    editable_cols = {
        COL["activity"],
        COL["intervention_type"],
        COL["company_id"],
        COL["office_code"],
        COL["gl_account"],
        COL["fund_code"],
        COL["lin_code"],
        COL["item_description"],
        COL["unit"],
        COL["qty"],
        COL["unit_cost_usd"],
        COL["target_award_date"],
        COL["pr_submit_date"],
        COL["local_international"],
        COL["requester_email"],
        COL["status"],
        COL["procurement_contact"],
        COL["notes"],
    }
    formula_cols = {
        COL["pr_id"],
        COL["total_cost_usd"],
        COL["threshold_class"],
        COL["sla_working_days"],
        COL["pr_deadline"],
    }

    for row in range(2, 501):
        formulas = _fmt_formulas(row, quarter_tag)
        for col_idx, formula in formulas.items():
            cell = ws.cell(row=row, column=col_idx, value=formula)
            apply_style(cell, style_f)
        for col_idx in editable_cols:
            cell = ws.cell(row=row, column=col_idx)
            apply_style(cell, style_e)
        # Default values
        ws.cell(row=row, column=COL["procurement_contact"]).value = "Donia Shadeed"

    add_dropdown(ws, "C", "=intervention_types")
    add_dropdown(ws, "G", "=fund_codes")
    add_dropdown(ws, "S", '"Local,International"')
    add_dropdown(ws, "U", "=procurement_status")
    add_dropdown(ws, "J", '"Unit,Hour,Day,Month,License,Seat,Trip,Lump Sum"')

    # Phase K polish: status tones, threshold tones, deadline date tone, banded rows.
    conditional_formatting_status(
        ws, "U",
        good=["Awarded", "Delivered"],
        warn=["Submitted", "Under Review"],
        bad=["Cancelled"],
        neutral=["Draft"],
        last_row=500,
    )
    conditional_formatting_threshold(ws, "N", last_row=500)
    conditional_formatting_date(ws, "R", last_row=500)
    band_rows(ws, last_col_letter="Y", last_row=500)

    set_column_widths(
        ws,
        {
            1: 16,
            2: 30,
            3: 22,
            4: 12,
            5: 14,
            6: 14,
            7: 18,
            8: 16,
            9: 36,
            10: 12,
            11: 8,
            12: 14,
            13: 16,
            14: 16,
            15: 12,
            16: 16,
            17: 16,
            18: 16,
            19: 18,
            20: 26,
            21: 16,
            22: 22,
            23: 40,
            24: 18,
            25: 20,
        },
    )


def _build_summary(ws):
    tab_color(ws, BRAND["navy"])
    write_header(
        ws,
        [
            "Fund code",
            "Intervention",
            "Q1 2026",
            "Q2 2026",
            "Q3 2026",
            "Q4 2026",
            "Total 2026",
        ],
    )
    freeze_header(ws)
    style_f = formula_style()
    style_c = cell_style()

    row = 2
    for fund in FUND_CODES:
        for intervention in INTERVENTION_TYPES:
            ws.cell(row=row, column=1, value=fund)
            ws.cell(row=row, column=2, value=intervention)
            for q_idx, quarter in enumerate(QUARTERS, start=3):
                formula = (
                    f"=SUMIFS('{quarter}'!M:M,'{quarter}'!G:G,A{row},"
                    f"'{quarter}'!C:C,B{row})"
                )
                cell = ws.cell(row=row, column=q_idx, value=formula)
                apply_style(cell, style_f)
            total = f"=SUM(C{row}:F{row})"
            cell = ws.cell(row=row, column=7, value=total)
            apply_style(cell, style_f)
            for col in (1, 2):
                apply_style(ws.cell(row=row, column=col), style_c)
            row += 1

    # Grand total row
    grand_row = row + 1
    ws.cell(row=grand_row, column=1, value="Grand total")
    for q_idx in range(3, 8):
        col_letter = chr(ord("A") + q_idx - 1)
        formula = f"=SUM({col_letter}2:{col_letter}{row - 1})"
        cell = ws.cell(row=grand_row, column=q_idx, value=formula)
        apply_style(cell, style_f)

    set_column_widths(ws, {1: 18, 2: 22, 3: 14, 4: 14, 5: 14, 6: 14, 7: 16})


def _build_dashboard(ws):
    """Mirrors pages/procurement/ProcurementPage.tsx in-sheet."""
    setup_dashboard_tab(ws, brand_tab_color=BRAND["red"])

    title = ws.cell(row=1, column=1, value="Procurement Dashboard")
    title.font = Font(name="Source Sans Pro", bold=True, size=18, color=BRAND["navy"])
    title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 36
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)

    sub = ws.cell(row=2, column=1, value="Live mirror of the Procurement module in the Elevate Portal.")
    sub.font = Font(name="Source Sans Pro", size=10, color=BRAND["muted_text"])
    sub.alignment = Alignment(horizontal="left", indent=1)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)

    row = 4
    row = add_section_header(ws, row, "Top metrics")
    # PRs across all 4 quarters; deadlines this week (across all 4)
    pr_count_formula = "+".join(f'COUNTA(\'Q{i} 2026\'!B2:B500)' for i in range(1, 5))
    total_value_formula = "+".join(f'SUM(\'Q{i} 2026\'!M2:M500)' for i in range(1, 5))
    awarded_formula = "+".join(
        f'SUMIFS(\'Q{i} 2026\'!M2:M500,\'Q{i} 2026\'!U2:U500,"Awarded")' for i in range(1, 5)
    )
    overdue_formula = "+".join(
        f'COUNTIFS(\'Q{i} 2026\'!R2:R500,"<"&TODAY(),\'Q{i} 2026\'!U2:U500,"Draft")' for i in range(1, 5)
    )
    row = add_kpi_row(ws, row, [
        {"label": "PRs", "formula": f"={pr_count_formula}", "tone": "navy"},
        {"label": "Total value (USD)", "formula": f"={total_value_formula}", "tone": "teal", "value_fmt": "$#,##0"},
        {"label": "Awarded (USD)", "formula": f"={awarded_formula}", "tone": "green", "value_fmt": "$#,##0"},
        {"label": "Past-due drafts", "formula": f"={overdue_formula}", "tone": "red"},
    ])

    # Per quarter breakdown
    row = add_section_header(ws, row, "By quarter")
    for i in range(1, 5):
        q = f"Q{i} 2026"
        ws.cell(row=row, column=1, value=q).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
        ws.cell(row=row, column=2, value=f"=COUNTA('{q}'!B2:B500)").font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        ws.cell(row=row, column=3, value=f"=SUM('{q}'!M2:M500)").font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        ws.cell(row=row, column=3).number_format = "$#,##0"
        bar = ws.cell(
            row=row, column=4,
            value=f'=IFERROR(REPT("█",MIN(40,ROUND(SUM(\'{q}\'!M2:M500)/MAX(1,{total_value_formula})*40,0))),"")',
        )
        bar.font = Font(name="Menlo", size=11, color=BRAND["red"])
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=12)
        row += 1
    row += 1

    # By status (across all quarters)
    row = add_section_header(ws, row, "By status")
    for status in PROCUREMENT_STATUS:
        per_quarter = "+".join(
            f'COUNTIF(\'Q{i} 2026\'!U2:U500,"{status}")' for i in range(1, 5)
        )
        ws.cell(row=row, column=1, value=status).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
        ws.cell(row=row, column=2, value=f"={per_quarter}").font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        bar = ws.cell(
            row=row, column=3,
            value=f'=IFERROR(REPT("█",MIN(40,ROUND(({per_quarter})/MAX(1,{pr_count_formula})*40,0))),"")',
        )
        bar.font = Font(name="Menlo", size=11, color=BRAND["teal"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        row += 1
    row += 1

    # By threshold class
    row = add_section_header(ws, row, "By threshold class")
    for cls in ["Micro", "Small", "Standard", "High Value"]:
        per_quarter = "+".join(
            f'COUNTIF(\'Q{i} 2026\'!N2:N500,"{cls}")' for i in range(1, 5)
        )
        ws.cell(row=row, column=1, value=cls).font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"], bold=True)
        ws.cell(row=row, column=2, value=f"={per_quarter}").font = Font(name="Source Sans Pro", size=10, color=BRAND["navy"])
        bar = ws.cell(
            row=row, column=3,
            value=f'=IFERROR(REPT("█",MIN(40,ROUND(({per_quarter})/MAX(1,{pr_count_formula})*40,0))),"")',
        )
        bar.font = Font(name="Menlo", size=11, color=BRAND["orange"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        row += 1


def build() -> str:
    wb = new_workbook()

    add_lookups_tab(
        wb,
        {
            "intervention_types": INTERVENTION_TYPES,
            "fund_codes": FUND_CODES,
            "procurement_status": PROCUREMENT_STATUS,
        },
    )

    dashboard = wb.create_sheet("Dashboard")
    for i in range(1, 5):
        ws = wb.create_sheet(f"Q{i} 2026")
        _build_quarter(ws, i)

    summary = wb.create_sheet("Annual Summary")
    _build_summary(summary)
    _build_dashboard(dashboard)

    # Order: Dashboard, Q1..Q4, Annual Summary, Lookups
    order = ["Dashboard", "Q1 2026", "Q2 2026", "Q3 2026", "Q4 2026", "Annual Summary", "Lookups"]
    for idx, name in enumerate(order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    path = save_workbook(wb, FILENAME)
    return str(path)


if __name__ == "__main__":
    print(build())
