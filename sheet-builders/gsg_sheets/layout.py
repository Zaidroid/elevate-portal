"""Layout helpers: headers, rows, dropdowns, column sizing, named ranges."""

from typing import Iterable, Sequence

from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from .brand import BRAND, HEADER_HEIGHT, ROW_HEIGHT
from .styles import apply_row_style, apply_style, cell_style, editable_style, header_style


def write_header(ws, headers: Sequence[str], row: int = 1) -> None:
    style = header_style()
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=title)
        apply_style(cell, style)
    ws.row_dimensions[row].height = HEADER_HEIGHT


def write_rows(
    ws,
    rows: Iterable[Sequence],
    start_row: int,
    editable_cols: Sequence[int] = (),
    formula_cols: dict = None,
) -> int:
    """Write data rows. editable_cols are 1-based indexes that get the blue fill.
    formula_cols is {col_index: lambda row_index: "=FORMULA(...)"}.
    Returns the last row written.
    """
    formula_cols = formula_cols or {}
    style_cell = cell_style()
    style_editable = editable_style()
    r = start_row
    for values in rows:
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col_idx, value=value)
            apply_style(cell, style_editable if col_idx in editable_cols else style_cell)
        for col_idx, formula_fn in formula_cols.items():
            cell = ws.cell(row=r, column=col_idx, value=formula_fn(r))
            apply_style(cell, style_cell)
        ws.row_dimensions[r].height = ROW_HEIGHT
        r += 1
    return r - 1


def add_dropdown(ws, col_letter: str, source_range: str, first_row: int = 2, last_row: int = 1000) -> None:
    """source_range example: 'Lookups!$A$2:$A$20' or a literal '"One,Two,Three"'."""
    dv = DataValidation(type="list", formula1=source_range, allow_blank=True, showDropDown=False)
    dv.error = "Value not in allowed list"
    dv.errorTitle = "Invalid selection"
    dv.prompt = "Select from list"
    dv.promptTitle = "Allowed values"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


def add_table(ws, name: str, first_row: int, last_row: int, first_col: int, last_col: int) -> None:
    ref = f"{get_column_letter(first_col)}{first_row}:{get_column_letter(last_col)}{last_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def set_column_widths(ws, widths: dict) -> None:
    """widths: {col_index_1based: width_in_chars}"""
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def autosize_columns(ws, headers: Sequence[str], min_width: int = 14, max_width: int = 40) -> None:
    for col_idx, title in enumerate(headers, start=1):
        w = max(min_width, min(max_width, len(title) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def add_named_range(wb, name: str, sheet_name: str, cell_range: str) -> None:
    ref = f"'{sheet_name}'!{cell_range}"
    dn = DefinedName(name=name, attr_text=ref)
    wb.defined_names[name] = dn


def tab_color(ws, hex_color: str) -> None:
    ws.sheet_properties.tabColor = hex_color
