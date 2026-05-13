"""Workbook-level helpers: new workbook, save, lookups tab."""

from pathlib import Path
from typing import Mapping, Sequence

from openpyxl import Workbook

from .brand import BRAND
from .layout import add_named_range, set_column_widths, tab_color, write_header, write_rows


OUTPUT_DIR = Path(__file__).resolve().parents[1] / "out"


def new_workbook() -> Workbook:
    wb = Workbook()
    # Remove the default sheet; callers add their own.
    default = wb.active
    wb.remove(default)
    return wb


def save_workbook(wb: Workbook, filename: str) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    path = OUTPUT_DIR / filename
    wb.save(path)
    return path


def add_lookups_tab(wb: Workbook, lookups: Mapping[str, Sequence[str]]) -> None:
    """Create a Lookups tab with one column per named list.

    Also registers a named range per list so other tabs can use
    `=Lookups!name` for data validation sources.
    """
    ws = wb.create_sheet("Lookups")
    tab_color(ws, BRAND["navy"])
    write_header(ws, list(lookups.keys()))
    max_len = max(len(values) for values in lookups.values()) if lookups else 0
    # Write each column
    for col_idx, (_name, values) in enumerate(lookups.items(), start=1):
        for row_idx, value in enumerate(values, start=2):
            ws.cell(row=row_idx, column=col_idx, value=value)
    # Column widths
    set_column_widths(ws, {i: 24 for i in range(1, len(lookups) + 1)})
    # Register named ranges
    from openpyxl.utils import get_column_letter

    for col_idx, (name, values) in enumerate(lookups.items(), start=1):
        col_letter = get_column_letter(col_idx)
        last_row = max(2, 1 + len(values))
        add_named_range(
            wb,
            name=_sanitize_name(name),
            sheet_name="Lookups",
            cell_range=f"${col_letter}$2:${col_letter}${last_row}",
        )
    ws.sheet_state = "visible"


def _sanitize_name(name: str) -> str:
    return name.replace(" ", "_").replace("-", "_").replace("/", "_")
