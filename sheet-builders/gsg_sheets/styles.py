"""Reusable openpyxl style objects honoring the GSG brand."""

from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .brand import BRAND, FONT_NAME, HEADER_HEIGHT, ROW_HEIGHT


def _side(color: str = None) -> Side:
    return Side(style="thin", color=color or BRAND["border_gray"])


def _border() -> Border:
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def header_style() -> dict:
    return {
        "font": Font(name=FONT_NAME, bold=True, color=BRAND["header_text"], size=11),
        "fill": PatternFill("solid", fgColor=BRAND["navy"]),
        "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "border": _border(),
    }


def subheader_style() -> dict:
    return {
        "font": Font(name=FONT_NAME, bold=True, color=BRAND["navy"], size=10),
        "fill": PatternFill("solid", fgColor=BRAND["subheader_fill"]),
        "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "border": _border(),
    }


def editable_style() -> dict:
    return {
        "font": Font(name=FONT_NAME, color=BRAND["navy"], size=10),
        "fill": PatternFill("solid", fgColor=BRAND["editable_blue"]),
        "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "border": _border(),
    }


def formula_style() -> dict:
    return {
        "font": Font(name=FONT_NAME, color=BRAND["muted_text"], size=10, italic=True),
        "fill": PatternFill("solid", fgColor=BRAND["white"]),
        "alignment": Alignment(horizontal="left", vertical="center"),
        "border": _border(),
    }


def cell_style() -> dict:
    return {
        "font": Font(name=FONT_NAME, color=BRAND["navy"], size=10),
        "fill": PatternFill("solid", fgColor=BRAND["white"]),
        "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "border": _border(),
    }


def link_style() -> dict:
    return {
        "font": Font(name=FONT_NAME, color=BRAND["teal"], size=10, underline="single"),
        "fill": PatternFill("solid", fgColor=BRAND["white"]),
        "alignment": Alignment(horizontal="left", vertical="center"),
        "border": _border(),
    }


def note_style() -> dict:
    return {
        "font": Font(name=FONT_NAME, color=BRAND["muted_text"], size=9, italic=True),
        "fill": PatternFill("solid", fgColor=BRAND["white"]),
        "alignment": Alignment(horizontal="left", vertical="top", wrap_text=True),
        "border": _border(),
    }


def apply_style(cell, style: dict) -> None:
    for attr, value in style.items():
        setattr(cell, attr, value)


def apply_row_style(ws, row: int, cols: int, style: dict, height: float = None) -> None:
    for col in range(1, cols + 1):
        apply_style(ws.cell(row=row, column=col), style)
    if height is not None:
        ws.row_dimensions[row].height = height


def freeze_header(ws, header_row: int = 1) -> None:
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.row_dimensions[header_row].height = HEADER_HEIGHT


# ----- Conditional formatting helpers (Phase K polish) -----

# Brand-tinted fills keyed by status semantic.
STATUS_FILLS = {
    "good":   PatternFill("solid", fgColor="E7F5EC"),  # green-tinted
    "warn":   PatternFill("solid", fgColor="FDF3DA"),  # amber-tinted
    "bad":    PatternFill("solid", fgColor="FCEDE7"),  # red-tinted
    "neutral":PatternFill("solid", fgColor="F1F4F6"),  # slate
    "info":   PatternFill("solid", fgColor="E6F3F7"),  # teal-tinted
}

STATUS_FONTS = {
    "good":   Font(name=FONT_NAME, color="2E7D4F", bold=True, size=10),
    "warn":   Font(name=FONT_NAME, color="8C6A12", bold=True, size=10),
    "bad":    Font(name=FONT_NAME, color=BRAND["red"], bold=True, size=10),
    "neutral":Font(name=FONT_NAME, color=BRAND["muted_text"], size=10),
    "info":   Font(name=FONT_NAME, color=BRAND["teal"], bold=True, size=10),
}


def _range_for(col_letter: str, first_row: int = 2, last_row: int = 1000) -> str:
    return f"{col_letter}{first_row}:{col_letter}{last_row}"


def conditional_formatting_status(
    ws,
    col_letter: str,
    *,
    good: list[str] | None = None,
    warn: list[str] | None = None,
    bad: list[str] | None = None,
    neutral: list[str] | None = None,
    info: list[str] | None = None,
    first_row: int = 2,
    last_row: int = 1000,
) -> None:
    """Apply tinted fills to status columns. Pass lists of literal status values.

    Example:
        conditional_formatting_status(ws, "M",
            good=["Active", "Graduated", "Paid"],
            warn=["Pending Approval", "Drafted"],
            bad=["Withdrawn", "Rejected"])
    """
    rng = _range_for(col_letter, first_row, last_row)
    bands = {"good": good, "warn": warn, "bad": bad, "neutral": neutral, "info": info}
    for tone, values in bands.items():
        if not values:
            continue
        for v in values:
            rule = CellIsRule(
                operator="equal",
                formula=[f'"{v}"'],
                stopIfTrue=False,
                fill=STATUS_FILLS[tone],
                font=STATUS_FONTS[tone],
            )
            ws.conditional_formatting.add(rng, rule)


def conditional_formatting_date(
    ws,
    col_letter: str,
    *,
    first_row: int = 2,
    last_row: int = 1000,
    today_ref: str = "TODAY()",
) -> None:
    """Tint date cells: red when overdue, amber when due in <=7 days."""
    rng = _range_for(col_letter, first_row, last_row)
    overdue = FormulaRule(
        formula=[f'AND(ISNUMBER({col_letter}{first_row}),{col_letter}{first_row}<{today_ref})'],
        stopIfTrue=False,
        fill=STATUS_FILLS["bad"],
        font=STATUS_FONTS["bad"],
    )
    soon = FormulaRule(
        formula=[
            f'AND(ISNUMBER({col_letter}{first_row}),{col_letter}{first_row}>={today_ref},'
            f'{col_letter}{first_row}<=({today_ref}+7))'
        ],
        stopIfTrue=False,
        fill=STATUS_FILLS["warn"],
        font=STATUS_FONTS["warn"],
    )
    ws.conditional_formatting.add(rng, overdue)
    ws.conditional_formatting.add(rng, soon)


def conditional_formatting_threshold(
    ws,
    col_letter: str,
    *,
    first_row: int = 2,
    last_row: int = 1000,
) -> None:
    """Tint procurement threshold_class column: Micro pale, Small mint, Standard amber,
    High Value red."""
    conditional_formatting_status(
        ws,
        col_letter,
        neutral=["Micro"],
        good=["Small"],
        warn=["Standard"],
        bad=["High Value"],
        first_row=first_row,
        last_row=last_row,
    )


def band_rows(
    ws,
    *,
    first_row: int = 2,
    last_row: int = 1000,
    last_col_letter: str,
    band_color: str | None = None,
) -> None:
    """Zebra fill on even rows. Light enough to keep navy text legible."""
    color = band_color or BRAND["row_zebra"]
    rng = f"A{first_row}:{last_col_letter}{last_row}"
    rule = FormulaRule(
        formula=[f"MOD(ROW(),2)=0"],
        stopIfTrue=False,
        fill=PatternFill("solid", fgColor=color),
    )
    ws.conditional_formatting.add(rng, rule)
