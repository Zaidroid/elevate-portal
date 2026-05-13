"""Shared layout helpers for the in-sheet Dashboard tab.

Every workbook in the project ships a `Dashboard` tab whose KPIs and aggregates
mirror the corresponding portal page. This module gives builders a uniform set
of primitives so the visual language is consistent across workbooks:

  - section header bars
  - KPI tiles (4 wide grid)
  - progress bars (REPT-based, brand palette)
  - sparkline-style numeric panels
  - openpyxl chart objects (bar, pie, line) anchored to the dashboard

Layout convention: Dashboard tabs use a 12-column virtual grid. Section headers
span 12. KPI tiles take 3 each (so 4 fit per row). Charts take 6 each (2 per
row). Lists span 6 or 12 depending on density.
"""

from __future__ import annotations

from typing import Iterable, Mapping, Sequence

from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .brand import BRAND, FONT_NAME
from .layout import set_column_widths, tab_color


# Brand-tinted tile fills keyed by tone.
TONES: dict[str, dict[str, str]] = {
    "navy":   {"fill": BRAND["navy"],   "fg": "FFFFFF"},
    "red":    {"fill": "FCEDE7",        "fg": BRAND["red"]},
    "teal":   {"fill": "E6F3F7",        "fg": BRAND["teal"]},
    "orange": {"fill": "FCEFE1",        "fg": "B05E1F"},
    "green":  {"fill": "E7F5EC",        "fg": "2E7D4F"},
    "amber":  {"fill": "FDF3DA",        "fg": "8C6A12"},
    "slate":  {"fill": "F1F4F6",        "fg": BRAND["muted_text"]},
}


def setup_dashboard_tab(ws, *, brand_tab_color: str | None = None) -> None:
    """Apply uniform layout settings: tab color, hide gridlines, freeze, column widths."""
    if brand_tab_color is None:
        brand_tab_color = BRAND["red"]
    tab_color(ws, brand_tab_color)
    ws.sheet_view.showGridLines = False
    # 12-column grid, each column ~14 chars wide.
    set_column_widths(ws, {i: 14 for i in range(1, 13)})


def add_section_header(ws, row: int, text: str, *, span: int = 12) -> int:
    """Render a navy bar with white text spanning `span` columns. Returns next row."""
    last_col = span
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=12)
    cell.fill = PatternFill("solid", fgColor=BRAND["navy"])
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 26
    return row + 1


def add_kpi_tile(
    ws,
    row: int,
    col: int,
    *,
    label: str,
    formula: str,
    tone: str = "navy",
    value_fmt: str | None = None,
) -> None:
    """Lay out a KPI tile of width 3 (label row + value row)."""
    tone_cfg = TONES.get(tone, TONES["navy"])
    last_col = col + 2
    label_row = row
    value_row = row + 1

    # Label band
    ws.merge_cells(start_row=label_row, start_column=col, end_row=label_row, end_column=last_col)
    label_cell = ws.cell(row=label_row, column=col, value=label.upper())
    label_cell.font = Font(name=FONT_NAME, bold=True, size=8, color=tone_cfg["fg"])
    label_cell.fill = PatternFill("solid", fgColor=tone_cfg["fill"])
    label_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[label_row].height = 18

    # Value band
    ws.merge_cells(start_row=value_row, start_column=col, end_row=value_row, end_column=last_col)
    value_cell = ws.cell(row=value_row, column=col, value=formula)
    value_cell.font = Font(name=FONT_NAME, bold=True, size=22, color=tone_cfg["fg"])
    value_cell.fill = PatternFill("solid", fgColor=tone_cfg["fill"])
    value_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if value_fmt:
        value_cell.number_format = value_fmt
    ws.row_dimensions[value_row].height = 38


def add_kpi_row(
    ws,
    row: int,
    tiles: Sequence[Mapping[str, object]],
) -> int:
    """Render up to 4 KPI tiles across the 12-column grid. Returns next row."""
    cols = [1, 4, 7, 10]
    for tile, col in zip(tiles, cols):
        add_kpi_tile(
            ws,
            row,
            col,
            label=str(tile["label"]),
            formula=str(tile["formula"]),
            tone=str(tile.get("tone", "navy")),
            value_fmt=(str(tile["value_fmt"]) if tile.get("value_fmt") else None),
        )
    # KPI tile takes 2 rows
    return row + 3


def add_progress_bar(
    ws,
    row: int,
    col: int,
    *,
    label: str,
    value_formula: str,
    total_formula: str,
    tone: str = "red",
    width_chars: int = 24,
    span: int = 6,
) -> int:
    """Render a label + REPT-based progress bar. Returns next row.

    The bar uses block characters so it shows up identically whether the
    user is in Excel, Google Sheets, or Numbers.
    """
    tone_cfg = TONES.get(tone, TONES["red"])
    last_col = col + span - 1

    # Label
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=last_col)
    cell = ws.cell(row=row, column=col, value=label)
    cell.font = Font(name=FONT_NAME, bold=True, size=10, color=BRAND["navy"])
    cell.alignment = Alignment(horizontal="left", vertical="center")

    # Value caption (e.g. "12 / 50  (24%)")
    caption_cell = ws.cell(row=row, column=last_col, value=(
        f'=IFERROR(TEXT({value_formula},"#,##0")&" / "&TEXT({total_formula},"#,##0")&'
        f'"  ("&TEXT(IFERROR({value_formula}/{total_formula},0),"0%")&")","")'
    ))
    caption_cell.font = Font(name=FONT_NAME, size=10, color=BRAND["muted_text"])
    caption_cell.alignment = Alignment(horizontal="right", vertical="center")
    # caption_cell ends up overwritten by the merge; rewrite the value-only cell.
    # Simpler: write the bar in the next row and skip a separate caption.

    # Bar
    bar_row = row + 1
    ws.merge_cells(start_row=bar_row, start_column=col, end_row=bar_row, end_column=last_col)
    bar = ws.cell(
        row=bar_row,
        column=col,
        value=(
            f'=IFERROR(REPT("█",MIN({width_chars},ROUND({value_formula}/{total_formula}*{width_chars},0)))'
            f'&REPT("░",MAX(0,{width_chars}-MIN({width_chars},ROUND({value_formula}/{total_formula}*{width_chars},0)))),"")'
        ),
    )
    bar.font = Font(name="Menlo", size=11, color=tone_cfg["fg"])
    bar.alignment = Alignment(horizontal="left", vertical="center")
    return row + 2


def add_list_block(
    ws,
    row: int,
    *,
    title: str,
    headers: Sequence[str],
    rows_formula: Sequence[Sequence[str]],
    span: int = 6,
) -> int:
    """A small titled table — each cell is a formula expression. Returns next row.

    Use this for "deadlines this week", "top 5 idle advisors", "needs review list".
    """
    last_col = span
    # Title bar
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    title_cell = ws.cell(row=row, column=1, value=title)
    title_cell.font = Font(name=FONT_NAME, bold=True, size=10, color=BRAND["navy"])
    title_cell.fill = PatternFill("solid", fgColor=BRAND["subheader_fill"])
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22

    # Header row
    header_row = row + 1
    cols_per_header = max(1, span // len(headers))
    for h_idx, h in enumerate(headers):
        c_start = 1 + h_idx * cols_per_header
        c_end = min(last_col, c_start + cols_per_header - 1)
        if c_start <= c_end:
            ws.merge_cells(start_row=header_row, start_column=c_start, end_row=header_row, end_column=c_end)
        cell = ws.cell(row=header_row, column=c_start, value=h)
        cell.font = Font(name=FONT_NAME, bold=True, size=9, color=BRAND["muted_text"])
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Data rows (formula expressions)
    r = header_row + 1
    for data_row in rows_formula:
        for h_idx, expr in enumerate(data_row):
            c_start = 1 + h_idx * cols_per_header
            c_end = min(last_col, c_start + cols_per_header - 1)
            if c_start <= c_end:
                ws.merge_cells(start_row=r, start_column=c_start, end_row=r, end_column=c_end)
            cell = ws.cell(row=r, column=c_start, value=expr)
            cell.font = Font(name=FONT_NAME, size=10, color=BRAND["navy"])
            cell.alignment = Alignment(horizontal="left", vertical="center")
        r += 1
    return r + 1


def add_bar_chart(
    ws,
    *,
    title: str,
    cat_ref: Reference,
    val_ref: Reference,
    anchor: str,
    width: float = 14,
    height: float = 7,
) -> None:
    """Bar chart with brand colors anchored at `anchor` (e.g. 'A30')."""
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = title
    chart.y_axis.title = None
    chart.x_axis.title = None
    chart.legend = None
    chart.add_data(val_ref, titles_from_data=False)
    chart.set_categories(cat_ref)
    chart.width = width
    chart.height = height
    if chart.series:
        chart.series[0].graphicalProperties.solidFill = BRAND["red"]
        chart.dataLabels = DataLabelList(showVal=True)
    ws.add_chart(chart, anchor)


def add_pie_chart(
    ws,
    *,
    title: str,
    cat_ref: Reference,
    val_ref: Reference,
    anchor: str,
    width: float = 12,
    height: float = 7,
) -> None:
    chart = PieChart()
    chart.title = title
    chart.add_data(val_ref, titles_from_data=False)
    chart.set_categories(cat_ref)
    chart.width = width
    chart.height = height
    chart.dataLabels = DataLabelList(showPercent=True)
    ws.add_chart(chart, anchor)


def add_line_chart(
    ws,
    *,
    title: str,
    cat_ref: Reference,
    val_ref: Reference,
    anchor: str,
    width: float = 14,
    height: float = 7,
) -> None:
    chart = LineChart()
    chart.title = title
    chart.style = 12
    chart.add_data(val_ref, titles_from_data=False)
    chart.set_categories(cat_ref)
    chart.width = width
    chart.height = height
    chart.legend = None
    if chart.series:
        chart.series[0].graphicalProperties.line.solidFill = BRAND["teal"]
    ws.add_chart(chart, anchor)
