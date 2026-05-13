"""Seed the new E3 Elevate Bridge Programme workbook from the four source
xlsx files in /Elevate 3.0/ElevateBridge/.

Idempotent: applicant_id is sha1(lowercased email)[:10] prefixed with EB- so
re-running this script over an existing workbook would (in theory) re-mint
the same ids. The migrator always rebuilds the workbook from scratch via
the builder, so re-running is safe.

Source tabs:
  ElevateBridge Freelancers Application Responses.xlsx
    -> Master Data           -> Applicants (one row per applicant)
    -> S1|| Killing Factor   -> S1 Killing Factor
    -> S2|| Tracks Sorting   -> S2 Tracks Sorting
    -> Interview Scoring - FLA -> Interview Scoring (track=FL, Q1..Q11)
    -> Interview Scoring - SM  -> Interview Scoring (track=SM, Q1..Q14)
  Top Freelancers.xlsx
    -> Upwork|| For Service-based tech -> Top Performers (track=FL, 15 rows)
    -> SDR&BDR|| For Product-Based Com -> Top Performers (track=SM, 11 rows)
  Elevated Bridge Capacity Material.xlsx
    -> Upwork                            -> Training Sessions (track=Upwork Agency Building)
    -> Business Development and tech s   -> Training Sessions (track=Business Development & Tech Sales)
    -> Google Maps Lead Generation       -> Training Sessions (track=Google Maps Lead Generation)

Also seeds:
  Mentors (3 rows from the report PDF)
  Final Decisions (rows where Track [assigned] != "" -> Admitted)
"""

import hashlib
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from builders.elevate_bridge_portal import FILENAME, build as build_template


SRC_DIR = Path(__file__).resolve().parents[2] / "ElevateBridge"
RESPONSES_XLSX  = SRC_DIR / "ElevateBridge Freelancers Application Responses.xlsx"
TOP_XLSX        = SRC_DIR / "Top Freelancers.xlsx"
CAPACITY_XLSX   = SRC_DIR / "Elevated Bridge Capacity Material.xlsx"


# Mentor metadata sourced from the Selection Process Report PDF
# (pages 16-18). Hours derived from the report's budget breakdown:
#   Google Maps Lead Gen: 12 hours @ $40/hr = $480
#   BD & Tech Sales:      30 hours @ $55/hr = $1,650
#   Upwork Agency Building: 5 sessions, hours not specified ~ 10 mentored hours
MENTORS = [
    {
        "mentor_id": "MTR-ANAS",
        "full_name": "Anas Albaghdadi",
        "email":     "dranassofian@gmail.com",
        "whatsapp":  "+27 69 826 4970",
        "track":     "Upwork Agency Building",
        "hourly_rate": "0",
        "total_hours": "10",
        "budget_total": "0",
        "bio": "Mentor for the Upwork track. 5 sessions covering individual vs agency accounts, agency plans & pricing, roles & permissions, agency economics, finding a niche, kickstarting the agency, portfolio & skills positioning, native KPIs and review process.",
    },
    {
        "mentor_id": "MTR-KARIM",
        "full_name": "Karim Alharazin",
        "email":     "karim.alharazin@gmail.com",
        "whatsapp":  "+20 10 64756940",
        "track":     "Google Maps Lead Generation",
        "hourly_rate": "40",
        "total_hours": "12",
        "budget_total": "480",
        "bio": "Mentor for the Google Maps lead-gen track. 6 sessions covering modern lead structures, list building, ready-made scrapers, custom scraping with JSON/HTML, n8n workflow automation, and project demo / Q&A.",
    },
    {
        "mentor_id": "MTR-YOUSEF",
        "full_name": "Yousef Yaqoub \"Mohammad Rabee\" Ashhab",
        "email":     "joseashhab@gmail.com",
        "whatsapp":  "+972-598831363",
        "track":     "Business Development & Tech Sales",
        "hourly_rate": "55",
        "total_hours": "30",
        "budget_total": "1650",
        "bio": "Mentor for the BD & tech sales track. 6 modules covering tech revenue landscape, software economics (CAC/LTV/churn), sales architecture & funnels, lead gen & prospecting, negotiation & closing, CRM & automation basics.",
    },
]


def _applicant_id(email: str) -> str:
    e = (email or "").strip().lower()
    if not e:
        return ""
    h = hashlib.sha1(e.encode("utf-8")).hexdigest()[:10]
    return f"EB-{h}"


def _region_of(location_category: str) -> str:
    s = (location_category or "").lower()
    if "gaza" in s:
        return "Gaza Strip"
    if "west bank" in s or " wb" in s or s.endswith("wb") or "- wb" in s:
        return "West Bank"
    if "outside" in s or "displaced" in s:
        return "Outside Palestine"
    return ""


def _normalize_kf(raw: str) -> tuple[str, str]:
    """Source workbook stores killing factor as Yes/No; portal expects
    Pass/Fail. Returns (result, reason_hint) where result is the
    canonical value and reason_hint is empty (the seeder uses col 18 for
    the actual reason)."""
    s = (raw or "").strip().lower()
    if s in ("yes", "pass", "passed", "in"):
        return ("Pass", "")
    if s in ("no", "fail", "failed", "out"):
        return ("Fail", "")
    return ("", "")


def _normalize_track(t: str) -> str:
    """Canonicalise the source workbook's track strings into FL / SM / FL+SM."""
    if not t:
        return ""
    s = str(t).upper().strip()
    if "FL" in s and "SM" in s:
        return "FL+SM"
    if "SM" in s:
        return "SM"
    if "FL" in s or "UPWORK" in s or "FREELANC" in s:
        return "FL"
    return ""


def _str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _clean(v) -> str:
    """Treat the source workbook's N/A placeholder as empty so the UI shows
    an em-dash instead of literal 'N/A'."""
    s = _str(v)
    if s.upper() == "N/A":
        return ""
    return s


def _format_earnings(raw) -> str:
    """Top Performers earnings: source has both raw float (100000.0) and a
    formatted column ('$100K'). Prefer the formatted; otherwise format ourselves."""
    if raw is None or raw == "":
        return ""
    if isinstance(raw, str):
        s = raw.strip()
        if s:
            return s
    try:
        n = float(raw)
    except (TypeError, ValueError):
        return _str(raw)
    if n >= 1000:
        if n >= 1_000_000:
            return f"${n/1_000_000:.1f}M"
        if n % 1000 == 0:
            return f"${int(n/1000)}K"
        return f"${n/1000:.1f}K"
    return f"${int(n)}"


def _seed_applicants(wb, master_rows):
    """Master Data -> Applicants. One row per of the 203."""
    ws = wb["Applicants"]
    # Column map: builders/elevate_bridge_portal.py APPLICANTS_HEADERS
    # 1 applicant_id, 2 full_name_en, 3 full_name_ar, 4 email, 5 phone,
    # 6 gender, 7 dob, 8 location, 9 region, 10 specialization,
    # 11 education_level, 12 university_name, 13 university_major,
    # 14 has_disability, 15 electricity_reliability, 16 internet_reliability,
    # 17 track_registered, 18 track_assigned, 19 current_stage, 20 decision,
    # 21 killing_factor_result, 22 killing_factor_reason, 23 ssi_score,
    # 24 response_score_fl, 25 response_score_sm, 26 interview_score_fl,
    # 27 interview_score_sm, 28 total_score, 29 hours_per_week,
    # 30 session_commitment, 31 notes, 32 updated_at, 33 updated_by

    now = datetime.utcnow().isoformat()
    row = 2
    ingested = 0
    for r in master_rows:
        email = _str(r[0])
        name  = _str(r[1])
        if not email and not name:
            continue
        aid = _applicant_id(email) or f"EB-row{row}"
        kf_raw      = _str(r[16])                # Killing Factor result (Yes/No)
        in_out      = _str(r[17])                # In/Out
        attended    = _str(r[23])                # Attended Interview?
        track_assigned = _normalize_track(_str(r[12]))
        kf_canon, _ = _normalize_kf(kf_raw)

        # Stage progression:
        #   In/Out = "In"            -> Decision
        #   Attended Interview "Yes" -> Interview
        #   track_assigned set       -> S3 Scoring
        #   kf = Pass                -> S2 Sort
        #   kf = Fail                -> S1 Filter (waitlisted)
        #   otherwise                -> Applied
        s_in_out = in_out.lower()
        s_attended = attended.lower()
        if s_in_out == "in":
            stage = "Decision"
        elif s_attended.startswith("yes"):
            stage = "Interview"
        elif track_assigned:
            stage = "S3 Scoring"
        elif kf_canon == "Pass":
            stage = "S2 Sort"
        elif kf_canon == "Fail":
            stage = "S1 Filter"
        else:
            stage = "Applied"

        # Decision: derive deterministically where possible.
        decision = ""
        if s_attended.startswith("yes-withdrew") or "withdrew" in s_attended:
            decision = "Withdrew"
        elif s_in_out == "in":
            decision = "Admitted"
        elif kf_canon == "Fail":
            decision = "Waitlisted"

        ws.cell(row=row, column=1,  value=aid)
        ws.cell(row=row, column=2,  value=name)
        ws.cell(row=row, column=4,  value=email)
        ws.cell(row=row, column=5,  value=_str(r[3]))           # phone
        ws.cell(row=row, column=8,  value=_clean(r[4]))         # location (strip N/A)
        ws.cell(row=row, column=9,  value=_region_of(_clean(r[5])))
        ws.cell(row=row, column=17, value=_normalize_track(_str(r[10] or r[11])))
        ws.cell(row=row, column=18, value=track_assigned)
        ws.cell(row=row, column=19, value=stage)
        ws.cell(row=row, column=20, value=decision)
        ws.cell(row=row, column=21, value=kf_canon)             # killing_factor_result (Pass/Fail)
        ws.cell(row=row, column=22, value=_str(r[18]))          # reason
        ws.cell(row=row, column=24, value=_str(r[20]))          # response_score_fl
        ws.cell(row=row, column=25, value=_str(r[21]))          # response_score_sm
        ws.cell(row=row, column=26, value=_str(r[24]))          # interview_score_fl
        ws.cell(row=row, column=27, value=_str(r[25]))          # interview_score_sm
        flsm = _str(r[26])
        fls  = _str(r[24])
        sms  = _str(r[25])
        total = flsm or fls or sms or _str(r[33])
        ws.cell(row=row, column=28, value=total)
        ws.cell(row=row, column=31, value=_str(r[19]))          # notes
        ws.cell(row=row, column=32, value=now)
        ws.cell(row=row, column=33, value="seeder")
        row += 1
        ingested += 1
    return ingested


def _seed_stage1(wb, rows):
    ws = wb["S1 Killing Factor"]
    # S1 headers (15 cols):
    # applicant_id, full_name, email, phone, location, track_registered,
    # killing_factor_result, reason, total_upwork_income_i, upwork_income_jh,
    # upwork_income_agency, sm_income_i, sm_income_agency,
    # total_income_normalized, notes
    row = 2
    n = 0
    for r in rows:
        email = _str(r[1])
        name  = _str(r[0])
        if not email and not name:
            continue
        kf_canon, _ = _normalize_kf(_str(r[7]))
        ws.cell(row=row, column=1,  value=_applicant_id(email) or f"EB-row{row}")
        ws.cell(row=row, column=2,  value=name)
        ws.cell(row=row, column=3,  value=email)
        ws.cell(row=row, column=4,  value=_str(r[2]))
        ws.cell(row=row, column=5,  value=_str(r[3]))
        ws.cell(row=row, column=6,  value=_normalize_track(_str(r[6])))
        ws.cell(row=row, column=7,  value=kf_canon)             # Pass/Fail
        ws.cell(row=row, column=8,  value=_str(r[8]))           # reason
        ws.cell(row=row, column=9,  value=_str(r[9]))
        ws.cell(row=row, column=10, value=_str(r[10]))
        ws.cell(row=row, column=11, value=_str(r[11]))
        ws.cell(row=row, column=12, value=_str(r[12]))
        ws.cell(row=row, column=13, value=_str(r[13]))
        ws.cell(row=row, column=14, value=_str(r[15]))
        ws.cell(row=row, column=15, value=_str(r[22]))
        row += 1
        n += 1
    return n


def _seed_stage2(wb, rows):
    ws = wb["S2 Tracks Sorting"]
    # S2 headers: applicant_id, full_name, email, track_registered, track_assigned,
    # skills_category, jh_sector_upwork, jh_sector_sm, notes
    row = 2
    n = 0
    for r in rows:
        email = _str(r[1])
        name  = _str(r[0])
        if not email and not name:
            continue
        ws.cell(row=row, column=1, value=_applicant_id(email) or f"EB-row{row}")
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=email)
        ws.cell(row=row, column=4, value=_normalize_track(_str(r[9] or r[10])))
        ws.cell(row=row, column=5, value=_normalize_track(_str(r[11])))
        ws.cell(row=row, column=6, value=_str(r[6]))
        ws.cell(row=row, column=7, value=_str(r[7]))
        ws.cell(row=row, column=8, value=_str(r[8]))
        ws.cell(row=row, column=9, value=_str(r[26]))
        row += 1
        n += 1
    return n


# Per-criterion response scoring from the S3||Response Scoring source tabs.
# The source has the 1-5 scores in cols 11..17 with a stable column order.
# These mappings must match the criterion_key in RUBRIC_ROWS_*_RESPONSE so
# the Scoring tab's grid joins them correctly.
FL_RESPONSE_CRITERIA = [
    # (criterion_key, criterion_label, category, weight, sub_weight, src_col_idx, src_label)
    ("ind_income",            "Total Individual Income ($)", "Upwork Earnings",  "40", "60", 11, "Individual Income"),
    ("agency_income",         "Agency Income ($)",           "Upwork Earnings",  "40", "40", 12, "Agency Income"),
    ("project_count",         "Project Count",               "Performance",      "30", "60", 13, "Project Count"),
    ("other_platforms",       "Other Freelancing Platforms", "Performance",      "30", "40", 14, "Other Platforms"),
    ("badges",                "Badges (Upwork)",             "Profile Quality",  "30", "35", 15, "Badge"),
    ("jss",                   "JSS (Job Success Score)",     "Profile Quality",  "30", "35", 16, "JSS"),
    ("agency_experience_resp","Agency Experience",           "Profile Quality",  "30", "30", 17, "Agency Experience"),
]
SM_RESPONSE_CRITERIA = [
    ("ind_income_sm",         "Total Individual Income ($)", "SM Earnings",     "35", "30", 11, "Individual Income"),
    ("agency_income_sm",      "Agency Income ($)",           "SM Earnings",     "35", "30", 12, "Agency Income"),
    ("verification",          "Verification Provided",       "SM Earnings",     "35", "40", 13, "Verification"),
    ("ssi",                   "Social Selling Index",        "Profile Quality", "50", "70", 14, "SSI"),
    ("agency_experience_sm",  "Agency Experience",           "Profile Quality", "50", "30", 15, "Agency Experience"),
    ("project_count_sm",      "Project Count",               "Performance",     "15", "50", 16, "Project Count"),
    ("other_sm_platforms",    "Other SM Platforms",          "Performance",     "15", "50", 17, "Other SM Platforms"),
]


def _seed_response_scoring(wb, src):
    """Seed the long-form S3 Response Scoring tab from the source's two
    per-applicant scoring tabs."""
    ws = wb["S3 Response Scoring"]
    row = 2
    n = 0
    now = datetime.utcnow().isoformat()

    def _emit(src_tab, track, criteria):
        nonlocal row, n
        if src_tab not in src.sheetnames:
            return
        rows = list(src[src_tab].iter_rows(values_only=True))[1:]
        for r in rows:
            name = _clean(r[0]) if len(r) > 0 else ""
            email = _clean(r[1]) if len(r) > 1 else ""
            if not name and not email:
                continue
            aid = _applicant_id(email) or f"EB-{name[:6]}"
            for (key, label, category, weight, sub_weight, src_col, _src_label) in criteria:
                if src_col >= len(r):
                    continue
                score = _str(r[src_col])
                if not score:
                    continue
                score_id = f"{aid}-{track}-{key}".lower().replace("+", "p")
                ws.cell(row=row, column=1,  value=score_id)
                ws.cell(row=row, column=2,  value=aid)
                ws.cell(row=row, column=3,  value=name)
                ws.cell(row=row, column=4,  value=track)
                ws.cell(row=row, column=5,  value=category)
                ws.cell(row=row, column=6,  value=key)
                ws.cell(row=row, column=7,  value=label)
                ws.cell(row=row, column=8,  value=weight)
                ws.cell(row=row, column=9,  value=sub_weight)
                ws.cell(row=row, column=10, value=score)
                ws.cell(row=row, column=12, value="seeder")
                ws.cell(row=row, column=13, value=now)
                row += 1
                n += 1

    _emit("S3||Response Scoring - FL", "FL", FL_RESPONSE_CRITERIA)
    _emit("S3||Response Scoring - SM", "SM", SM_RESPONSE_CRITERIA)
    return n


def _seed_interview_fl(wb, rows):
    """Interview Scoring - FLA: Q1..Q11 → 11 long-form rows per applicant.
    Source cols: [0] Name [1] Email [2] Attended [9..19] Q1..Q11
    """
    ws = wb["Interview Scoring"]
    # Walk the worksheet to find the next empty row.
    row = 2
    while ws.cell(row=row, column=1).value:
        row += 1
    n = 0
    for r in rows:
        email = _str(r[1])
        name  = _str(r[0])
        if not email and not name:
            continue
        attended = _str(r[2])
        if not attended and not any(_str(r[i]) for i in range(9, 20)):
            continue
        aid = _applicant_id(email) or f"EB-{name[:6]}"
        for q in range(1, 12):
            cell_val = r[8 + q] if len(r) > 8 + q else None
            score = _str(cell_val)
            if not score:
                continue
            ws.cell(row=row, column=1,  value=f"{aid}-FL-Q{q}")
            ws.cell(row=row, column=2,  value=aid)
            ws.cell(row=row, column=3,  value=name)
            ws.cell(row=row, column=4,  value="FL")
            ws.cell(row=row, column=5,  value=str(q))
            ws.cell(row=row, column=6,  value="Freelancing track")
            ws.cell(row=row, column=7,  value=f"Q{q}")
            ws.cell(row=row, column=10, value=score)
            ws.cell(row=row, column=11, value=attended)
            row += 1
            n += 1
    return n


def _seed_interview_sm(wb, rows):
    """Interview Scoring - SM: Q1..Q14 → 14 long-form rows per applicant.
    Source cols: [0] Name [1] Email [2] Attended [3] Match track [4] Earnings
                 [5] Proof of income [6] Proof Scoring [7] Agency Exp [8..21] Q1..Q14
    """
    ws = wb["Interview Scoring"]
    row = 2
    while ws.cell(row=row, column=1).value:
        row += 1
    n = 0
    for r in rows:
        email = _str(r[1])
        name  = _str(r[0])
        if not email and not name:
            continue
        attended = _str(r[2])
        if not attended and not any(_str(r[i]) for i in range(8, 22) if i < len(r)):
            continue
        aid = _applicant_id(email) or f"EB-{name[:6]}"
        for q in range(1, 15):
            cell_val = r[7 + q] if len(r) > 7 + q else None
            score = _str(cell_val)
            if not score:
                continue
            ws.cell(row=row, column=1,  value=f"{aid}-SM-Q{q}")
            ws.cell(row=row, column=2,  value=aid)
            ws.cell(row=row, column=3,  value=name)
            ws.cell(row=row, column=4,  value="SM")
            ws.cell(row=row, column=5,  value=str(q))
            ws.cell(row=row, column=6,  value="Social Media track")
            ws.cell(row=row, column=7,  value=f"Q{q}")
            ws.cell(row=row, column=10, value=score)
            ws.cell(row=row, column=11, value=attended)
            row += 1
            n += 1
    return n


def _seed_top_performers(wb, master_rows):
    """Top Performers from Top Freelancers.xlsx. Withdrew status is derived
    from the master "Attended Interview?" column (look up by email)."""
    # Build a (lowercase email) -> Attended Interview map.
    withdrew_lookup = {}
    for r in master_rows:
        email = _str(r[0]).lower()
        attended = _str(r[23]).lower()
        if email and "withdrew" in attended:
            withdrew_lookup[email] = "Yes"

    ws = wb["Top Performers"]
    row = 2
    n = 0
    src = load_workbook(TOP_XLSX, read_only=True, data_only=True)

    # FL track: "Upwork|| For Service-based tech"
    fl_tab = next((t for t in src.sheetnames if t.startswith("Upwork||")), None)
    if fl_tab:
        rows = src[fl_tab].iter_rows(values_only=True)
        next(rows)  # skip header
        for r in rows:
            name = _str(r[5])
            email = _str(r[7])
            if not name and not email:
                continue
            ws.cell(row=row, column=1,  value=_applicant_id(email) or f"EB-{name[:6]}")
            ws.cell(row=row, column=2,  value=_str(r[0]))
            ws.cell(row=row, column=3,  value=_str(r[1]))
            ws.cell(row=row, column=4,  value=_str(r[2]))
            ws.cell(row=row, column=5,  value="FL")
            ws.cell(row=row, column=6,  value=name)
            ws.cell(row=row, column=7,  value=_str(r[6]))
            ws.cell(row=row, column=8,  value=email)
            ws.cell(row=row, column=9,  value=_str(r[8]))
            ws.cell(row=row, column=10, value=_str(r[9]))
            ws.cell(row=row, column=11, value=_str(r[10]))
            ws.cell(row=row, column=12, value=_str(r[11]))
            ws.cell(row=row, column=13, value=_str(r[12]))
            ws.cell(row=row, column=14, value=_clean(r[13]))
            # FL track has no separate performance score column; use the
            # formatted earnings ($XK) when present, else format the raw float.
            formatted = _str(r[22]) if len(r) > 22 and r[22] else _format_earnings(r[4])
            ws.cell(row=row, column=15, value=formatted)         # total_earnings
            ws.cell(row=row, column=16, value=formatted)         # performance_score (display)
            ws.cell(row=row, column=17, value=_str(r[21]) if len(r) > 21 else "")
            ws.cell(row=row, column=18, value=withdrew_lookup.get(email.lower(), "No"))
            row += 1
            n += 1

    # SM track: "SDR&BDR|| For Product-Based Com"
    sm_tab = next((t for t in src.sheetnames if t.startswith("SDR&BDR")), None)
    if sm_tab:
        rows = src[sm_tab].iter_rows(values_only=True)
        next(rows)
        for r in rows:
            name = _str(r[5])
            email = _str(r[7])
            if not name and not email:
                continue
            ws.cell(row=row, column=1,  value=_applicant_id(email) or f"EB-{name[:6]}")
            ws.cell(row=row, column=2,  value=_str(r[0]))
            ws.cell(row=row, column=3,  value=_str(r[1]))
            ws.cell(row=row, column=4,  value=_str(r[2]))
            ws.cell(row=row, column=5,  value="SM")
            ws.cell(row=row, column=6,  value=name)
            ws.cell(row=row, column=7,  value=_str(r[6]))
            ws.cell(row=row, column=8,  value=email)
            ws.cell(row=row, column=9,  value=_str(r[8]))
            ws.cell(row=row, column=10, value=_str(r[9]))
            ws.cell(row=row, column=11, value=_str(r[10]))
            ws.cell(row=row, column=12, value=_str(r[11]))
            ws.cell(row=row, column=13, value=_str(r[12]))
            ws.cell(row=row, column=14, value=_clean(r[13]))
            # SM track has Performance Score in col 3 and Total Income in col 30/31.
            sm_earnings = _format_earnings(r[31]) if len(r) > 31 and r[31] not in (None, "") else _format_earnings(r[30])
            ws.cell(row=row, column=15, value=sm_earnings)       # total_earnings
            ws.cell(row=row, column=16, value=_str(r[3]))        # performance_score (real)
            ws.cell(row=row, column=17, value=_str(r[25]) if len(r) > 25 else "")
            ws.cell(row=row, column=18, value=withdrew_lookup.get(email.lower(), "No"))
            row += 1
            n += 1
    return n


# Scoring rubrics distilled from the Selection Process Report (pages 4-11).
# Each row defines one criterion: stage (Response | Interview), track, category,
# weight (category weight %), sub_weight (criterion weight within category %),
# and the 1-5 score descriptors that drive the tooltip in the Scoring tab.

RUBRIC_ROWS_FL_RESPONSE = [
    # Upwork Earnings 40%
    ("FL", "Response", "Upwork Earnings", "ind_income",
     "Total Individual Income ($)", 40, 60,
     "≥ 30,000", "10,000-29,999", "5,000-9,999", "2,000-4,999", "0-1,999"),
    ("FL", "Response", "Upwork Earnings", "agency_income",
     "Agency Income ($)", 40, 40,
     "≥ 4,000", "3,000-3,999", "2,000-2,999", "1,000-1,999", "< 1,000"),
    # Performance 30%
    ("FL", "Response", "Performance", "project_count",
     "Project Count", 30, 60,
     "≥ 50", "40-49", "30-39", "20-29", "< 20"),
    ("FL", "Response", "Performance", "other_platforms",
     "Other Freelancing Platforms", 30, 40,
     "≥ 4", "3", "2", "1", "0"),
    # Profile Quality 30%
    ("FL", "Response", "Profile Quality", "badges",
     "Badges (Upwork)", 30, 35,
     "Expert-Vetted", "Top-Rated Plus", "Top-Rated", "Rising Talent", "None"),
    ("FL", "Response", "Profile Quality", "jss",
     "JSS (Job Success Score)", 30, 35,
     "100%", "97-99%", "94-96%", "91-93%", "> 90% or N/A"),
    ("FL", "Response", "Profile Quality", "agency_experience_resp",
     "Agency Experience", 30, 30,
     "Yes", "", "", "", "No"),
]

RUBRIC_ROWS_SM_RESPONSE = [
    # SM Earnings 35%
    ("SM", "Response", "SM Earnings", "ind_income_sm",
     "Total Individual Income ($)", 35, 30,
     "≥ 30,000", "15,000-29,999", "6,000-14,999", "2,000-5,999", "0-1,999"),
    ("SM", "Response", "SM Earnings", "agency_income_sm",
     "Agency Income ($)", 35, 30,
     "≥ 25,000", "10,000-24,999", "5,000-9,999", "2,000-4,999", "0-1,999"),
    ("SM", "Response", "SM Earnings", "verification",
     "Verification Provided", 35, 40,
     "Yes", "", "", "", "No"),
    # Profile Quality 50%
    ("SM", "Response", "Profile Quality", "ssi",
     "Social Selling Index (LinkedIn)", 50, 70,
     "≥ 50", "30-49", "20-29", "10-19", "0-9"),
    ("SM", "Response", "Profile Quality", "agency_experience_sm",
     "Agency Experience", 50, 30,
     "Yes", "", "", "", "No"),
    # Performance 15%
    ("SM", "Response", "Performance", "project_count_sm",
     "Project Count", 15, 50,
     "≥ 50", "30-49", "20-29", "10-19", "0-9"),
    ("SM", "Response", "Performance", "other_sm_platforms",
     "Other SM Platforms", 15, 50,
     "≥ 5", "4", "3", "2", "1-0"),
]

RUBRIC_ROWS_FL_INTERVIEW = [
    ("FL", "Interview", "Freelance Platform Earnings", "q1_earning",
     "Q1: Earning", 50, 100,
     "≥ 30,000", "10,000-29,999", "5,000-9,999", "2,000-4,999", "0-1,999"),
    ("FL", "Interview", "Agency Experience", "q2_agency_experience",
     "Q2: Agency Experience", 50, 50,
     "Strong Knowledge", "", "Basic Theoretical Understanding", "", "Weak Knowledge"),
    ("FL", "Interview", "Agency Experience", "q3_prequalify",
     "Q3: How do you Prequalify Projects?", 50, 50,
     "Data-Based qualification", "", "Partial Process / Missing Structure", "", "Applies Randomly"),
    ("FL", "Interview", "Processes & Methods", "q4_tools",
     "Q4: Tools/Methods for Leads Analysis", 50, 20,
     "Uses Actual tools (Upwork Analytics)", "", "Manual Tracking", "", "No Tracking"),
    ("FL", "Interview", "Processes & Methods", "q5_inbound",
     "Q5: Inbound Strategies", 50, 20,
     "Invitations, Gigs, Consultations", "", "Invitations and Profile Optimizations", "", "No real understanding of inbound mechanics"),
    ("FL", "Interview", "Processes & Methods", "q6_pricing",
     "Q6: Pricing Models", 50, 20,
     "Provides a clear pricing strategy", "", "Attempts negotiation but misses structure", "", "Accepts or declines poorly"),
    ("FL", "Interview", "Processes & Methods", "q7_low_budget",
     "Q7: Handling Low-Budget Offers", 50, 20,
     "Re-scope, offer alternatives", "", "Unstructured negotiation", "", "Accepts or declines poorly"),
    ("FL", "Interview", "Processes & Methods", "q8_jss",
     "Q8: JSS (Job Success Score)", 50, 20,
     "Calculation Drivers, end visibility effect", "", "Knows what JSS is but lacks depth", "", "Doesn't know how it works"),
    ("FL", "Interview", "Negotiation & Problem Solving", "q9_response_rate",
     "Q9: Response Rate dropped — root cause", 50, 33,
     "Root cause analysis", "", "One area only (e.g. Proposals)", "", "No analytical method"),
    ("FL", "Interview", "Negotiation & Problem Solving", "q10_walk_through",
     "Q10: Walk me through challenging Upwork project", 50, 33,
     "Clear Example with Ownership", "", "Vague, incomplete example", "", "No meaningful Experience"),
    ("FL", "Interview", "Negotiation & Problem Solving", "q11_scope_creep",
     "Q11: Managing last-minute scope creep before signing", 50, 34,
     "Re-scope, Re-price, maintains profitability", "", "Basic negotiation, no process", "", "Accepts change without protection"),
]

RUBRIC_ROWS_SM_INTERVIEW = [
    ("SM", "Interview", "SM Earnings", "q1_earning_sm",
     "Q1: Earning", 20, 100,
     "≥ 30,000", "15,000-29,999", "6,000-14,999", "2,000-5,999", "0-1,999"),
    ("SM", "Interview", "Agency Experience & Proof of Income", "q2_agency_exp_sm",
     "Q2: Agency Experience", 0, 50,
     "Clear Agent Experience + Defined Responsibilities", "", "Partial Agent Exposure", "", "No Agency Experience"),
    ("SM", "Interview", "Agency Experience & Proof of Income", "q3_proof",
     "Q3: Proof of Income", 0, 50,
     "Proof of income provided", "", "Proof of income not provided", "", "Proof of income not provided"),
    ("SM", "Interview", "SM Acquisition Knowledge", "q4_inbound_outbound",
     "Q4: Inbound vs Outbound", 80, 12,
     "Correct definitions + structured strategies for both", "", "Basic Understanding", "", "Vague Answer"),
    ("SM", "Interview", "SM Acquisition Knowledge", "q5_selling_outreach",
     "Q5: SM Selling vs Outreach", 80, 12,
     "Clear differentiation + real social selling examples", "", "Understands concept but lacks examples", "", "Vague Answer"),
    ("SM", "Interview", "SM Acquisition Knowledge", "q6_endtoend",
     "Q6: End-to-End SM Acquisition Process", 80, 12,
     "Provides a full funnel", "", "Partial Process without structure", "", "No real process"),
    ("SM", "Interview", "SM Acquisition Knowledge", "q7_design",
     "Q7: Design SM Acquisition System", 80, 12,
     "Clear, Logical System Design", "", "Some Elements but not organized", "", "No Structure"),
    ("SM", "Interview", "SM Acquisition Knowledge", "q8_platform_prio",
     "Q8: Platform Prioritization — B2B vs B2C", 80, 12,
     "Correct Platform Mapping", "", "Basic Mapping", "", "Generic Answer"),
    ("SM", "Interview", "SM Acquisition Knowledge", "q9_icp",
     "Q9: ICP", 80, 12,
     "Correct Definition + clear example of ICP", "", "Knows definition only", "", "Incorrect Answer"),
    ("SM", "Interview", "Analytical Thinking", "q10_leads_no_close",
     "Q10: You had 20 leads but closed 0% — Diagnose root cause & KPIs", 80, 14,
     "Uses MFA & structured diagnosis steps", "", "Basic analysis", "", "No real analysis"),
    ("SM", "Interview", "Analytical Thinking", "q11_engagement_drop",
     "Q11: Engagement dropped by 50% despite consistent posting — What do you check?", 80, 14,
     "Checks Content structure, Performance, algorithm factors, or conversion rates", "", "Content Timing", "", "No real analysis"),
    ("SM", "Interview", "Communication & Sales", "q12_lead_cold",
     "Q12: High-value lead goes cold for 5 days. How do you re-engage?", 0, 33,
     "Add value, re-opens conversation professionally", "", "Basic followup", "", "Pushy or passive approach"),
    ("SM", "Interview", "Communication & Sales", "q13_high_price",
     "Q13: Client says 'Your price is too high; others are cheaper'. How do you respond?", 0, 33,
     "Reframe Value and offers alternatives", "", "Polite but weak negotiation", "", "Accepts or responds poorly"),
    ("SM", "Interview", "Communication & Sales", "q14_posting",
     "Q14: Posting schedule & content architecture to convert strangers into leads", 0, 34,
     "Clear hands-on experience, posting cadence, mapping content to funnel stages", "", "Surface level experience, posting cadence and some content types", "", "Vague Answer"),
]

RUBRIC_ROWS_FLSM_INTERVIEW = [
    # Combined track applicants are interviewed under the SM rubric (per
    # report page 8); the same rows are duplicated under FL+SM so the
    # Scoring tab's "Interview · FL+SM" sub-tab renders.
    ("FL+SM", "Interview", category, key, label, weight, sub_weight, s5, s4, s3, s2, s1)
    for (_, _, category, key, label, weight, sub_weight, s5, s4, s3, s2, s1)
    in RUBRIC_ROWS_SM_INTERVIEW
]

RUBRIC_ROWS_FLSM_RESPONSE = [
    # Combined-track response scoring uses the FL response rubric + the
    # SM response rubric weighted equally. Listed as a single set so the
    # Scoring tab's "Response · FL+SM" sub-tab renders.
    ("FL+SM", "Response", category, key, label, weight, sub_weight, s5, s4, s3, s2, s1)
    for (_, _, category, key, label, weight, sub_weight, s5, s4, s3, s2, s1)
    in (RUBRIC_ROWS_FL_RESPONSE + RUBRIC_ROWS_SM_RESPONSE)
]


def _seed_rubrics(wb):
    ws = wb["Scoring Rubrics"]
    row = 2
    n = 0
    all_rows = (
        RUBRIC_ROWS_FL_RESPONSE
        + RUBRIC_ROWS_SM_RESPONSE
        + RUBRIC_ROWS_FLSM_RESPONSE
        + RUBRIC_ROWS_FL_INTERVIEW
        + RUBRIC_ROWS_SM_INTERVIEW
        + RUBRIC_ROWS_FLSM_INTERVIEW
    )
    for (track, stage, category, key, label, weight, sub_weight, s5, s4, s3, s2, s1) in all_rows:
        rubric_id = f"RUB-{track}-{stage[:3]}-{key}".lower().replace("+", "p")
        ws.cell(row=row, column=1,  value=rubric_id)
        ws.cell(row=row, column=2,  value=track)
        ws.cell(row=row, column=3,  value=stage)
        ws.cell(row=row, column=4,  value=category)
        ws.cell(row=row, column=5,  value=key)
        ws.cell(row=row, column=6,  value=label)
        ws.cell(row=row, column=7,  value=str(weight))
        ws.cell(row=row, column=8,  value=str(sub_weight))
        ws.cell(row=row, column=9,  value=s5)
        ws.cell(row=row, column=10, value=s4)
        ws.cell(row=row, column=11, value=s3)
        ws.cell(row=row, column=12, value=s2)
        ws.cell(row=row, column=13, value=s1)
        row += 1
        n += 1
    return n


def _seed_mentors(wb):
    ws = wb["Mentors"]
    row = 2
    for m in MENTORS:
        for col_idx, key in enumerate(
            ["mentor_id", "full_name", "email", "whatsapp", "track",
             "hourly_rate", "total_hours", "budget_total", "bio"],
            start=1,
        ):
            ws.cell(row=row, column=col_idx, value=m[key])
        row += 1
    return len(MENTORS)


def _seed_sessions(wb):
    ws = wb["Training Sessions"]
    row = 2
    n = 0
    src = load_workbook(CAPACITY_XLSX, read_only=True, data_only=True)
    now = datetime.utcnow().isoformat()

    # Source workbook's Upwork tab has no Date/Topic columns. The Selection
    # Process Report (page 16) lists the 4 module names; fill them in here
    # so the Capacity tab isn't empty. Session 5 (if present) defaults to a
    # wrap-up label.
    UPWORK_TOPICS = {
        "1": "Module 1: How Upwork Agencies Work — Individual vs Agency, plans & pricing, roles & economics",
        "2": "Module 2: Building an Agency — Finding a niche, preparing the offer, kickstart, portfolio positioning",
        "3": "Module 3: Managing & Optimizing — Upwork KPIs, tracking tools, workflow design, weekly review",
        "4": "Module 4: Tools & Applications — Tech stack, inbound sales, profile-as-storefront, payment methods",
        "5": "Module 5: Wrap-up & Q&A",
    }

    # Track names match the Mentors.track values so the portal can
    # join attendance / budget by track.
    track_tabs = [
        ("Upwork", "Upwork Agency Building"),
        ("Business Development and tech s", "Business Development & Tech Sales"),
        ("Google Maps Lead Generation", "Google Maps Lead Generation"),
    ]
    for src_tab, canonical_track in track_tabs:
        if src_tab not in src.sheetnames:
            continue
        rows = src[src_tab].iter_rows(values_only=True)
        header = next(rows)
        # Discover positions by header name (some tabs have Date/Topic, some don't).
        idx = {str(h or "").strip().lower(): i for i, h in enumerate(header)}
        i_num   = idx.get("session #", 0)
        i_date  = idx.get("date")
        i_topic = idx.get("topic")
        i_rec   = idx.get("session recordings")
        i_pass  = idx.get("passcode")
        i_curr  = idx.get("curriculum") or idx.get("trainng plan") or idx.get("training plan")
        for r in rows:
            session_num = _str(r[i_num]) if i_num is not None and len(r) > i_num else ""
            if not session_num:
                continue
            date = _str(r[i_date]) if i_date is not None and len(r) > i_date else ""
            topic = _str(r[i_topic]) if i_topic is not None and len(r) > i_topic else ""
            recording = _str(r[i_rec]) if i_rec is not None and len(r) > i_rec else ""
            passcode = _str(r[i_pass]) if i_pass is not None and len(r) > i_pass else ""
            curriculum = _str(r[i_curr]) if i_curr is not None and len(r) > i_curr else ""
            # Upwork tab has no Topic column; populate from the PDF modules.
            if not topic and src_tab == "Upwork":
                topic = UPWORK_TOPICS.get(session_num, "")
            # Strip the noisy "Passcode: " prefix when present.
            if passcode.lower().startswith("passcode:"):
                passcode = passcode.split(":", 1)[1].strip()
            sid = f"{canonical_track[:3].lower()}-s{session_num}".replace(" ", "")
            ws.cell(row=row, column=1,  value=sid)
            ws.cell(row=row, column=2,  value=canonical_track)
            ws.cell(row=row, column=3,  value=session_num)
            ws.cell(row=row, column=4,  value=date)
            ws.cell(row=row, column=5,  value=topic)
            ws.cell(row=row, column=6,  value=recording)
            ws.cell(row=row, column=7,  value=passcode)
            ws.cell(row=row, column=8,  value=curriculum)
            # Default hours: 2 per session (typical training cadence). Admins
            # can edit once actual durations are confirmed.
            ws.cell(row=row, column=9,  value="2")
            ws.cell(row=row, column=10, value="Completed" if recording else "Scheduled")
            ws.cell(row=row, column=12, value=now)
            ws.cell(row=row, column=13, value="seeder")
            row += 1
            n += 1
    return n


def _seed_decisions(wb, master_rows):
    """Build Final Decisions. Heuristic:
       In/Out = "In"        -> Admitted
       Attended starts Yes-Withdrew -> Withdrew
       Track assigned set, In/Out = "Out", Attended = N/A -> Waitlisted (interview cut)
       No assigned track, KF = No   -> Waitlisted (S1 killing factor)
       Otherwise -> blank (admin decides explicitly)."""
    ws = wb["Final Decisions"]
    row = 2
    n = 0
    today = datetime.utcnow().strftime("%Y-%m-%d")
    for r in master_rows:
        email = _str(r[0])
        name  = _str(r[1])
        if not email and not name:
            continue
        track_assigned = _normalize_track(_str(r[12]))
        kf_canon, _ = _normalize_kf(_str(r[16]))
        in_out = _str(r[17]).lower()
        attended = _str(r[23]).lower()

        flsm = _str(r[26])
        fls  = _str(r[24])
        sms  = _str(r[25])
        score = flsm or fls or sms or ""

        if "withdrew" in attended:
            decision = "Withdrew"
        elif in_out == "in":
            decision = "Admitted"
        elif track_assigned and in_out == "out":
            decision = "Waitlisted"
        elif not track_assigned and kf_canon == "Fail":
            decision = "Waitlisted"
        else:
            decision = ""

        ws.cell(row=row, column=1, value=_applicant_id(email) or f"EB-row{row}")
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=email)
        ws.cell(row=row, column=4, value=track_assigned)
        ws.cell(row=row, column=5, value=score)
        ws.cell(row=row, column=6, value=decision)
        if decision:
            ws.cell(row=row, column=7, value=today)
            ws.cell(row=row, column=8, value="seeder")
        row += 1
        n += 1
    return n


def run():
    if not RESPONSES_XLSX.exists():
        raise SystemExit(f"Missing source file: {RESPONSES_XLSX}")
    if not TOP_XLSX.exists():
        raise SystemExit(f"Missing source file: {TOP_XLSX}")
    if not CAPACITY_XLSX.exists():
        raise SystemExit(f"Missing source file: {CAPACITY_XLSX}")

    # Rebuild template so the schema is current.
    build_template()
    out_path = Path(__file__).resolve().parents[1] / "out" / FILENAME

    print(f"Loading template: {out_path}")
    wb = load_workbook(out_path)

    src = load_workbook(RESPONSES_XLSX, read_only=True, data_only=True)
    master = list(src["Master Data"].iter_rows(values_only=True))
    master_header, master_rows = master[0], master[1:]
    print(f"Master Data: {len(master_rows)} rows")

    s1 = list(src["S1|| Killing Factor"].iter_rows(values_only=True))[1:]
    s2 = list(src["S2|| Tracks Sorting"].iter_rows(values_only=True))[1:]
    print(f"S1: {len(s1)} rows · S2: {len(s2)} rows")

    fla_rows = list(src["Interview Scoring - FLA"].iter_rows(values_only=True))[1:]
    sm_rows  = list(src["Interview Scoring - SM"].iter_rows(values_only=True))[1:]
    # Trim trailing empty rows
    fla_rows = [r for r in fla_rows if any(c is not None and str(c).strip() for c in r)]
    sm_rows  = [r for r in sm_rows  if any(c is not None and str(c).strip() for c in r)]
    print(f"Interview FL: {len(fla_rows)} rows · Interview SM: {len(sm_rows)} rows")

    seeded = {}
    seeded["applicants"] = _seed_applicants(wb, master_rows)
    seeded["s1"]         = _seed_stage1(wb, s1)
    seeded["s2"]         = _seed_stage2(wb, s2)
    seeded["response"]   = _seed_response_scoring(wb, src)
    seeded["int_fl"]     = _seed_interview_fl(wb, fla_rows)
    seeded["int_sm"]     = _seed_interview_sm(wb, sm_rows)
    seeded["rubrics"]    = _seed_rubrics(wb)
    seeded["mentors"]    = _seed_mentors(wb)
    seeded["sessions"]   = _seed_sessions(wb)
    seeded["top"]        = _seed_top_performers(wb, master_rows)
    seeded["decisions"]  = _seed_decisions(wb, master_rows)

    wb.save(out_path)

    print()
    print("Seeded:")
    for k, v in seeded.items():
        print(f"  {k:12s}  {v}")
    print(f"Output: {out_path}")
    return seeded


if __name__ == "__main__":
    run()
