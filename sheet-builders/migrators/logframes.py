"""Consolidate 2026 Targets Dutch SIDA.xlsx and Dutch SIDA Companies Logframes.xlsx
into a single E3 — Logframes workbook.

Prefers the richer row set from each source. Tabs:
  Dutch Log Frame, SIDA TechRise Log Frame, Program Budget, Monthly Budget by LIN Code, Lookups.
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook

from gsg_sheets.brand import BRAND
from gsg_sheets.layout import set_column_widths, tab_color, write_header
from gsg_sheets.styles import apply_style, cell_style, freeze_header


LOGFRAMES_PATH = "../Dutch SIDA Companies Logframes.xlsx"
TARGETS_PATH = "../2026 Targets Dutch SIDA.xlsx"
OUTPUT_NAME = "E3 - Logframes.xlsx"


def _copy_tab(src_ws, dst_wb, new_name, color=None):
    new_ws = dst_wb.create_sheet(new_name)
    if color:
        tab_color(new_ws, color)
    max_col = src_ws.max_column
    max_row = src_ws.max_row
    if max_col == 0:
        return new_ws

    headers = [src_ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
    write_header(new_ws, headers)
    freeze_header(new_ws)
    style = cell_style()
    for r in range(2, max_row + 1):
        vals = [src_ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if all(v in (None, "") for v in vals):
            continue
        for c_idx, v in enumerate(vals, start=1):
            cell = new_ws.cell(row=r, column=c_idx, value=v)
            apply_style(cell, style)
    set_column_widths(new_ws, {i: 22 for i in range(1, max_col + 1)})
    return new_ws


def run():
    root = Path(__file__).resolve().parents[1]
    logframes = load_workbook((root / LOGFRAMES_PATH).resolve(), data_only=True)
    targets = load_workbook((root / TARGETS_PATH).resolve(), data_only=True)

    out = Workbook()
    out.remove(out.active)

    # Prefer targets file for Dutch Log Frame (richer rows per analysis), logframes file for SIDA.
    if "Dutch Log Frame (Target only)" in targets.sheetnames:
        _copy_tab(targets["Dutch Log Frame (Target only)"], out, "Dutch Log Frame", BRAND["red"])
    elif "Dutch Log Frame (Target only)" in logframes.sheetnames:
        _copy_tab(logframes["Dutch Log Frame (Target only)"], out, "Dutch Log Frame", BRAND["red"])

    if "New SIDA (TechRise) Log Frame" in logframes.sheetnames:
        _copy_tab(logframes["New SIDA (TechRise) Log Frame"], out, "SIDA TechRise Log Frame", BRAND["orange"])
    elif "New SIDA (TechRise) Log Frame" in targets.sheetnames:
        _copy_tab(targets["New SIDA (TechRise) Log Frame"], out, "SIDA TechRise Log Frame", BRAND["orange"])

    # Copy remaining budget tabs from targets file if present.
    for name in targets.sheetnames:
        if name in ("Dutch Log Frame (Target only)", "New SIDA (TechRise) Log Frame"):
            continue
        if name in ("Dutch Legend", "New Sida Legend"):
            continue
        _copy_tab(targets[name], out, name, BRAND["teal"])

    dst = root / "out" / OUTPUT_NAME
    out.save(dst)
    print(f"Logframes consolidated to {dst}")
    return str(dst)


if __name__ == "__main__":
    run()
