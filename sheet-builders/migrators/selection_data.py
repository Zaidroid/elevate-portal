"""Tidy migrator for Elevate 3.0 Selection Data.xlsx.

Rather than regenerating the workbook from scratch (already consumed by
selection-tool), this:
  1. Copies all non-empty tabs into a new E3 — Selection Data.xlsx.
  2. Adds a company_id column on Source Data, 1st Filtration, Doc Reviews,
     Company Needs, Committee Interviews, Shortlists, Selection Votes so all
     cross-tab and cross-workbook joins key on the canonical id.
  3. Activates the empty Final Cohort tab with a header row matching Source
     Data plus a selection_decision column.
  4. Flags the Sellenvo vs. Inspire IT Solutions duplicate on the relevant
     row with a note.
"""

from pathlib import Path

from openpyxl import Workbook, load_workbook

from gsg_sheets.brand import BRAND
from gsg_sheets.layout import set_column_widths, tab_color, write_header
from gsg_sheets.styles import apply_style, cell_style, editable_style, formula_style, freeze_header


LEGACY_RELPATH = "../Elevate 3.0 Selection Data.xlsx"
OUTPUT_NAME = "E3 - Selection Data.xlsx"

ID_INJECT_TABS = {
    "Source Data",
    "1st Filtration",
    "Doc Reviews",
    "Company Needs",
    "Committee Interviews",
    "Shortlists",
    "Selection Votes",
    "Interview Assessments",
    "Interview Discussion",
    "ElevateBridge Assessments",
    "grey area",
    "Scoring Matrix",
    "Custom Presets",
    "Additional Factors Filtration 1",
}

DUPLICATE_FLAG_COMPANIES = {"Sellenvo", "Inspire IT Solutions"}


def run():
    src = Path(__file__).resolve().parents[1] / LEGACY_RELPATH
    dst = Path(__file__).resolve().parents[1] / "out" / OUTPUT_NAME

    legacy = load_workbook(src.resolve(), data_only=True)
    out = Workbook()
    out.remove(out.active)

    style_f = formula_style()
    style_c = cell_style()

    for tab_name in legacy.sheetnames:
        src_ws = legacy[tab_name]
        max_col = src_ws.max_column
        max_row = src_ws.max_row
        if max_col == 0:
            continue

        new_ws = out.create_sheet(tab_name)
        tab_color(new_ws, BRAND["teal"])

        # Decide if we're injecting a company_id column in position 1.
        inject = tab_name in ID_INJECT_TABS

        # Header row
        header = [src_ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]
        if inject:
            write_header(new_ws, ["company_id", *header])
        else:
            write_header(new_ws, header)
        freeze_header(new_ws)

        # Data rows
        for r in range(2, max_row + 1):
            row_vals = [src_ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            if all(v in (None, "") for v in row_vals):
                continue
            if inject:
                # company_id formula: look up in Source Data by company_name (column
                # varies per tab; we emit a formula that pulls the first string cell
                # from this row and resolves through Source Data. If Source Data is
                # this very tab, inject a sequential id instead.)
                new_row = r
                if tab_name == "Source Data":
                    id_formula = f'=IF(B{new_row}<>"", "E3-"&TEXT(ROW()-1,"0000"), "")'
                else:
                    id_formula = f'=IFERROR(INDEX(\'Source Data\'!A:A,MATCH(B{new_row},\'Source Data\'!B:B,0)),"")'
                id_cell = new_ws.cell(row=new_row, column=1, value=id_formula)
                apply_style(id_cell, style_f)
                for col_idx, val in enumerate(row_vals, start=2):
                    cell = new_ws.cell(row=new_row, column=col_idx, value=val)
                    apply_style(cell, style_c)
            else:
                for col_idx, val in enumerate(row_vals, start=1):
                    cell = new_ws.cell(row=r, column=col_idx, value=val)
                    apply_style(cell, style_c)

        # Autosize columns roughly
        cols = max_col + (1 if inject else 0)
        set_column_widths(new_ws, {i: 20 for i in range(1, cols + 1)})

    # Activate the empty Final Cohort tab with a useful schema.
    if "Final Cohort" in out.sheetnames:
        out.remove(out["Final Cohort"])
    fc = out.create_sheet("Final Cohort")
    tab_color(fc, BRAND["red"])
    write_header(
        fc,
        [
            "company_id",
            "company_name",
            "cohort",
            "fund_code",
            "selection_date",
            "primary_interventions",
            "selection_decision",
            "decision_rationale",
            "committee_vote_count",
            "drive_folder_url",
            "notes",
        ],
    )
    freeze_header(fc)
    set_column_widths(fc, {1: 12, 2: 32, 3: 8, 4: 18, 5: 14, 6: 32, 7: 18, 8: 40, 9: 14, 10: 36, 11: 40})
    style_e = editable_style()
    for r in range(2, 200):
        for c in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11]:
            apply_style(fc.cell(row=r, column=c), style_e)
    # company_id formula
    for r in range(2, 200):
        cell = fc.cell(row=r, column=1, value=f'=IFERROR(INDEX(\'Source Data\'!A:A,MATCH(B{r},\'Source Data\'!B:B,0)),"")')
        apply_style(cell, style_f)

    # Flag duplicate company for Sellenvo / Inspire IT Solutions in Source Data notes column.
    if "Source Data" in out.sheetnames:
        sd = out["Source Data"]
        hdr = [sd.cell(row=1, column=c).value for c in range(1, sd.max_column + 1)]
        # Find a name-like column (inject shifted by 1).
        name_col = 2  # after injection, column 2 is the original first column
        notes_col = sd.max_column + 1
        sd.cell(row=1, column=notes_col, value="migration_notes")
        for r in range(2, sd.max_row + 1):
            val = sd.cell(row=r, column=name_col).value
            if not val:
                continue
            for keyword in DUPLICATE_FLAG_COMPANIES:
                if keyword.lower() in str(val).lower():
                    sd.cell(row=r, column=notes_col, value="Potential duplicate: review Sellenvo vs. Inspire IT Solutions before assigning interventions")

    out.save(dst)
    print(f"Selection Data migrated to {dst}")
    return str(dst)


if __name__ == "__main__":
    run()
