"""Port 2026 YTD procurement activities from the legacy monthly plan.

Source: GSG WB Procurement Plan 2025-2026.xlsx
        Tabs: Jan 2026, Feb 2026, Mar 2026, April 2026, May 2026

Target: E3 - Procurement Plan.xlsx
        Jan/Feb/Mar 2026 -> Q1 2026
        April/May 2026   -> Q2 2026

Two header variants exist in the source:
    - Jan 2026: Number | Activity | Office Code | GL Account | Fund | LIN | Item | Unit | Qty | Unit Cost | Total | Delivery Date | Type | PR Deadline | Local/Intl | Program POC | PR # | Procurement POC | Status | Completion | Notes
    - Feb/Mar/April/May 2026: Number | Activity | PR# If Any | Office Code | GL Account | Fund | LIN | Item | Unit | Qty | Unit Cost | Total | Delivery Date | Type | PR Deadline | Local/Intl | Program POC | Procurement POC | Status | (Notes)

Formula columns in the target (pr_id, total_cost_usd, threshold_class,
sla_working_days, pr_deadline) are left alone; we only populate user-editable
cells so the existing formulas remain live.

Usage:
    python -m migrators.procurement_ytd
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from gsg_sheets import save_workbook


ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "sheet-builders" / "out" / "E3 - Procurement Plan.xlsx"
SRC = ROOT / "GSG WB Procurement Plan 2025-2026.xlsx"

TAB_TO_QUARTER = {
    "Jan 2026": "Q1 2026",
    "Feb 2026": "Q1 2026",
    "Mar 2026": "Q1 2026",
    "April 2026": "Q2 2026",
    "May 2026": "Q2 2026",
}

# Target column indexes (1-based) in the E3 Procurement Plan quarter tab.
TGT = {
    "activity": 2,
    "intervention_type": 3,
    "company_id": 4,
    "office_code": 5,
    "gl_account": 6,
    "fund_code": 7,
    "lin_code": 8,
    "item_description": 9,
    "unit": 10,
    "qty": 11,
    "unit_cost_usd": 12,
    "target_award_date": 16,
    "pr_submit_date": 17,
    "local_international": 19,
    "requester_email": 20,
    "status": 21,
    "procurement_contact": 22,
    "notes": 23,
    "updated_at": 24,
    "updated_by": 25,
}


def _clean(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].replace("-", "").isdigit():
        s = s[:-2]
    return s


def _num(v):
    """Return a numeric value if the cell holds one, else empty string."""
    if v is None or v == "":
        return ""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return ""


def _status_map(raw: str) -> str:
    """Map legacy status strings to E3 procurement_status taxonomy."""
    if not raw:
        return "Draft"
    low = raw.strip().lower()
    if "not started" in low or "draft" in low:
        return "Draft"
    if "submit" in low:
        return "Submitted"
    if "review" in low or "under" in low:
        return "Under Review"
    if "award" in low:
        return "Awarded"
    if "deliver" in low or "complete" in low or "done" in low:
        return "Delivered"
    if "cancel" in low:
        return "Cancelled"
    return "Draft"


def _read_month(wb, tab: str) -> list[dict]:
    """Parse a source monthly tab. Returns a list of activity dicts.

    Auto-detects the two known header variants by checking whether
    column 3 on row 1 is "Office Code" (Jan layout) or "PR# If Any" style.
    """
    ws = wb[tab]
    hdr_c3 = _clean(ws.cell(row=1, column=3).value)
    jan_layout = hdr_c3.lower().startswith("office")

    # Column indexes per layout (1-based in source).
    if jan_layout:
        C = {
            "activity": 2, "office_code": 3, "gl_account": 4, "fund": 5,
            "lin": 6, "item": 7, "unit": 8, "qty": 9, "unit_cost": 10,
            "total": 11, "target_date": 12, "type": 13, "pr_deadline": 14,
            "local_intl": 15, "program_poc": 16, "pr_num": 17,
            "procurement_poc": 18, "status": 19, "notes": 21,
        }
    else:
        C = {
            "activity": 2, "pr_num": 3, "office_code": 4, "gl_account": 5,
            "fund": 6, "lin": 7, "item": 8, "unit": 9, "qty": 10,
            "unit_cost": 11, "total": 12, "target_date": 13, "type": 14,
            "pr_deadline": 15, "local_intl": 16, "program_poc": 17,
            "procurement_poc": 18, "status": 19, "notes": 20,
        }

    rows: list[dict] = []
    for r in range(5, ws.max_row + 1):
        activity = _clean(ws.cell(row=r, column=C["activity"]).value)
        if not activity:
            continue
        fund_raw = _clean(ws.cell(row=r, column=C["fund"]).value)
        notes_bits = [
            f"Legacy from {tab}",
            f"Type: {_clean(ws.cell(row=r, column=C['type']).value)}" if _clean(ws.cell(row=r, column=C['type']).value) else "",
            f"PR #: {_clean(ws.cell(row=r, column=C['pr_num']).value)}" if _clean(ws.cell(row=r, column=C['pr_num']).value) else "",
            f"Legacy deadline: {_clean(ws.cell(row=r, column=C['pr_deadline']).value)}" if _clean(ws.cell(row=r, column=C['pr_deadline']).value) else "",
            _clean(ws.cell(row=r, column=C['notes']).value),
        ]
        rows.append({
            "activity": activity,
            "office_code": _clean(ws.cell(row=r, column=C["office_code"]).value),
            "gl_account": _clean(ws.cell(row=r, column=C["gl_account"]).value),
            "fund_code": fund_raw,
            "lin_code": _clean(ws.cell(row=r, column=C["lin"]).value),
            "item_description": _clean(ws.cell(row=r, column=C["item"]).value),
            "unit": _clean(ws.cell(row=r, column=C["unit"]).value),
            "qty": _num(ws.cell(row=r, column=C["qty"]).value),
            "unit_cost_usd": _num(ws.cell(row=r, column=C["unit_cost"]).value),
            "target_award_date": _clean(ws.cell(row=r, column=C["target_date"]).value),
            "local_international": _clean(ws.cell(row=r, column=C["local_intl"]).value),
            "requester_email": _clean(ws.cell(row=r, column=C["program_poc"]).value),
            "procurement_contact": _clean(ws.cell(row=r, column=C["procurement_poc"]).value) or "Donia Shadeed",
            "status": _status_map(_clean(ws.cell(row=r, column=C["status"]).value)),
            "notes": " | ".join(b for b in notes_bits if b),
            "source_tab": tab,
        })
    return rows


def _find_first_empty_row(ws, key_col: int = TGT["activity"], start: int = 2) -> int:
    r = start
    while ws.cell(row=r, column=key_col).value:
        r += 1
    return r


def populate():
    if not OUT.exists():
        raise FileNotFoundError(
            f"{OUT} missing. Run `python -m builders.procurement_plan` first."
        )

    src = load_workbook(SRC, read_only=False, data_only=True)
    by_quarter: dict[str, list[dict]] = {"Q1 2026": [], "Q2 2026": []}
    for tab, quarter in TAB_TO_QUARTER.items():
        rows = _read_month(src, tab)
        by_quarter[quarter].extend(rows)
        print(f"[procurement_ytd] {tab}: {len(rows)} rows -> {quarter}")
    src.close()

    tgt = load_workbook(OUT)
    now = datetime.utcnow().strftime("%Y-%m-%d")

    for quarter, rows in by_quarter.items():
        ws = tgt[quarter]
        write_row = _find_first_empty_row(ws, TGT["activity"])
        for rec in rows:
            ws.cell(row=write_row, column=TGT["activity"]).value = rec["activity"]
            ws.cell(row=write_row, column=TGT["office_code"]).value = rec["office_code"]
            ws.cell(row=write_row, column=TGT["gl_account"]).value = rec["gl_account"]
            ws.cell(row=write_row, column=TGT["fund_code"]).value = rec["fund_code"]
            ws.cell(row=write_row, column=TGT["lin_code"]).value = rec["lin_code"]
            ws.cell(row=write_row, column=TGT["item_description"]).value = rec["item_description"]
            ws.cell(row=write_row, column=TGT["unit"]).value = rec["unit"]
            if rec["qty"] != "":
                ws.cell(row=write_row, column=TGT["qty"]).value = rec["qty"]
            if rec["unit_cost_usd"] != "":
                ws.cell(row=write_row, column=TGT["unit_cost_usd"]).value = rec["unit_cost_usd"]
            ws.cell(row=write_row, column=TGT["target_award_date"]).value = rec["target_award_date"]
            ws.cell(row=write_row, column=TGT["local_international"]).value = rec["local_international"]
            ws.cell(row=write_row, column=TGT["requester_email"]).value = rec["requester_email"]
            ws.cell(row=write_row, column=TGT["status"]).value = rec["status"]
            ws.cell(row=write_row, column=TGT["procurement_contact"]).value = rec["procurement_contact"]
            ws.cell(row=write_row, column=TGT["notes"]).value = rec["notes"]
            ws.cell(row=write_row, column=TGT["updated_at"]).value = now
            ws.cell(row=write_row, column=TGT["updated_by"]).value = "migrator"
            write_row += 1
        print(f"[procurement_ytd] {quarter}: wrote {len(rows)} rows")

    save_workbook(tgt, OUT.name)


if __name__ == "__main__":
    populate()
