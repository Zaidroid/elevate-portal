"""Port Doc Tracker 2025.csv into E3 - Docs and Agreements > Agreements.

Source columns:
    Number | Company | Intervention | Agreement | COC | CL | RPS | Ref | POC | CL Ref

One source row becomes one Agreements row (agreement_type = MJPSA), with the
state of COC, CL, RPS carried in the notes column so nothing is lost and a
future sweep can split commitment letters into their own tab.

FK resolution: company_id best-effort matched against E3 Companies Master by
normalized company_name.

Usage:
    python -m migrators.docs_tracker
"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from gsg_sheets import save_workbook


ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "sheet-builders" / "out" / "E3 - Docs and Agreements.xlsx"
COMPANIES_WB = ROOT / "sheet-builders" / "out" / "E3 - Companies Master.xlsx"
SRC = ROOT / "Doc Tracker 2025.csv"


def _norm(v: str) -> str:
    return (v or "").strip().lower()


def _clean(v: str) -> str:
    return (v or "").strip()


def _read_rows() -> list[dict]:
    rows: list[dict] = []
    with open(SRC, newline="", encoding="utf-8") as fh:
        reader = csv.reader(fh)
        for raw in reader:
            # skip blank lines and the two-line preamble before the real header
            if not raw or not raw[0].strip():
                continue
            if raw[0].strip() == "Number" or not raw[0].strip().isdigit():
                continue
            # pad to 10 columns
            raw = raw + [""] * (10 - len(raw))
            rows.append({
                "number": raw[0].strip(),
                "company_name": _clean(raw[1]),
                "intervention": _clean(raw[2]),
                "agreement_status": _clean(raw[3]),
                "coc_status": _clean(raw[4]),
                "cl_status": _clean(raw[5]),
                "rps_status": _clean(raw[6]),
                "ref": _clean(raw[7]),
                "poc": _clean(raw[8]),
                "cl_ref": _clean(raw[9]),
            })
    return rows


def _read_companies_map() -> dict[str, str]:
    if not COMPANIES_WB.exists():
        return {}
    wb = load_workbook(COMPANIES_WB)
    ws = wb["Companies"]
    out: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        if not name:
            continue
        out[_norm(str(name))] = f"E3-{r - 1:04d}"
    wb.close()
    return out


def _resolve_status(raw: str) -> str:
    low = raw.lower()
    if "fully" in low:
        return "Executed"
    if "mc signed" in low or "gsg signed" in low:
        return "Signed"
    if "shared" in low:
        return "Sent"
    if "draft" in low:
        return "Drafted"
    if "countersigned" in low:
        return "Countersigned"
    if not raw:
        return "Drafted"
    return "Drafted"


def _build_notes(rec: dict) -> str:
    bits = [
        f"Source: Doc Tracker 2025 row {rec['number']}",
        f"Intervention: {rec['intervention']}" if rec["intervention"] else "",
        f"COC: {rec['coc_status']}" if rec["coc_status"] else "",
        f"CL: {rec['cl_status']}" if rec["cl_status"] else "",
        f"RPS: {rec['rps_status']}" if rec["rps_status"] else "",
        f"Ref: {rec['ref']}" if rec["ref"] else "",
        f"POC: {rec['poc']}" if rec["poc"] else "",
        f"CL Ref: {rec['cl_ref']}" if rec["cl_ref"] else "",
    ]
    return " | ".join(b for b in bits if b)


def populate():
    if not OUT.exists():
        raise FileNotFoundError(
            f"{OUT} missing. Run `python -m builders.docs_agreements` first."
        )

    rows = _read_rows()
    print(f"[docs_tracker] parsed {len(rows)} rows from Doc Tracker 2025.csv")

    companies = _read_companies_map()
    resolved = 0
    for rec in rows:
        rec["company_id"] = companies.get(_norm(rec["company_name"]), "")
        if rec["company_id"]:
            resolved += 1
    print(f"[docs_tracker] resolved {resolved}/{len(rows)} company_id FKs")

    tgt = load_workbook(OUT)
    ws = tgt["Agreements"]

    # Target columns: 1 agreement_id (formula), 2 company_id, 3 company_name,
    # 4 agreement_type, 5 signed_date, 6 signatory_name, 7 signatory_title,
    # 8 gsg_signatory, 9 drive_url, 10 status, 11 related_intervention,
    # 12 assignment_id, 13 notes, 14 updated_at, 15 updated_by
    now = datetime.utcnow().strftime("%Y-%m-%d")

    # find first empty row in Agreements (company_name column 3)
    r = 2
    while ws.cell(row=r, column=3).value:
        r += 1

    for rec in rows:
        ws.cell(row=r, column=2).value = rec["company_id"]
        ws.cell(row=r, column=3).value = rec["company_name"]
        ws.cell(row=r, column=4).value = "MJPSA"
        ws.cell(row=r, column=8).value = rec["poc"]
        ws.cell(row=r, column=10).value = _resolve_status(rec["agreement_status"])
        ws.cell(row=r, column=11).value = rec["intervention"]
        ws.cell(row=r, column=13).value = _build_notes(rec)
        ws.cell(row=r, column=14).value = now
        ws.cell(row=r, column=15).value = "migrator"
        r += 1

    save_workbook(tgt, OUT.name)
    print(f"[docs_tracker] wrote {len(rows)} rows to Agreements tab")


if __name__ == "__main__":
    populate()
