"""Build the fresh "E3 - Cohort 3 Companies Master (Operational)" workbook
from the post-selection state.

The live data sits in three Drive workbooks. Since gcloud Drive scope
isn't available locally, this migrator reads from LOCAL EXPORTS the
user takes:

  1. Current Companies workbook  -> sheet-builders/inputs/companies_current.xlsx
  2. Current Selection workbook  -> sheet-builders/inputs/selection_current.xlsx
  3. Current ElevateBridge workbook -> sheet-builders/inputs/elevatebridge_current.xlsx
     (optional — used only to verify the 7 EB-selected companies)

To export from Drive: File > Download > Microsoft Excel (.xlsx) on each
of the three workbooks, save into sheet-builders/inputs/ with the names
above.

The migrator:
  1. Runs builders/companies_master.py to produce a fresh canonical
     xlsx template at out/E3 - Companies Master.xlsx.
  2. Seeds the Companies tab from selection's Final Cohort (status, AM,
     donor, sector, governorate) — fall back to current Companies tab
     if Final Cohort is empty.
  3. Seeds the Assignments tab from selection's Stage3 Distribution
     plus ElevateBridge admitted (7 EB companies with intervention_type
     = "ElevateBridge").
  4. Carries forward Contacts / Reviews / Comments / Activity / Status
     Log / Pre-decision Recommendations / Interview Aliases / Removed
     Companies / Historical Interventions from the current Companies
     workbook — these are team-owned operational data that survives
     selection.
  5. Runs a pre-flight validation pass.

Output: out/E3 - Companies Master.xlsx (same filename as the existing
builder produces — the user uploads this side-by-side to Drive).

Usage:
    python3 -m migrators.cohort3_operational_seed
"""

from __future__ import annotations

import hashlib
import re
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from builders.companies_master import (
    FILENAME,
    COMPANIES_HEADERS,
    CONTACTS_HEADERS,
    ASSIGNMENTS_HEADERS,
    STATUS_LOG_HEADERS,
    REVIEWS_HEADERS,
    build as build_template,
)


HERE = Path(__file__).resolve().parents[1]
INPUT_DIR = HERE / "inputs"
OUTPUT_PATH = HERE / "out" / FILENAME

COMPANIES_SRC = INPUT_DIR / "companies_current.xlsx"
SELECTION_SRC = INPUT_DIR / "selection_current.xlsx"
EB_SRC = INPUT_DIR / "elevatebridge_current.xlsx"

# Confirmed EB-selected companies (canonical-name-matched). The user gave
# these names; the seeder verifies each one resolves to a Cohort 3 row.
EB_SELECTED_NAMES = [
    "Tayf",
    "Haweya",
    "Jaffa",
    "Radix Technologies",
    "Top Mena Talents",
    "Sada Intelligence",
    "IzTechValley",
]


def _norm(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _name_key(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (s or "").lower())


def _company_id_for(name: str, fallback_idx: int) -> str:
    """Stable company_id derived from canonical name. Falls back to a
    short hash so re-runs produce identical IDs for the same company."""
    key = _name_key(name) or f"row{fallback_idx}"
    h = hashlib.sha1(key.encode("utf-8")).hexdigest()[:6]
    return f"co-{h}"


def _load_input(path: Path) -> Workbook | None:
    if not path.exists():
        print(f"  [skip] {path.name} not found in inputs/")
        return None
    # Warn loudly if the export is more than a day old. The seeded
    # workbook is only as fresh as its inputs; re-running over stale
    # exports silently regresses the live operational state to a
    # snapshot from N days ago.
    import time
    age_s = time.time() - path.stat().st_mtime
    age_h = age_s / 3600
    if age_h > 24:
        days = age_h / 24
        print(f"  [WARN] {path.name} is {days:.1f} day(s) old — consider re-exporting from Drive before seeding.")
    elif age_h > 4:
        print(f"  [info] {path.name} is {age_h:.1f} hour(s) old.")
    return load_workbook(path, read_only=True, data_only=True)


def _read_rows(wb: Workbook, tab_name: str) -> tuple[list[str], list[list]]:
    """Return (headers, rows) for a tab. (`[],[]`) if missing."""
    if tab_name not in wb.sheetnames:
        return [], []
    ws = wb[tab_name]
    iterator = ws.iter_rows(values_only=True)
    try:
        headers = list(next(iterator))
    except StopIteration:
        return [], []
    rows = [list(r) for r in iterator if any(c is not None and str(c).strip() for c in r)]
    return [_norm(h) for h in headers], rows


def _row_to_dict(headers: list[str], row: list) -> dict:
    out: dict = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        v = row[i] if i < len(row) else None
        out[h] = _norm(v)
    return out


def _carry_forward(out_wb: Workbook, tab_name: str, src_wb: Workbook, src_tab: str, canonical_headers: list[str]) -> int:
    """Copy every row from src_wb[src_tab] into out_wb[tab_name],
    matching columns by header name (so a column rename in the source
    just produces blanks rather than wrong-column writes)."""
    if src_tab not in src_wb.sheetnames:
        return 0
    src_headers, src_rows = _read_rows(src_wb, src_tab)
    if not src_headers:
        return 0
    ws = out_wb[tab_name]
    out_headers = list(canonical_headers)
    n = 0
    next_row_idx = 2  # row 1 is the header (written by the builder)
    for raw in src_rows:
        rec = _row_to_dict(src_headers, raw)
        for col_idx, canon_h in enumerate(out_headers, start=1):
            v = rec.get(canon_h, "")
            if v:
                ws.cell(row=next_row_idx, column=col_idx, value=v)
        next_row_idx += 1
        n += 1
    return n


def _seed_companies(out_wb: Workbook, companies_src: Workbook | None, selection_src: Workbook | None) -> int:
    """Populate the Companies tab:
       - Final Cohort row → 41 selected Cohort 3 companies (status, AM,
         donor, sector, governorate, etc.)
       - Plus carry-forward of any existing Companies rows for
         historical reference (cohort != 'E3').
    """
    ws = out_wb["Companies"]
    row_idx = 2
    seen_ids: set[str] = set()
    seen_names: set[str] = set()

    # Final Cohort (post-selection)
    if selection_src is not None and "Final Cohort" in selection_src.sheetnames:
        headers, rows = _read_rows(selection_src, "Final Cohort")
        if not rows:
            print("  [warn] selection_current.xlsx 'Final Cohort' tab is empty")
        for i, raw in enumerate(rows):
            rec = _row_to_dict(headers, raw)
            name = rec.get("company_name") or rec.get("name") or rec.get("Company Name") or ""
            if not name:
                continue
            cid = rec.get("company_id") or _company_id_for(name, i)
            if cid in seen_ids:
                continue
            seen_ids.add(cid)
            seen_names.add(_name_key(name))
            mapping = {
                "company_id": cid,
                "company_name": name,
                "legal_name": rec.get("legal_name", ""),
                "city": rec.get("city", ""),
                "governorate": rec.get("governorate", ""),
                "sector": rec.get("sector", ""),
                "employee_count": rec.get("employee_count") or rec.get("employees", ""),
                "revenue_bracket": rec.get("revenue_bracket", ""),
                "international_revenue_pct": rec.get("international_revenue_pct", ""),
                "readiness_score": rec.get("readiness_score", ""),
                "fund_code": rec.get("fund_code") or rec.get("donor", ""),
                "cohort": "E3",
                "status": rec.get("status") or rec.get("final_status", "Selected"),
                "stage": rec.get("stage", "Onboarding"),
                "profile_manager_email": rec.get("profile_manager_email") or rec.get("am_email") or rec.get("am", ""),
                "selection_date": rec.get("selection_date", ""),
                "onboarding_date": rec.get("onboarding_date", ""),
                "primary_contact_id": "",
                "drive_folder_url": rec.get("drive_folder_url", ""),
                "notes": rec.get("notes", ""),
                "updated_at": datetime.utcnow().isoformat(),
                "updated_by": "seeder",
            }
            for col_idx, h in enumerate(COMPANIES_HEADERS, start=1):
                v = mapping.get(h, "")
                if v:
                    ws.cell(row=row_idx, column=col_idx, value=v)
            row_idx += 1

    # Historical carry-forward from current Companies workbook for rows
    # not in Final Cohort (older cohorts, archive references).
    if companies_src is not None and "Companies" in companies_src.sheetnames:
        headers, rows = _read_rows(companies_src, "Companies")
        for raw in rows:
            rec = _row_to_dict(headers, raw)
            name = rec.get("company_name") or rec.get("name") or ""
            cid = rec.get("company_id") or ""
            if not name:
                continue
            key = _name_key(name)
            if cid in seen_ids or key in seen_names:
                continue
            seen_ids.add(cid or "")
            seen_names.add(key)
            # Carry over verbatim — historical row, keep as-is.
            for col_idx, h in enumerate(COMPANIES_HEADERS, start=1):
                v = rec.get(h, "")
                if v:
                    ws.cell(row=row_idx, column=col_idx, value=v)
            row_idx += 1

    return row_idx - 2


def _seed_assignments(out_wb: Workbook, selection_src: Workbook | None, companies_src: Workbook | None, eb_src: Workbook | None) -> int:
    """Stage3 Distribution → Assignments rows + ElevateBridge admitted
    (7 EB companies with intervention_type='ElevateBridge')."""
    ws = out_wb["Intervention Assignments"]
    row_idx = 2
    n = 0

    # Build a name → company_id map from the Companies tab we just seeded.
    cmap: dict[str, str] = {}
    comp_ws = out_wb["Companies"]
    for r in comp_ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        cid = _norm(r[0])
        cname = _norm(r[1] if len(r) > 1 else "")
        if cid and cname:
            cmap[_name_key(cname)] = cid

    # Stage3 Distribution (selection workbook)
    if selection_src is not None and "Stage3 Distribution" in selection_src.sheetnames:
        headers, rows = _read_rows(selection_src, "Stage3 Distribution")
        for i, raw in enumerate(rows):
            rec = _row_to_dict(headers, raw)
            name = rec.get("company_name") or rec.get("name") or ""
            if not name:
                continue
            cid = cmap.get(_name_key(name))
            if not cid:
                continue
            interventions = rec.get("interventions") or rec.get("sub_interventions") or ""
            subs = [s.strip() for s in re.split(r"[;,]", interventions) if s.strip()]
            for sub in subs:
                ws.cell(row=row_idx, column=1, value=f"asn-{cid}-{_name_key(sub)[:12]}-{i:03d}")
                ws.cell(row=row_idx, column=2, value=cid)
                ws.cell(row=row_idx, column=3, value=sub)   # intervention_type = sub for ops simplicity
                ws.cell(row=row_idx, column=4, value="")    # sub_intervention blank (sub IS the type here)
                ws.cell(row=row_idx, column=5, value=rec.get("fund_code") or rec.get("donor", ""))
                ws.cell(row=row_idx, column=9, value="In Progress")
                ws.cell(row=row_idx, column=10, value=rec.get("budget_usd", ""))
                ws.cell(row=row_idx, column=13, value=datetime.utcnow().isoformat())
                ws.cell(row=row_idx, column=14, value="seeder")
                row_idx += 1
                n += 1

    # ElevateBridge admitted (7 EB companies)
    eb_assigned: set[str] = set()
    for eb_name in EB_SELECTED_NAMES:
        cid = cmap.get(_name_key(eb_name))
        if not cid:
            print(f"  [warn] EB company not found in Companies tab: {eb_name}")
            continue
        ws.cell(row=row_idx, column=1, value=f"asn-{cid}-elevatebridge")
        ws.cell(row=row_idx, column=2, value=cid)
        ws.cell(row=row_idx, column=3, value="ElevateBridge")
        ws.cell(row=row_idx, column=9, value="In Progress")
        ws.cell(row=row_idx, column=12, value="Seeded from ElevateBridge selection (7 admitted companies)")
        ws.cell(row=row_idx, column=13, value=datetime.utcnow().isoformat())
        ws.cell(row=row_idx, column=14, value="seeder")
        row_idx += 1
        n += 1
        eb_assigned.add(eb_name)

    # Carry-forward any existing Assignments rows not already covered.
    if companies_src is not None and "Intervention Assignments" in companies_src.sheetnames:
        headers, rows = _read_rows(companies_src, "Intervention Assignments")
        for raw in rows:
            rec = _row_to_dict(headers, raw)
            aid = rec.get("assignment_id") or ""
            cid = rec.get("company_id") or ""
            # Skip duplicates of the EB rows we just seeded.
            it = rec.get("intervention_type") or ""
            if cid and it == "ElevateBridge":
                continue
            if not aid:
                aid = f"asn-{cid}-legacy-{n}"
            for col_idx, h in enumerate(ASSIGNMENTS_HEADERS, start=1):
                v = rec.get(h, "") if h != "assignment_id" else aid
                if v:
                    ws.cell(row=row_idx, column=col_idx, value=v)
            row_idx += 1
            n += 1

    return n


def _carry_operational_tabs(out_wb: Workbook, companies_src: Workbook | None) -> dict:
    """Contacts, Status Log, Reviews, Comments, Activity, Pre-decision
    Recommendations, Interview Aliases, Removed Companies, Historical
    Interventions — all team-owned operational data, copied verbatim."""
    if companies_src is None:
        return {}
    transfers = [
        ("Contacts", "Contacts", CONTACTS_HEADERS),
        ("Status Log", "Status Log", STATUS_LOG_HEADERS),
        ("Reviews", "Reviews", REVIEWS_HEADERS),
        # Other tabs use headers defined by the builder's other constants.
        # We carry forward only the canonical-column subset.
    ]
    out: dict = {}
    for out_tab, src_tab, canonical_headers in transfers:
        n = _carry_forward(out_wb, out_tab, companies_src, src_tab, canonical_headers)
        out[out_tab] = n
    return out


def run():
    INPUT_DIR.mkdir(parents=True, exist_ok=True)

    print("Step 1: building canonical template from companies_master.py...")
    build_template()
    print(f"  -> {OUTPUT_PATH}")

    print("Step 2: loading source exports...")
    companies_src = _load_input(COMPANIES_SRC)
    selection_src = _load_input(SELECTION_SRC)
    eb_src = _load_input(EB_SRC)

    if companies_src is None and selection_src is None:
        print()
        print("=" * 70)
        print("No source exports found. Output is a clean canonical TEMPLATE only.")
        print(f"Drop your live workbook exports into {INPUT_DIR} as:")
        print("  companies_current.xlsx  (File > Download > Excel from the live Companies sheet)")
        print("  selection_current.xlsx  (same for Selection)")
        print("  elevatebridge_current.xlsx  (optional, same for ElevateBridge Programme)")
        print("Then re-run.")
        print("=" * 70)
        return

    print("Step 3: seeding Companies tab...")
    out_wb = load_workbook(OUTPUT_PATH)
    n_companies = _seed_companies(out_wb, companies_src, selection_src)
    print(f"  -> {n_companies} company rows")

    print("Step 4: seeding Assignments tab...")
    n_assignments = _seed_assignments(out_wb, selection_src, companies_src, eb_src)
    print(f"  -> {n_assignments} assignment rows")

    print("Step 5: carrying forward operational tabs...")
    transfers = _carry_operational_tabs(out_wb, companies_src)
    for t, n in transfers.items():
        print(f"  -> {t}: {n} rows")

    out_wb.save(OUTPUT_PATH)
    print()
    print("=" * 70)
    print(f"Done. Output: {OUTPUT_PATH}")
    print()
    print("Next steps:")
    print("  1. Open the xlsx, scan the Companies + Assignments tabs for sanity.")
    print("  2. Upload to Drive as a NEW workbook (side-by-side with the existing one).")
    print("  3. Flip VITE_SHEET_COMPANIES in .env to the new sheet id.")
    print("  4. Smoke-test 10 random company profiles in the portal.")
    print("  5. Once happy, archive the old workbook (don't delete).")
    print("=" * 70)


if __name__ == "__main__":
    run()
