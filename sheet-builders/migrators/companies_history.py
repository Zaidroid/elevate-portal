"""Populate E3 Companies Master with historical company references.

Pulls from the two legacy workbooks:
    - Companies Data Tracker (1).xlsx > 3. Company Profile  (~48 rows with POCs)
    - Companies Database 2025.xlsx > Companies 2024         (~116 rows with POCs)
    - Companies Database 2025.xlsx > Companies 2025         (~120 rows)
    - Companies Database 2025.xlsx > Not active-available   (~17 rows)

Dedupes by normalized company name. Writes to the Companies tab and
populates Contacts for rows that have POC info. Leaves cohort / status
blank so the data is clearly visible as reference material and does not
contaminate the Cohort 3 roster.

Usage:
    python -m migrators.companies_history
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from gsg_sheets import save_workbook
from gsg_sheets.styles import editable_style


ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "sheet-builders" / "out" / "E3 - Companies Master.xlsx"
SRC_TRACKER = ROOT / "Companies Data Tracker (1).xlsx"
SRC_DB = ROOT / "Companies Database 2025.xlsx"


def _norm(v) -> str:
    if v is None:
        return ""
    return str(v).strip().lower()


def _clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    # openpyxl returns floats for integer-valued cells; strip the .0
    if s.endswith(".0") and s[:-2].replace("-", "").isdigit():
        s = s[:-2]
    return s


def _read_company_profile() -> list[dict]:
    wb = load_workbook(SRC_TRACKER, read_only=False, data_only=True)
    ws = wb["3. Company Profile"]
    out: list[dict] = []
    for r in range(2, ws.max_row + 1):
        name = _clean(ws.cell(row=r, column=2).value)
        if not name:
            continue
        out.append({
            "company_name": name,
            "sector": _clean(ws.cell(row=r, column=3).value),
            "city": _clean(ws.cell(row=r, column=4).value),
            "company_phone": _clean(ws.cell(row=r, column=5).value),
            "company_email": _clean(ws.cell(row=r, column=6).value),
            "website": _clean(ws.cell(row=r, column=7).value),
            "year_established": _clean(ws.cell(row=r, column=8).value),
            "poc_name": _clean(ws.cell(row=r, column=9).value),
            "poc_title": _clean(ws.cell(row=r, column=10).value),
            "poc_email": _clean(ws.cell(row=r, column=11).value),
            "poc_phone": _clean(ws.cell(row=r, column=12).value),
            "source": "Companies Data Tracker — Company Profile",
        })
    wb.close()
    return out


def _read_companies_2024() -> list[dict]:
    wb = load_workbook(SRC_DB, read_only=False, data_only=True)
    ws = wb["Companies 2024"]
    out: list[dict] = []
    for r in range(2, ws.max_row + 1):
        name = _clean(ws.cell(row=r, column=2).value)
        if not name:
            continue
        city = _clean(ws.cell(row=r, column=5).value)
        if city == "#VALUE!":
            city = ""
        out.append({
            "company_name": name,
            "sector": _clean(ws.cell(row=r, column=4).value),
            "city": city,
            "company_phone": _clean(ws.cell(row=r, column=7).value),
            "company_email": _clean(ws.cell(row=r, column=8).value),
            "website": _clean(ws.cell(row=r, column=6).value),
            "poc_name": _clean(ws.cell(row=r, column=9).value),
            "poc_title": _clean(ws.cell(row=r, column=10).value),
            "poc_phone": _clean(ws.cell(row=r, column=11).value),
            "poc_email": _clean(ws.cell(row=r, column=12).value),
            "sb_pb": _clean(ws.cell(row=r, column=1).value),
            "applied": _clean(ws.cell(row=r, column=3).value),
            "source": "Companies Database 2025 — Companies 2024",
        })
    wb.close()
    return out


def _read_simple_database(tab: str, source_label: str) -> list[dict]:
    """Read Companies 2025 / Not active-available style tabs (header on row 2)."""
    wb = load_workbook(SRC_DB, read_only=False, data_only=True)
    ws = wb[tab]
    out: list[dict] = []
    # header row is row 2: Name, Specilaization, Email, Number, Location, Year, URL/Website, Notes
    for r in range(3, ws.max_row + 1):
        name = _clean(ws.cell(row=r, column=1).value)
        if not name or name.lower() in ("name", "name "):
            continue
        out.append({
            "company_name": name,
            "sector": _clean(ws.cell(row=r, column=2).value),
            "company_email": _clean(ws.cell(row=r, column=3).value),
            "company_phone": _clean(ws.cell(row=r, column=4).value),
            "city": _clean(ws.cell(row=r, column=5).value),
            "year_established": _clean(ws.cell(row=r, column=6).value),
            "website": _clean(ws.cell(row=r, column=7).value),
            "notes": _clean(ws.cell(row=r, column=8).value),
            "source": source_label,
        })
    wb.close()
    return out


def _merge(records: list[dict]) -> list[dict]:
    """Dedupe by normalized company_name. Later records fill gaps in earlier."""
    by_name: dict[str, dict] = {}
    for rec in records:
        key = _norm(rec["company_name"])
        if key in by_name:
            existing = by_name[key]
            for k, v in rec.items():
                if v and not existing.get(k):
                    existing[k] = v
            existing_sources = existing.get("source", "")
            if rec["source"] not in existing_sources:
                existing["source"] = f"{existing_sources}; {rec['source']}"
        else:
            by_name[key] = dict(rec)
    return sorted(by_name.values(), key=lambda r: r["company_name"].lower())


def _build_notes(rec: dict) -> str:
    bits = [f"Reference — {rec['source']}"]
    if rec.get("website"):
        bits.append(f"web: {rec['website']}")
    if rec.get("company_email"):
        bits.append(f"email: {rec['company_email']}")
    if rec.get("company_phone"):
        bits.append(f"phone: {rec['company_phone']}")
    if rec.get("year_established"):
        bits.append(f"established: {rec['year_established']}")
    if rec.get("sb_pb"):
        bits.append(f"track: {rec['sb_pb']}")
    if rec.get("applied"):
        bits.append(f"applied: {rec['applied']}")
    if rec.get("notes"):
        bits.append(rec["notes"])
    return " | ".join(bits)


def populate():
    companies = _merge([
        *_read_company_profile(),
        *_read_companies_2024(),
        *_read_simple_database("Companies 2025", "Companies Database 2025 — Companies 2025"),
        *_read_simple_database("Not active-available ", "Companies Database 2025 — Not active-available"),
    ])

    if not OUT.exists():
        raise FileNotFoundError(
            f"{OUT} missing. Run `python -m builders.companies_master` first."
        )

    wb = load_workbook(OUT)
    companies_ws = wb["Companies"]
    contacts_ws = wb["Contacts"]

    now = datetime.utcnow().strftime("%Y-%m-%d")
    style_e = editable_style()

    # Column indices in Companies (1-based):
    # 1 company_id (formula, skip), 2 company_name, 3 legal_name, 4 city,
    # 5 governorate, 6 sector, 7 employee_count, 8 revenue_bracket,
    # 9 international_revenue_pct, 10 readiness_score, 11 fund_code, 12 cohort,
    # 13 status, 14 stage, 15 profile_manager_email, 16 selection_date,
    # 17 onboarding_date, 18 primary_contact_id, 19 drive_folder_url,
    # 20 notes, 21 updated_at, 22 updated_by
    COL_NAME, COL_CITY, COL_SECTOR = 2, 4, 6
    COL_STAGE, COL_NOTES = 14, 20
    COL_UPDATED_AT, COL_UPDATED_BY = 21, 22

    # Extend the id formula + editable styling beyond row 500 if we have more.
    needed_rows = 1 + len(companies)
    current_max = 500
    if needed_rows > current_max:
        for row in range(current_max + 1, needed_rows + 1):
            companies_ws.cell(
                row=row, column=1,
                value=f'=IF(B{row}<>"", "E3-"&TEXT(ROW()-1, "0000"), "")',
            )
            for col in range(2, 21):
                cell = companies_ws.cell(row=row, column=col)
                for attr, value in style_e.items():
                    setattr(cell, attr, value)

    for idx, rec in enumerate(companies, start=2):
        companies_ws.cell(row=idx, column=COL_NAME, value=rec["company_name"])
        if rec.get("city"):
            companies_ws.cell(row=idx, column=COL_CITY, value=rec["city"])
        if rec.get("sector"):
            companies_ws.cell(row=idx, column=COL_SECTOR, value=rec["sector"])
        # Historical reference rows are not part of the Cohort 3 funnel — leave
        # stage blank so they are visually distinct from live applicants.
        companies_ws.cell(row=idx, column=COL_NOTES, value=_build_notes(rec))
        companies_ws.cell(row=idx, column=COL_UPDATED_AT, value=now)
        companies_ws.cell(row=idx, column=COL_UPDATED_BY, value="migrator")

    # Populate Contacts with POC info where we have it.
    # contact_id has a formula (col 1), so fill from col 2 onwards.
    # 2 company_id (we leave blank — portal can link via name search), 3 full_name,
    # 4 title, 5 email, 6 phone, 7 is_signatory, 8 notes, 9 updated_at, 10 updated_by
    contact_row = 2
    for idx, rec in enumerate(companies, start=2):
        poc_name = rec.get("poc_name")
        if not poc_name:
            continue
        company_id = f"E3-{idx - 1:04d}"  # mirrors the formula result
        contacts_ws.cell(row=contact_row, column=2, value=company_id)
        contacts_ws.cell(row=contact_row, column=3, value=poc_name)
        if rec.get("poc_title"):
            contacts_ws.cell(row=contact_row, column=4, value=rec["poc_title"])
        if rec.get("poc_email"):
            contacts_ws.cell(row=contact_row, column=5, value=rec["poc_email"])
        if rec.get("poc_phone"):
            contacts_ws.cell(row=contact_row, column=6, value=rec["poc_phone"])
        contacts_ws.cell(row=contact_row, column=8, value=f"Reference POC from {rec['source']}")
        contacts_ws.cell(row=contact_row, column=9, value=now)
        contacts_ws.cell(row=contact_row, column=10, value="migrator")
        contact_row += 1

    save_workbook(wb, OUT.name)
    print(f"wrote {len(companies)} company references and {contact_row - 2} contacts to {OUT}")


if __name__ == "__main__":
    populate()
