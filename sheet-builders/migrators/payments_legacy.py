"""Port legacy Aug-Dec 2025 payments into E3 - Payments Tracker.

Source: Payments Aug 1 to Dec 31.xlsx
    - CB tab:  pillar x month matrix (TTH West Bank, TTH Gaza, Upskilling,
               MKT Resources TTH) with per-company monthly disbursements and
               status flags.
    - M&B tab: Digify Marketing Agency first/final + Sponsorship.
    - MA tab:  flat list of Market Access payments.

Target: E3 - Payments Tracker.xlsx > Payments tab.

FK resolution:
    - company_id: best-effort normalized match against E3 Companies Master
      rows; unmatched stays blank.
    - pr_id / assignment_id: cannot be recovered from legacy matrix.
      Flagged via needs_review = TRUE.

Per decision: import with flag so nothing is lost.

Usage:
    python -m migrators.payments_legacy
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from gsg_sheets import save_workbook


ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "sheet-builders" / "out" / "E3 - Payments Tracker.xlsx"
COMPANIES_WB = ROOT / "sheet-builders" / "out" / "E3 - Companies Master.xlsx"
SRC = ROOT / "Payments Aug 1 to Dec 31.xlsx"

# Month name -> YYYY-MM-DD (end-of-month sentinel used by the legacy matrix).
MONTH_2025 = {
    "August": "2025-08-31",
    "September": "2025-09-30",
    "October": "2025-10-31",
    "November": "2025-11-30",
    "December": "2025-12-31",
}

PILLAR_TO_INTERVENTION = {
    "TTH  West Bank": "TTH",
    " TTH Gaza Strip": "TTH",
    "Upskilling": "Upskilling",
    "Upskilling ": "Upskilling",
    "MKT Resources TTH": "MKG",
}

FUND_DEFAULT = ""  # legacy sheet does not encode fund per row


def _norm(v) -> str:
    return "" if v is None else str(v).strip().lower()


def _clean(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def _read_companies_map() -> dict[str, str]:
    if not COMPANIES_WB.exists():
        print("[payments_legacy] warning: Companies Master workbook not found, "
              "company_id FKs will be left blank")
        return {}
    wb = load_workbook(COMPANIES_WB)
    ws = wb["Companies"]
    out: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        if not name:
            continue
        out[_norm(name)] = f"E3-{r - 1:04d}"
    wb.close()
    return out


def _map_company_id(name: str, companies: dict[str, str]) -> str:
    key = _norm(name)
    if not key:
        return ""
    if key in companies:
        return companies[key]
    # Light fuzzy: strip trailing punctuation / spaces, try common variants.
    variants = [key.rstrip(" .,"), key.replace(" ", ""), key.split()[0]]
    for v in variants:
        if v in companies:
            return companies[v]
    return ""


def _read_cb(wb) -> list[dict]:
    """Parse the CB tab into flat payment rows. Column layout:
        A #, B Company, C Total Amount, D-H August-December amounts,
        I-M August-December status flags.
    Pillar headers appear in column B of section header rows (e.g. row 2).
    """
    ws = wb["CB"]
    out: list[dict] = []
    current_pillar: Optional[str] = None
    months = ["August", "September", "October", "November", "December"]

    for r in range(1, ws.max_row + 1):
        a_val = ws.cell(row=r, column=1).value
        b_val = ws.cell(row=r, column=2).value

        # Pillar header: single text cell in either column A or column B with
        # the other empty and no numeric row-index. Skip known column-labels and
        # side-note rows.
        header_candidate: Optional[str] = None
        if isinstance(a_val, str) and a_val.strip() and (b_val is None or b_val == ""):
            header_candidate = a_val
        elif (a_val is None or a_val == "") and isinstance(b_val, str) and b_val.strip():
            header_candidate = b_val

        if header_candidate is not None:
            norm = header_candidate.strip().lower()
            if norm in ("company", "training", "total") or norm.startswith("side note") or norm.startswith("cb in"):
                # not a pillar header, ignore
                pass
            else:
                current_pillar = header_candidate
            continue

        if not isinstance(a_val, (int, float)):
            continue

        company = _clean(b_val)
        if not company or _norm(company) == "total":
            continue

        for m_idx, month in enumerate(months):
            amount = ws.cell(row=r, column=4 + m_idx).value
            if amount in (None, "") or not isinstance(amount, (int, float)):
                continue
            status_flag = ws.cell(row=r, column=9 + m_idx).value
            # Flag semantics from the legacy sheet:
            #   TRUE = paid / disbursed
            #   FALSE = in flight / submitted to finance
            #   blank = pending / not yet processed
            if status_flag is True:
                status = "Paid"
            elif status_flag is False:
                status = "Sent to Finance"
            else:
                status = "Pending Approval"

            intervention = PILLAR_TO_INTERVENTION.get(current_pillar or "", "TTH")

            out.append({
                "company_name": company,
                "payee_type": "Vendor",
                "payee_name": company,
                "intervention_type": intervention,
                "fund_code": FUND_DEFAULT,
                "amount_usd": float(amount),
                "currency": "USD",
                "payment_date": MONTH_2025[month],
                "status": status,
                "notes": f"Legacy CB {current_pillar or ''} - {month} 2025",
                "source_tab": "CB",
            })

    return out


def _read_mb(wb) -> list[dict]:
    """Parse M&B tab. Digify first/final + sponsorship."""
    ws = wb["M&B"]
    out: list[dict] = []
    payee = _clean(ws.cell(row=1, column=1).value)  # "Digify Marketing Agency"
    for r in range(2, ws.max_row + 1):
        label = _clean(ws.cell(row=r, column=1).value)
        amount = ws.cell(row=r, column=2).value
        month = ws.cell(row=r, column=3).value
        status_raw = _clean(ws.cell(row=r, column=4).value)
        if label.lower() in ("", "total", "mb in 2025"):
            continue
        if not isinstance(amount, (int, float)):
            continue
        payment_date = ""
        if isinstance(month, datetime):
            payment_date = month.strftime("%Y-%m-%d")
        if "submit" in status_raw.lower():
            status = "Sent to Finance"
        elif status_raw:
            status = "Paid"
        else:
            status = "Pending Approval"

        out.append({
            "company_name": "",
            "payee_type": "Vendor",
            "payee_name": f"{payee} - {label}" if label.lower() != "sponorship" else "Sponsorship - M&B",
            "intervention_type": "MKG",
            "fund_code": FUND_DEFAULT,
            "amount_usd": float(amount),
            "currency": "USD",
            "payment_date": payment_date,
            "status": status,
            "notes": f"Legacy M&B 2025 - {label}",
            "source_tab": "M&B",
        })
    return out


def _read_ma(wb) -> list[dict]:
    """Parse MA tab. Flat list of MA payments (Andersen, Conferences, advisors)."""
    ws = wb["MA"]
    out: list[dict] = []
    for r in range(5, ws.max_row + 1):
        idx = ws.cell(row=r, column=1).value
        if not isinstance(idx, (int, float)):
            continue
        label = _clean(ws.cell(row=r, column=2).value)
        amount = ws.cell(row=r, column=3).value
        month = _clean(ws.cell(row=r, column=5).value)
        if not label or not isinstance(amount, (int, float)):
            continue
        payment_date = MONTH_2025.get(month, "")

        lower = label.lower()
        if "conference" in lower:
            intervention = "Conferences"
            payee_type = "Conference"
        elif any(t in lower for t in ("payment", "advance")) and "andersen" in lower:
            intervention = "MA-Legal"
            payee_type = "Vendor"
        else:
            intervention = "MA"
            payee_type = "Advisor"

        # Extract company from "Conference (X)" style labels.
        company = ""
        if "(" in label and ")" in label:
            company = label.split("(", 1)[1].rsplit(")", 1)[0].strip()

        out.append({
            "company_name": company,
            "payee_type": payee_type,
            "payee_name": label,
            "intervention_type": intervention,
            "fund_code": FUND_DEFAULT,
            "amount_usd": float(amount),
            "currency": "USD",
            "payment_date": payment_date,
            "status": "Paid",
            "notes": f"Legacy MA 2025 - {label}",
            "source_tab": "MA",
        })
    return out


def _find_first_empty_row(ws, key_col: int = 6, start: int = 2) -> int:
    r = start
    while ws.cell(row=r, column=key_col).value:
        r += 1
    return r


def populate():
    if not OUT.exists():
        raise FileNotFoundError(
            f"{OUT} missing. Run `python -m builders.payments` first."
        )

    src = load_workbook(SRC, read_only=False, data_only=True)
    rows = _read_cb(src) + _read_mb(src) + _read_ma(src)
    src.close()
    print(f"[payments_legacy] parsed {len(rows)} legacy payment rows")

    companies = _read_companies_map()
    resolved = 0
    for rec in rows:
        rec["company_id"] = _map_company_id(rec["company_name"], companies)
        if rec["company_id"]:
            resolved += 1
    print(f"[payments_legacy] resolved {resolved}/{len(rows)} company_id FKs")

    tgt = load_workbook(OUT)
    ws = tgt["Payments"]

    # Target column indexes (1-based) per updated Payments builder schema:
    # 1 payment_id (formula), 2 pr_id, 3 company_id, 4 assignment_id,
    # 5 payee_type, 6 payee_name, 7 intervention_type, 8 fund_code,
    # 9 amount_usd, 10 currency, 11 payment_date, 12 status,
    # 13 finance_contact, 14 invoice_url, 15 receipt_url,
    # 16 needs_review, 17 notes, 18 updated_at, 19 updated_by
    now = datetime.utcnow().strftime("%Y-%m-%d")
    row = _find_first_empty_row(ws, key_col=6)
    for rec in rows:
        needs_review = "FALSE"
        review_reasons: list[str] = []
        if not rec["company_id"] and rec["payee_type"] != "Vendor":
            needs_review = "TRUE"
            review_reasons.append("company_id unresolved")
        if not rec["fund_code"]:
            needs_review = "TRUE"
            review_reasons.append("fund_code missing")
        # pr_id / assignment_id never recoverable from legacy
        review_reasons.append("pr_id/assignment_id not in legacy source")

        ws.cell(row=row, column=3).value = rec["company_id"]
        ws.cell(row=row, column=5).value = rec["payee_type"]
        ws.cell(row=row, column=6).value = rec["payee_name"]
        ws.cell(row=row, column=7).value = rec["intervention_type"]
        ws.cell(row=row, column=8).value = rec["fund_code"]
        ws.cell(row=row, column=9).value = rec["amount_usd"]
        ws.cell(row=row, column=10).value = rec["currency"]
        ws.cell(row=row, column=11).value = rec["payment_date"]
        ws.cell(row=row, column=12).value = rec["status"]
        ws.cell(row=row, column=13).value = "Khamis Eweis"
        ws.cell(row=row, column=16).value = needs_review
        notes = rec["notes"]
        if review_reasons:
            notes = f"{notes} | REVIEW: {'; '.join(review_reasons)}"
        ws.cell(row=row, column=17).value = notes
        ws.cell(row=row, column=18).value = now
        ws.cell(row=row, column=19).value = "migrator"
        row += 1

    save_workbook(tgt, OUT.name)
    print(f"[payments_legacy] wrote {len(rows)} rows to Payments tab")


if __name__ == "__main__":
    populate()
