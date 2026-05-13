"""Migrate ElevateBridge Freelancers Application Responses > Deduplicated_Final_Report
into the new E3 Freelancers schema.

Conservative: extracts identity and track fields only. Total income is preserved
but is not decomposed into platform-level records (the legacy columns are
bilingual and overlap with normalized variants; manual review recommended).
Writes a fresh output xlsx so no legacy data is mutated.

All migrated rows land as `status='Available'` because the legacy workbook
IS the active ElevateBridge freelancer pool — pre-vetted hunters who are
ready to be matched with Cohort 3 companies as their sales funnel. The
team's job in the portal is to pair these freelancers with the right
company, not to triage applications.
"""

from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from builders.freelancers import FILENAME, build as build_template
from gsg_sheets.styles import apply_style, cell_style, editable_style


LEGACY_RELPATH = "../ElevateBridge Freelancers Application Responses.xlsx"
LEGACY_TAB = "Deduplicated_Final_Report"


def _iter_legacy_rows(legacy_path: Path) -> Iterable[dict]:
    wb = load_workbook(legacy_path, read_only=True, data_only=True)
    ws = wb[LEGACY_TAB]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    # Fixed positions based on analysis; all other columns are heterogeneous.
    POS = {
        "full_name": 0,
        "email": 1,
        "phone": 2,
        "sub_track": 3,
        "total_income_usd": 10,
    }
    for row in rows:
        if not row:
            continue
        name = row[POS["full_name"]] if POS["full_name"] < len(row) else None
        if not name:
            continue
        yield {
            "full_name": str(name).strip(),
            "email": (row[POS["email"]] if POS["email"] < len(row) else "") or "",
            "phone": (row[POS["phone"]] if POS["phone"] < len(row) else "") or "",
            "sub_track": (row[POS["sub_track"]] if POS["sub_track"] < len(row) else "") or "",
            "total_income_usd": (row[POS["total_income_usd"]] if POS["total_income_usd"] < len(row) else "") or "",
        }


def _infer_track(sub_track: str) -> str:
    s = (sub_track or "").lower()
    if "upwork" in s or "fl" in s or "freelanc" in s:
        return "Upwork"
    if "sm" in s or "social" in s or "linkedin" in s or "instagram" in s:
        return "Social Media"
    return "Other"


def _infer_role(sub_track: str) -> str:
    s = (sub_track or "").lower()
    if "agency" in s or "agen" in s:
        return "Agency"
    if "job hunter" in s or "jh" in s or "hunter" in s:
        return "Job Hunter"
    return "Individual"


def run():
    # Regenerate template first so the output reflects current schema.
    build_template()
    out_path = Path(__file__).resolve().parents[1] / "out" / FILENAME
    legacy_path = Path(__file__).resolve().parents[1] / LEGACY_RELPATH

    wb = load_workbook(out_path)
    freelancers = wb["Freelancers"]
    style_e = editable_style()
    style_c = cell_style()
    today = datetime.utcnow().strftime("%Y-%m-%d")

    # Column indexes (1-based) — keep in sync with FREELANCERS_HEADERS in
    # builders/freelancers.py.
    COL_NAME = 2
    COL_EMAIL = 3
    COL_PHONE = 4
    COL_TRACK = 6
    COL_ROLE = 7
    COL_STATUS = 10
    COL_SOURCE = 12
    COL_NOTES = 13
    COL_UPDATED_AT = 19
    COL_UPDATED_BY = 20

    row = 2
    ingested = 0
    for record in _iter_legacy_rows(legacy_path.resolve()):
        # Skip column 1 (formula auto-populates the id).
        freelancers.cell(row=row, column=COL_NAME, value=record["full_name"])
        freelancers.cell(row=row, column=COL_EMAIL, value=record["email"])
        freelancers.cell(row=row, column=COL_PHONE, value=record["phone"])
        freelancers.cell(row=row, column=COL_TRACK, value=_infer_track(record["sub_track"]))
        freelancers.cell(row=row, column=COL_ROLE, value=_infer_role(record["sub_track"]))
        # Pre-vetted pool — every row here is ready to be matched with a
        # Cohort 3 company. New form submissions land here too as Available.
        freelancers.cell(row=row, column=COL_STATUS, value="Available")
        freelancers.cell(row=row, column=COL_SOURCE, value=f"{LEGACY_TAB} row {row}")
        notes = (
            f"Total income (legacy, unverified): {record['total_income_usd']}"
            if record["total_income_usd"]
            else ""
        )
        if notes:
            freelancers.cell(row=row, column=COL_NOTES, value=notes)
        freelancers.cell(row=row, column=COL_UPDATED_AT, value=today)
        freelancers.cell(row=row, column=COL_UPDATED_BY, value="migrator")
        # Apply editable styling to the form-response columns.
        for col in [2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]:
            apply_style(freelancers.cell(row=row, column=col), style_e)
        # Audit columns get the readonly cell style.
        for col in [COL_UPDATED_AT, COL_UPDATED_BY]:
            apply_style(freelancers.cell(row=row, column=col), style_c)
        row += 1
        ingested += 1

    wb.save(out_path)
    print(f"Migrated {ingested} freelancers into {out_path} (all marked Available — ready to match)")
    return ingested


if __name__ == "__main__":
    run()
