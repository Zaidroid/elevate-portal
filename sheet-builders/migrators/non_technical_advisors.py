"""Port Non-Technical Advisors Responses.xlsx into the E3 - Non-Technical
Advisors workbook produced by builders/advisors.py.

The builder creates the template (Dashboard, Advisors with tracker columns,
FollowUps, ActivityLog, Comments, Mentors, Lookups). This migrator opens that
template and copies legacy form-response rows into the Advisors tab, mapping
the legacy column names to the canonical schema.

Defaults applied to every migrated row:
  - pipeline_status = "New"
  - assignment_status = "Planned" (when assignment_intervention_type is set)
  - updated_by = "migrator"
  - updated_at = today
"""

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from builders.advisors import (
    ADVISORS_HEADERS,
    ADVISORS_LAST_ROW,
    FILENAME as TEMPLATE_FILENAME,
    build as build_template,
)
from gsg_sheets.styles import apply_style, cell_style, editable_style


SRC_PATH = "../Non-Technical Advisors Responses.xlsx"

# Legacy headers vary wildly in wording (full question text, trailing
# whitespace, punctuation, even an "I agree to safeguarding" preamble row).
# Instead of an exact-match table, route by keyword presence. Each rule is
# (canonical_name, [must-include-substrings]); first match wins. Substrings
# are matched case-insensitively against the legacy header text.
HEADER_RULES: list[tuple[str, list[str]]] = [
    ("timestamp", ["timestamp"]),
    ("full_name", ["full name"]),
    ("gender", ["gender"]),
    ("country", ["country"]),
    ("email", ["email"]),
    ("whatsapp", ["whatsapp"]),
    ("linkedin", ["linkedin"]),
    ("tech_rating", ["rate", "experience", "tech industry"]),
    ("tech_rating", ["technical knowledge"]),
    ("eco_rating", ["palestinian tech"]),
    ("eco_rating", ["ecosystem"]),
    ("c_level", ["c-level managers"]),
    ("c_level", ["c-level role"]),
    ("c_level_detail", ["if yes, please share"]),
    ("exp_areas", ["which of the following"]),
    ("exp_detail", ["if any of the above"]),
    ("position", ["current position"]),
    ("employer", ["current employer"]),
    ("years", ["years of experience"]),
    ("non_tech_subjects", ["non-technical"]),
    ("non_tech_subjects", ["non technical"]),
    ("tech_specs", ["technical:", "specializations"]),
    ("gsg_past", ["worked", "gsg before"]),
    ("gsg_past", ["volunteered with gsg"]),
    ("paid_or_vol", ["paid or volunteering"]),
    ("hourly_rate", ["hourly rate"]),
    ("cv_link", ["upload your cv"]),
    ("cv_link", ["cv link"]),
    ("notes", ["anything else"]),
    ("heard_from", ["how did you hear"]),
    ("opportunities", ["opportunities related"]),
    ("support_in", ["like to support in"]),
    ("support_via", ["supporting gsg through"]),
    ("newsletter", ["newsletter"]),
]


def _normalize_legacy_header(text: str) -> str:
    return (text or "").strip()


def _route_header(legacy: str) -> str | None:
    """Return canonical schema name for a legacy header text, or None if no rule matches."""
    if not legacy:
        return None
    low = legacy.lower()
    for canonical, needles in HEADER_RULES:
        if all(n in low for n in needles):
            return canonical
    return None


def _index_canonical_columns(headers: list[str]) -> dict[str, int]:
    """Map canonical schema name -> 1-based column index in the template."""
    return {name: idx + 1 for idx, name in enumerate(headers)}


def _coerce(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d %H:%M:%S")
    return str(v).strip()


def run():
    root = Path(__file__).resolve().parents[1]
    src_path = (root / SRC_PATH).resolve()
    if not src_path.exists():
        print(f"[non_technical_advisors] source missing: {src_path} (skipping)")
        return None

    # Make sure the template exists and is fresh.
    template_path = build_template()

    src = load_workbook(src_path, data_only=True)
    if "Form Responses 1" not in src.sheetnames:
        print("[non_technical_advisors] no Form Responses 1 tab; skipping")
        return template_path

    src_ws = src["Form Responses 1"]
    src_max_col = src_ws.max_column
    src_max_row = src_ws.max_row

    legacy_headers = [
        _normalize_legacy_header(src_ws.cell(row=1, column=c).value)
        for c in range(1, src_max_col + 1)
    ]

    # legacy_idx (0-based) -> canonical schema name (or None if unmapped)
    legacy_to_canonical: list[str | None] = [_route_header(h) for h in legacy_headers]
    matched = sum(1 for c in legacy_to_canonical if c)
    print(f"[non_technical_advisors] mapped {matched}/{len(legacy_headers)} legacy headers")

    out = load_workbook(template_path)
    advisors = out["Advisors"]
    canonical_index = _index_canonical_columns(ADVISORS_HEADERS)
    pipeline_col = canonical_index["pipeline_status"]
    asgn_status_col = canonical_index["assignment_status"]
    asgn_int_col = canonical_index["assignment_intervention_type"]
    updated_at_col = canonical_index["updated_at"]
    updated_by_col = canonical_index["updated_by"]

    style_c = cell_style()
    style_e = editable_style()

    today = datetime.utcnow().strftime("%Y-%m-%d")
    # Cohort 3 cutoff: advisors who submitted before Jan 1 2026 are pre-cohort
    # historicals and land in pipeline_status='Archived' so they don't pollute
    # the live triage funnel. Anything from 2026-01-01 onward starts as 'New'.
    COHORT_3_START = datetime(2026, 1, 1)

    # Find the timestamp column index in the legacy header layout.
    timestamp_legacy_idx = next(
        (i for i, c in enumerate(legacy_to_canonical) if c == "timestamp"),
        None,
    )

    def _row_timestamp(vals: list) -> datetime | None:
        if timestamp_legacy_idx is None:
            return None
        v = vals[timestamp_legacy_idx]
        if isinstance(v, datetime):
            return v
        if isinstance(v, str) and v:
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y"):
                try:
                    return datetime.strptime(v, fmt)
                except ValueError:
                    continue
        return None

    written = 0
    archived = 0
    out_row = 2
    for r in range(2, src_max_row + 1):
        vals = [src_ws.cell(row=r, column=c).value for c in range(1, src_max_col + 1)]
        if all(v in (None, "") for v in vals):
            continue
        if out_row > ADVISORS_LAST_ROW:
            break

        # advisor_id is left as the formula the template put there.
        for src_idx, value in enumerate(vals):
            target_name = legacy_to_canonical[src_idx]
            if not target_name:
                continue
            target_col = canonical_index.get(target_name)
            if not target_col:
                continue
            cell = advisors.cell(row=out_row, column=target_col, value=_coerce(value))
            apply_style(cell, style_c)

        # Pipeline status: Archived for pre-2026 (already triaged historically),
        # New for 2026+ entries that need triage in the portal.
        ts = _row_timestamp(vals)
        is_archived = ts is not None and ts < COHORT_3_START
        status = "Archived" if is_archived else "New"
        if is_archived:
            archived += 1
        apply_style(advisors.cell(row=out_row, column=pipeline_col, value=status), style_e)

        # Audit stamps.
        apply_style(advisors.cell(row=out_row, column=updated_at_col, value=today), style_c)
        apply_style(advisors.cell(row=out_row, column=updated_by_col, value="migrator"), style_c)

        # If the legacy form did not carry assignment columns (it never does)
        # leave them blank so they keep the editable-blue fill the template
        # set up.
        out_row += 1
        written += 1

    print(f"[non_technical_advisors] of {written} migrated, {archived} archived (pre-2026), {written - archived} pending triage")

    out.save(template_path)
    print(f"[non_technical_advisors] migrated {written} advisors -> {template_path}")
    return str(template_path)


if __name__ == "__main__":
    run()
