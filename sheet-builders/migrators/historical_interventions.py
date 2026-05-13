"""Port Elevate 1 and Elevate 2 company x intervention rows into the
Historical Interventions tab of E3 - Companies Master.

Source: Companies Data Tracker (1).xlsx > 2. Participating Companies
        and                             > 1. Intervention Data (offering catalogue
        used to enrich offering_name / specialization via cohort_name join).

The target tab is intentionally separate from Intervention Assignments so
Cohort 3 roster queries stay clean. company_id is populated where a legacy
company_name matches a row already in the Companies tab (normalized match).

Usage:
    python -m migrators.historical_interventions
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from gsg_sheets import save_workbook
from gsg_sheets.styles import editable_style


ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "sheet-builders" / "out" / "E3 - Companies Master.xlsx"
SRC = ROOT / "Companies Data Tracker (1).xlsx"

DONOR_TO_FUND = {"Dutch": "97060", "SIDA": "91763"}


def _norm(v) -> str:
    if v is None:
        return ""
    return str(v).strip().lower()


def _clean(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].replace("-", "").isdigit():
        s = s[:-2]
    return s


def _read_offering_catalogue() -> dict[str, dict]:
    """cohort_name -> {offering_name, specialization}."""
    wb = load_workbook(SRC, read_only=False, data_only=True)
    ws = wb["1. Intervention Data"]
    out: dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        cohort_name = _clean(ws.cell(row=r, column=4).value)
        if not cohort_name:
            continue
        out[cohort_name.lower()] = {
            "offering_name": _clean(ws.cell(row=r, column=2).value),
            "specialization": _clean(ws.cell(row=r, column=3).value),
        }
    wb.close()
    return out


def _read_participating() -> list[dict]:
    wb = load_workbook(SRC, read_only=False, data_only=True)
    ws = wb["2. Participating Companies"]
    catalogue = _read_offering_catalogue()
    out: list[dict] = []
    for r in range(2, ws.max_row + 1):
        company = _clean(ws.cell(row=r, column=3).value)
        if not company:
            continue
        cohort_name = _clean(ws.cell(row=r, column=2).value)
        offering = catalogue.get(cohort_name.lower(), {})
        cohort = _clean(ws.cell(row=r, column=1).value)
        cohort_code = "E1" if "elevate 1" in cohort.lower() else "E2" if "elevate 2" in cohort.lower() else cohort
        donor = _clean(ws.cell(row=r, column=4).value)
        out.append({
            "cohort": cohort_code,
            "company_name": company,
            "offering_name": offering.get("offering_name", ""),
            "specialization": offering.get("specialization", ""),
            "cohort_name": cohort_name,
            "donor": donor,
            "fund_code": DONOR_TO_FUND.get(donor, ""),
            "start_date": _clean(ws.cell(row=r, column=5).value),
            "end_date": _clean(ws.cell(row=r, column=6).value),
            "year": _clean(ws.cell(row=r, column=7).value),
            "agreement_link": _clean(ws.cell(row=r, column=8).value),
            "source": "Companies Data Tracker — 2. Participating Companies",
        })
    wb.close()
    return out


def _read_existing_companies(wb) -> dict[str, str]:
    """Map normalized company_name -> company_id from the Companies tab."""
    ws = wb["Companies"]
    out: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        if not name:
            continue
        # company_id is a formula result; stored value is the formula string until
        # opened in a spreadsheet engine, so reconstruct the expected id.
        company_id = f"E3-{r - 1:04d}"
        out[_norm(name)] = company_id
    return out


def populate():
    if not OUT.exists():
        raise FileNotFoundError(
            f"{OUT} missing. Run `python -m builders.companies_master` first."
        )

    rows = _read_participating()
    print(f"[historical] read {len(rows)} rows from 2. Participating Companies")

    wb = load_workbook(OUT)
    if "Historical Interventions" not in wb.sheetnames:
        raise RuntimeError(
            "Historical Interventions tab not found. Rebuild Companies Master "
            "from the updated builder first."
        )

    company_ids = _read_existing_companies(wb)
    ws = wb["Historical Interventions"]

    now = datetime.utcnow().strftime("%Y-%m-%d")
    matched = 0
    for idx, rec in enumerate(rows, start=2):
        # Column 1 has the historical_id formula — do not overwrite.
        ws.cell(row=idx, column=2, value=rec["cohort"])
        ws.cell(row=idx, column=3, value=rec["company_name"])
        cid = company_ids.get(_norm(rec["company_name"]), "")
        if cid:
            matched += 1
        ws.cell(row=idx, column=4, value=cid)
        ws.cell(row=idx, column=5, value=rec["offering_name"])
        ws.cell(row=idx, column=6, value=rec["specialization"])
        ws.cell(row=idx, column=7, value=rec["cohort_name"])
        ws.cell(row=idx, column=8, value=rec["donor"])
        ws.cell(row=idx, column=9, value=rec["fund_code"])
        ws.cell(row=idx, column=10, value=rec["start_date"])
        ws.cell(row=idx, column=11, value=rec["end_date"])
        ws.cell(row=idx, column=12, value=rec["year"])
        ws.cell(row=idx, column=13, value=rec["agreement_link"])
        ws.cell(row=idx, column=14, value=rec["source"])
        ws.cell(row=idx, column=16, value=now)
        ws.cell(row=idx, column=17, value="migrator")

    save_workbook(wb, OUT.name)
    print(
        f"[historical] wrote {len(rows)} rows, matched {matched} company_ids "
        f"from Companies roster"
    )


if __name__ == "__main__":
    populate()
