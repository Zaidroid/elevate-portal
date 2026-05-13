"""Orphan-tab audit for the legacy ElevateBridge workbook.

The legacy `ElevateBridge Freelancers Application Responses.xlsx` ships with
35 tabs; only `Deduplicated_Final_Report` was migrated into the new schema by
migrators/elevatebridge.py. The other 34 are unaccounted for.

This script opens that workbook, classifies each non-canonical tab, and
writes a Markdown report to out/elevatebridge_orphans.md so the team can
decide archive / merge / keep on each.

Classification heuristics:
  - Headers contain "income" or platform names ("upwork", "fiverr") → income
  - Headers contain "assessment", "score", "evaluation" → assessment
  - Headers contain "mentor" → mentor
  - Tab name starts with a number ("01_", "02_") → ad-hoc / sequenced
  - Otherwise → applicant (default)

Usage:
    python -m tools.orphan_audit
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


CANONICAL = {"Deduplicated_Final_Report"}


ROOT = Path(__file__).resolve().parents[2]
LEGACY_PATH = ROOT / "ElevateBridge Freelancers Application Responses.xlsx"
OUT_PATH = ROOT / "sheet-builders" / "out" / "elevatebridge_orphans.md"


def _classify(tab_name: str, headers: list[str]) -> tuple[str, str]:
    """Return (class_label, recommended_disposition)."""
    blob = " ".join((h or "").lower() for h in headers)
    name_lower = tab_name.lower()

    if any(k in blob for k in ["income", "earnings", "monthly revenue"]):
        return "income tracking", "merge into Income Tracking tab"
    if any(k in blob for k in ["assessment", "score", "evaluation", "rubric"]):
        return "assessment", "merge into Assessments tab"
    if "mentor" in blob or "mentor" in name_lower:
        return "mentor", "merge into Mentors / Track Assignments"
    if any(k in blob for k in ["upwork", "fiverr", "freelancer.com", "platform"]):
        return "platform record", "merge into Income Tracking tab"
    if any(name_lower.startswith(prefix) for prefix in ["sheet", "copy", "draft", "wip"]):
        return "scratch", "archive"
    if any(c.isdigit() for c in name_lower[:3]):
        return "ad-hoc / sequenced", "review then archive"
    return "applicant", "verify against Deduplicated_Final_Report; archive if duplicate"


def run() -> str:
    if not LEGACY_PATH.exists():
        print(f"[orphan_audit] legacy file missing: {LEGACY_PATH}")
        return ""

    # read_only mode reports 0/0 dimensions until cells are scanned, so use
    # the regular mode at the cost of a slower load.
    wb = load_workbook(LEGACY_PATH, data_only=True)

    rows = []
    for tab_name in wb.sheetnames:
        ws = wb[tab_name]
        max_col = min(ws.max_column or 0, 30)
        headers = [
            ws.cell(row=1, column=c).value or "" for c in range(1, max_col + 1)
        ]
        # Strip trailing empties
        while headers and not str(headers[-1]).strip():
            headers.pop()
        row_count = max(0, (ws.max_row or 1) - 1)
        cls, disp = _classify(tab_name, [str(h) for h in headers])
        is_canonical = tab_name in CANONICAL
        rows.append({
            "tab": tab_name,
            "rows": row_count,
            "cols": len(headers),
            "headers": [str(h) for h in headers[:10]],
            "class": "canonical" if is_canonical else cls,
            "disposition": "keep — already migrated" if is_canonical else disp,
        })

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with OUT_PATH.open("w", encoding="utf-8") as f:
        f.write("# ElevateBridge legacy workbook — orphan tab audit\n\n")
        f.write(f"Source: `{LEGACY_PATH.name}`\n")
        f.write(f"Tabs found: {len(rows)} (1 canonical, {len(rows) - 1} orphans)\n\n")

        f.write("## Summary\n\n")
        from collections import Counter
        counts = Counter(r["class"] for r in rows)
        for cls, n in counts.most_common():
            f.write(f"- **{cls}**: {n}\n")
        f.write("\n")

        f.write("## Per-tab classification\n\n")
        f.write("| # | Tab | Class | Rows | Cols | Recommended disposition |\n")
        f.write("|---|-----|-------|------|------|-------------------------|\n")
        for i, r in enumerate(rows, 1):
            f.write(f"| {i} | `{r['tab']}` | {r['class']} | {r['rows']} | {r['cols']} | {r['disposition']} |\n")
        f.write("\n")

        f.write("## Header samples\n\n")
        for r in rows:
            if r["class"] == "canonical":
                continue
            f.write(f"### `{r['tab']}`\n")
            if r["headers"]:
                f.write("Top headers: " + ", ".join(f"`{h}`" for h in r["headers"]) + "\n\n")
            else:
                f.write("(no usable headers)\n\n")

    wb.close()
    print(f"[orphan_audit] wrote {OUT_PATH}")
    return str(OUT_PATH)


if __name__ == "__main__":
    run()
